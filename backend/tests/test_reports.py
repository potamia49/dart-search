"""app/reports/audit_proposal.py + `POST /api/jobs/{id}/generate-report` 테스트.

보고서 생성은 **로컬 파일 시스템에 폴더/파일을 떨구는** 기능이라, 테스트는
`get_settings()`를 스텁으로 갈아끼워 산출물 경로를 `tmp_path`로 돌린다
(개발자 저장소의 실제 `backend/report/`를 더럽히지 않기 위함).

템플릿은 저장소 루트의 실제 `tamplate/audit_proposal_template.html`을 그대로
쓴다 — 치환 대상(`const EMBEDDED_DATA = {...};`)이 실제 템플릿에서도 정확히
잡히는지가 이 기능의 핵심 계약이기 때문이다.
"""

from __future__ import annotations

import json
import re
import shutil
import subprocess
from datetime import date
from pathlib import Path
from types import SimpleNamespace

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app import main as app_main
from app.api.results import _load_job_peer_pool
from app.core.db import get_db
from app.exporters.excel import FINANCIAL_SNAPSHOT_ACCOUNT_LABELS
from app.models import Base
from app.models.financial_snapshot import FinancialSnapshot
from app.models.job import Job, JobStatus
from app.models.result import ParseStatus, Result
from app.reports import audit_proposal, firm_profile
from app.reports.audit_proposal import (
    LABEL_FILENAME,
    MAX_COMPARISON_ABS_RATIO,
    MAX_PEERS,
    MAX_REGION_GROUP,
    MAX_WARNING_NAME_SAMPLES,
    MIN_COMPARISON_SAMPLE,
    UNKNOWN_OPINION_LABEL,
    UNRELIABLE_OPINIONS,
    RATIO_REQUIRED_KEYS,
    SNAPSHOT_FIELD_TO_REPORT_KEY,
    PeerPool,
    ReportGenerationError,
    ReportInput,
    allocate_output_dir,
    build_company_payload,
    build_comparison_year_note,
    build_financial_rows,
    build_industry_average,
    build_peer_candidate,
    build_peer_pool,
    build_peer_rows,
    build_report_payload,
    collect_comparison_warnings,
    collect_warnings,
    compute_report_ratios,
    match_by_industry_prefix,
    select_peers,
    select_region_candidates,
    select_region_group,
    excel_safe_text,
    find_embedded_data_block,
    generate_reports,
    missing_ratio_inputs,
    render_report_html,
    resolve_template_path,
    sanitize_filename_stem,
    select_financial_rows,
    unique_filename,
    write_label_workbook,
)

REPO_ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_PATH = REPO_ROOT / "tamplate" / "audit_proposal_template.html"


@pytest.fixture
def template_text() -> str:
    assert TEMPLATE_PATH.is_file(), f"보고서 템플릿이 없습니다: {TEMPLATE_PATH}"
    return TEMPLATE_PATH.read_text(encoding="utf-8")


@pytest.fixture
def report_settings(tmp_path, monkeypatch):
    """산출물 폴더를 tmp_path로, 템플릿을 저장소 실제 파일로 고정한다."""
    stub = SimpleNamespace(
        report_output_dir=str(tmp_path / "report"),
        report_template_path=str(TEMPLATE_PATH),
    )
    monkeypatch.setattr(audit_proposal, "get_settings", lambda: stub)
    return stub


def _extract_embedded_json(html: str) -> dict:
    """생성된 HTML에서 EMBEDDED_DATA 객체를 JSON으로 다시 읽어온다."""
    start, end = find_embedded_data_block(html)
    block = html[start:end]
    body = block.split("=", 1)[1].strip().rstrip(";")
    return json.loads(body)


# ---------------------------------------------------------------------------
# 매핑 드리프트 가드
# ---------------------------------------------------------------------------


def test_snapshot_field_mapping_matches_real_snapshot_columns():
    """보고서가 쓰는 재무 필드는 전부 `financial_snapshots`의 실제 계정 컬럼이어야 한다.

    엑셀 내보내기의 `FINANCIAL_SNAPSHOT_ACCOUNT_LABELS`(같은 테이블의 계정 컬럼
    정의)를 기준으로 대조한다 — 컬럼명을 오타로 적으면 조용히 전부 null인 보고서가
    나오므로 여기서 먼저 깨지게 한다.
    """
    unknown = set(SNAPSHOT_FIELD_TO_REPORT_KEY) - set(FINANCIAL_SNAPSHOT_ACCOUNT_LABELS)
    assert not unknown, f"스냅샷에 없는 컬럼을 매핑하고 있습니다: {sorted(unknown)}"
    # 템플릿이 요구하는 13항목(재무상태표 7 + 손익계산서 6)이 모두 있어야 한다.
    assert len(SNAPSHOT_FIELD_TO_REPORT_KEY) == 13


def test_template_is_resolvable_from_repo(report_settings):
    assert resolve_template_path() == TEMPLATE_PATH


# ---------------------------------------------------------------------------
# 템플릿 치환
# ---------------------------------------------------------------------------


def test_render_replaces_embedded_data_only(template_text):
    payload = {
        "firm": {"name": "금바다세무회계"},
        "company": {"name": "㈜테스트"},
        "financials": [],
        "peers": [],
        "industryAverage": None,
        "regionGroup": [],
        "opinionSummary": "",
    }
    html = render_report_html(template_text, payload)

    # EMBEDDED_DATA 블록은 딱 하나 남고, 그 안의 값이 교체돼야 한다.
    assert html.count("const EMBEDDED_DATA") == 1
    assert _extract_embedded_json(html) == payload
    # 템플릿 원문(HTML/CSS/렌더 로직)은 건드리지 않는다.
    assert "async function loadData()" in html
    assert "</html>" in html
    # 자리표시자 주석("예: ...")이 남은 옛 블록이 그대로 있지 않아야 한다.
    assert '"firm": {\n    "name": "",' not in html


def test_render_escapes_script_terminator_in_company_name(template_text):
    """회사명에 `</script>`가 들어가도 HTML이 조기 종료되면 안 된다."""
    payload = {
        "firm": {"name": "F"},
        "company": {"name": '㈜악의</script><script>alert("x")</script>'},
        "financials": [],
        "peers": [],
        "industryAverage": None,
        "regionGroup": [],
        "opinionSummary": "",
    }
    html = render_report_html(template_text, payload)

    start, end = find_embedded_data_block(html)
    block = html[start:end]
    assert "</script" not in block
    assert "\\u003c/script" in block
    # 이스케이프해도 JSON 값 자체는 원본 그대로 복원돼야 한다.
    assert _extract_embedded_json(html)["company"]["name"] == payload["company"]["name"]


def test_find_embedded_data_block_ignores_braces_in_comments_and_strings():
    text = (
        "prefix\n"
        "const EMBEDDED_DATA = {\n"
        '  // 주석 속 중괄호 { } 는 무시한다\n'
        '  "note": "문자열 속 } 도 무시",\n'
        "  /* 블록 주석 } */\n"
        '  "nested": { "a": 1 }\n'
        "};\n"
        "suffix"
    )
    start, end = find_embedded_data_block(text)
    assert text[:start] == "prefix\n"
    assert text[end:] == "\nsuffix"


def test_find_embedded_data_block_raises_when_marker_missing():
    with pytest.raises(ReportGenerationError):
        find_embedded_data_block("<html><script>var x = 1;</script></html>")


# ---------------------------------------------------------------------------
# 데이터 조립
# ---------------------------------------------------------------------------


def test_build_financial_rows_maps_korean_keys_in_year_order():
    snapshots = [
        SimpleNamespace(
            fiscal_year="2024",
            current_assets=11,
            noncurrent_assets=12,
            total_assets=23,
            current_liab=5,
            noncurrent_liab=3,
            total_liab=8,
            total_equity=15,
            revenue=100,
            cogs=60,
            gross_profit=40,
            sga=25,
            operating_income=15,
            net_income=10,
        ),
        SimpleNamespace(
            fiscal_year="2023",
            current_assets=1,
            noncurrent_assets=2,
            total_assets=3,
            current_liab=1,
            noncurrent_liab=1,
            total_liab=2,
            total_equity=1,
            revenue=50,
            cogs=None,
            gross_profit=20,
            sga=10,
            operating_income=10,
            net_income=5,
        ),
    ]
    rows = build_financial_rows(snapshots)

    assert [r["year"] for r in rows] == [2023, 2024]  # 연도 오름차순 + 숫자 변환
    assert rows[1]["매출원가"] == 60
    assert rows[1]["판관비"] == 25
    assert rows[0]["매출원가"] is None  # 결측은 null(템플릿이 "-"로 처리)
    assert set(rows[0]) == {"year", *SNAPSHOT_FIELD_TO_REPORT_KEY.values()}


def test_build_report_payload_leaves_unavailable_sections_empty():
    result = SimpleNamespace(
        id=1,
        corp_name="㈜테스트",
        induty_name="금속가공제품 제조업",
        address="경남 김해시 삼계로 1",
        ceo_name="홍길동",
        fiscal_date="20251231",
        audit_opinion="적정",
        auditor_name="안경회계법인",
    )
    payload = build_report_payload(ReportInput(result=result, snapshots=[]))

    assert payload["company"] == {
        "name": "㈜테스트",
        "industry": "금속가공제품 제조업",
        "address": "경남 김해시 삼계로 1",
        "ceo": "홍길동",
        "fiscalDate": "20251231",
        "opinion": "적정",
        "auditor": "안경회계법인",
    }
    assert payload["peers"] == []
    assert payload["industryAverage"] is None
    assert payload["regionGroup"] == []
    assert payload["opinionSummary"] == ""
    assert payload["firm"]["name"]  # 사무소 상수가 실려야 한다


# ---------------------------------------------------------------------------
# [H2/H3] 연도 선별 — 결측/0분모/FAILED는 제외, PARTIAL/전기유래는 싣되 경고
# ---------------------------------------------------------------------------


def test_missing_ratio_inputs_flags_nulls_and_zero_denominators():
    complete = {key: 10 for key in RATIO_REQUIRED_KEYS}
    complete["year"] = 2024
    assert missing_ratio_inputs(complete) == []

    # null 필수항목
    null_row = dict(complete, 매출총이익=None)
    assert missing_ratio_inputs(null_row) == ["매출총이익"]

    # 0 분모(Infinity 방지) — 분모가 아닌 항목의 0은 정상 값이라 걸리지 않는다.
    assert missing_ratio_inputs(dict(complete, 매출액=0)) == ["매출액"]
    assert missing_ratio_inputs(dict(complete, 자본총계=0)) == ["자본총계"]
    assert missing_ratio_inputs(dict(complete, 매출총이익=0)) == []
    # 결측/0이 아닌 항목(매출원가·판관비·비유동자산)은 애초에 필수가 아니다.
    assert "매출원가" not in RATIO_REQUIRED_KEYS
    assert "판관비" not in RATIO_REQUIRED_KEYS
    assert "비유동자산" not in RATIO_REQUIRED_KEYS


def test_select_financial_rows_excludes_failed_and_incomplete_years():
    snapshots = [
        _snapshot_stub("2021", parse_status=ParseStatus.FAILED),
        _snapshot_stub("2022", total_equity=None),  # 결측 → 제외
        _snapshot_stub("2023", parse_status=ParseStatus.PARTIAL),  # 실음 + 경고
        _snapshot_stub("2024", from_current_period=0),  # 실음 + 경고
    ]
    selection = select_financial_rows(snapshots)

    assert [row["year"] for row in selection.rows] == [2023, 2024]
    assert selection.total_years == 4
    assert selection.failed_years == ["2021"]
    assert selection.incomplete_years == [("2022", ["자본총계"])]
    assert selection.partial_years == ["2023"]
    assert selection.prior_period_years == ["2024"]
    # 실린 연도에는 비율 계산 필수항목이 하나도 비어 있지 않아야 한다.
    for row in selection.rows:
        assert missing_ratio_inputs(row) == []


def test_collect_warnings_reports_partial_prior_period_and_excluded_years():
    item = ReportInput(
        result=_result_stub(3, "㈜검수필요", "주소", parse_status=ParseStatus.PARTIAL),
        snapshots=[
            _snapshot_stub("2021", parse_status=ParseStatus.FAILED),
            _snapshot_stub("2022", revenue=None),
            _snapshot_stub("2023", parse_status=ParseStatus.PARTIAL),
            _snapshot_stub("2024", from_current_period=0),
        ],
    )
    messages = [w.message for w in collect_warnings(item, select_financial_rows(item.snapshots))]
    joined = " / ".join(messages)

    assert "파싱 실패(FAILED)한 회계연도를 보고서에서 제외" in joined and "2021" in joined
    assert "전체 4개년 중 1개년은 재무 항목이 결측" in joined and "2022년(매출액)" in joined
    assert "부분 파싱(PARTIAL)" in joined and "2023년" in joined
    assert "전기 항목에서 가져온 참고값" in joined and "2024년" in joined
    assert "이 회사의 파싱 상태가 PARTIAL입니다" in joined
    # 생성은 되므로 "생성하지 않았습니다"가 있으면 안 된다.
    assert "생성하지 않았습니다" not in joined


def test_collect_warnings_says_not_generated_when_all_years_unusable():
    item = ReportInput(
        result=_result_stub(4, "㈜전부실패", "주소"),
        snapshots=[
            _snapshot_stub("2023", parse_status=ParseStatus.FAILED),
            _snapshot_stub("2024", total_assets=None),
        ],
    )
    messages = [w.message for w in collect_warnings(item, select_financial_rows(item.snapshots))]

    assert messages[-1].endswith("이 회사는 보고서를 생성하지 않았습니다.")
    assert "전체 2개년이 모두" in messages[-1]


def test_collect_warnings_says_not_generated_when_history_is_empty():
    item = ReportInput(result=_result_stub(5, "㈜이력없음", "주소"), snapshots=[])
    messages = [w.message for w in collect_warnings(item, select_financial_rows([]))]

    assert messages == ["재무 이력이 없어 이 회사는 보고서를 생성하지 않았습니다."]


# ---------------------------------------------------------------------------
# 폴더/파일명
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    ("raw", "expected"),
    [
        ("(주)홍성산업", "(주)홍성산업"),
        ('㈜a/b:c*d?e"f<g>h|i', "㈜a_b_c_d_e_f_g_h_i"),
        ("이름 끝 점...", "이름 끝 점"),
        ("   ", "회사명없음"),
        (None, "회사명없음"),
        ("con", "_con"),
    ],
)
def test_sanitize_filename_stem(raw, expected):
    assert sanitize_filename_stem(raw) == expected


def test_unique_filename_suffixes_duplicates_case_insensitively():
    taken: set[str] = set()
    assert unique_filename("㈜동명", ".html", taken) == "㈜동명.html"
    assert unique_filename("㈜동명", ".html", taken) == "㈜동명_2.html"
    assert unique_filename("㈜동명", ".html", taken) == "㈜동명_3.html"


def test_allocate_output_dir_never_overwrites(tmp_path):
    today = date(2026, 8, 3)
    first = allocate_output_dir(tmp_path, today)
    (first / "keep.txt").write_text("keep", encoding="utf-8")
    second = allocate_output_dir(tmp_path, today)

    assert first.name == "2026-08-03"
    assert second.name == "2026-08-03_2"
    assert (first / "keep.txt").read_text(encoding="utf-8") == "keep"


# ---------------------------------------------------------------------------
# generate_reports (파일 생성)
# ---------------------------------------------------------------------------


def _result_stub(
    result_id: int,
    name: str,
    address: str,
    parse_status: str = ParseStatus.OK,
    **overrides,
) -> SimpleNamespace:
    fields = {
        "id": result_id,
        "corp_name": name,
        "induty_code": "25110",
        "induty_name": "금속가공제품 제조업",
        "address": address,
        "ceo_name": "홍길동",
        "fiscal_date": "20251231",
        "audit_opinion": "적정",
        "auditor_name": "안경회계법인",
        "parse_status": parse_status,
        "excluded_by_stale_disclosure": 0,
    }
    fields.update(overrides)
    return SimpleNamespace(**fields)


def _snapshot_stub(year: str, revenue: int | None = 100, **overrides) -> SimpleNamespace:
    fields = {
        "fiscal_year": year,
        "current_assets": 10,
        "noncurrent_assets": 10,
        "total_assets": 20,
        "current_liab": 5,
        "noncurrent_liab": 5,
        "total_liab": 10,
        "total_equity": 10,
        "revenue": revenue,
        "cogs": 60,
        "gross_profit": 40,
        "sga": 20,
        "operating_income": 20,
        "net_income": 10,
        "parse_status": ParseStatus.OK,
        "from_current_period": 1,
    }
    fields.update(overrides)
    return SimpleNamespace(**fields)


def test_generate_reports_writes_html_per_company_and_label_excel(report_settings, tmp_path):
    items = [
        ReportInput(
            result=_result_stub(1, "(주)이력있음", "경남 김해시 1"),
            snapshots=[_snapshot_stub("2023"), _snapshot_stub("2024")],
        ),
        ReportInput(result=_result_stub(2, "(주)이력없음", "경남 김해시 2"), snapshots=[]),
    ]

    outcome = generate_reports(items, today=date(2026, 8, 3))

    assert outcome.output_dir == tmp_path / "report" / "2026-08-03"
    # 재무이력이 없는 회사는 **파일을 만들지 않는다**(템플릿 렌더가 중단돼 연락처까지
    # 빈 반쪽 문서가 나오기 때문).
    assert [f.filename for f in outcome.files] == ["(주)이력있음.html"]
    assert not (outcome.output_dir / "(주)이력없음.html").exists()

    html = (outcome.output_dir / "(주)이력있음.html").read_text(encoding="utf-8")
    data = _extract_embedded_json(html)
    assert data["company"]["name"] == "(주)이력있음"
    assert [row["year"] for row in data["financials"]] == [2023, 2024]

    assert [(s.result_id, s.corp_name) for s in outcome.skipped] == [(2, "(주)이력없음")]
    warned = [w for w in outcome.warnings if w.result_id == 2]
    assert warned and "보고서를 생성하지 않았습니다" in warned[-1].message
    assert not [w for w in outcome.warnings if w.result_id == 1]

    # 발송처 라벨 엑셀 — 실제로 생성된 회사만(보고서 없는 회사에 우편을 보내면 안 된다).
    wb = openpyxl.load_workbook(outcome.output_dir / LABEL_FILENAME)
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows == [
        ("회사명", "주소"),
        ("(주)이력있음", "경남 김해시 1"),
    ]


def test_generate_reports_skips_company_whose_only_year_is_incomplete(report_settings):
    """[H1/H2] 유일한 연도의 매출액이 없으면 → 실을 연도 0건 → 생성 자체를 건너뛴다."""
    items = [
        ReportInput(
            result=_result_stub(7, "(주)한해만", "경남 김해시 7"),
            snapshots=[_snapshot_stub("2024", revenue=None)],
        )
    ]
    outcome = generate_reports(items, today=date(2026, 8, 3))
    messages = " / ".join(w.message for w in outcome.warnings)

    assert outcome.files == []
    assert list(outcome.output_dir.glob("*.html")) == []
    assert [s.result_id for s in outcome.skipped] == [7]
    assert "2024년(매출액)" in messages  # 어떤 항목 때문에 제외됐는지
    assert "보고서를 생성하지 않았습니다" in messages


def test_generate_reports_disambiguates_same_company_names(report_settings):
    items = [
        ReportInput(
            result=_result_stub(1, "㈜동명", "주소1"),
            snapshots=[_snapshot_stub("2023"), _snapshot_stub("2024")],
        ),
        ReportInput(
            result=_result_stub(2, "㈜동명", "주소2"),
            snapshots=[_snapshot_stub("2023"), _snapshot_stub("2024")],
        ),
    ]
    outcome = generate_reports(items, today=date(2026, 8, 3))

    assert [f.filename for f in outcome.files] == ["㈜동명.html", "㈜동명_2.html"]
    assert len(list(outcome.output_dir.glob("*.html"))) == 2


def test_generate_reports_rejects_empty_selection(report_settings):
    with pytest.raises(ReportGenerationError):
        generate_reports([])


def test_generated_html_never_contains_unusable_financials(report_settings):
    """[H1/H2] 생성된 HTML의 `financials`는 항상 비어있지 않고 전부 완전한 연도다.

    템플릿 렌더 메인은 `financials[financials.length-1].매출액`을 무조건 읽고
    `calcRatios()`가 null-safe하지 않으므로, 이 두 조건이 깨지면 문서 뒷부분
    (사무소 소개·담당자·연락처 포함)이 통째로 빈 채 인쇄된다.
    """
    items = [
        ReportInput(result=_result_stub(1, "㈜정상", "주소1"), snapshots=[_snapshot_stub("2024")]),
        ReportInput(result=_result_stub(2, "㈜이력없음", "주소2"), snapshots=[]),
        ReportInput(
            result=_result_stub(3, "㈜일부결측", "주소3"),
            snapshots=[_snapshot_stub("2023", current_liab=None), _snapshot_stub("2024")],
        ),
        ReportInput(
            result=_result_stub(4, "㈜매출0", "주소4"),
            snapshots=[_snapshot_stub("2024", revenue=0)],
        ),
    ]
    outcome = generate_reports(items, today=date(2026, 8, 3))

    assert [f.result_id for f in outcome.files] == [1, 3]
    assert [s.result_id for s in outcome.skipped] == [2, 4]
    for entry in outcome.files:
        data = _extract_embedded_json(
            (outcome.output_dir / entry.filename).read_text(encoding="utf-8")
        )
        assert data["financials"], "빈 financials가 실리면 템플릿 렌더가 중단된다"
        for row in data["financials"]:
            assert missing_ratio_inputs(row) == []

    # 결측 연도를 제외한 회사는 그 사실이 경고에 남는다(내용은 있지만 1개년뿐).
    joined = " / ".join(w.message for w in outcome.warnings if w.result_id == 3)
    assert "전체 2개년 중 1개년은 재무 항목이 결측" in joined


def test_generate_reports_warns_once_when_firm_contact_is_placeholder(
    report_settings, monkeypatch
):
    """[M1] 자리표시자 연락처는 회사 수와 무관하게 요청당 1건만 경고한다."""
    items = [
        ReportInput(result=_result_stub(1, "㈜가", "주소1"), snapshots=[_snapshot_stub("2024")]),
        ReportInput(result=_result_stub(2, "㈜나", "주소2"), snapshots=[_snapshot_stub("2024")]),
    ]
    outcome = generate_reports(items, today=date(2026, 8, 3))
    contact_warnings = [w for w in outcome.warnings if "사무소 연락처" in w.message]

    assert len(contact_warnings) == 1
    assert contact_warnings[0].result_id is None

    # 실제 연락처를 채우면 경고가 사라진다.
    monkeypatch.setitem(audit_proposal.FIRM_PROFILE, "contact", "T. 055-000-0000")
    filled = generate_reports(items, today=date(2026, 8, 3))
    assert not [w for w in filled.warnings if "사무소 연락처" in w.message]


def test_firm_profile_placeholder_detection():
    assert firm_profile.is_placeholder_contact(firm_profile.FIRM_PROFILE["contact"]) is True
    assert firm_profile.is_placeholder_contact(None) is True
    assert firm_profile.is_placeholder_contact("   ") is True
    assert firm_profile.is_placeholder_contact("T. 055-000-0000") is False


# ---------------------------------------------------------------------------
# [H4] 자본잠식 회사의 재무안정성 등급 (템플릿 `scoreStability`)
# ---------------------------------------------------------------------------
#
# 템플릿은 원칙적으로 "수정 불필요 영역"이지만, 완전자본잠식(자본총계 < 0) 회사의
# 부채비율(=부채총계/자본총계)이 **음수**가 되어 `last.부채비율 < 0.5` 임계값 비교에
# 걸려 +30점을 받는 결함은 백엔드 필터로 막을 수 없어(음수는 null도 0도 NaN도 아니라
# `select_financial_rows()`의 결측/0분모 가드에 걸리지 않는다) 템플릿을 직접 고쳤다
# (2026-08-03, 사용자 승인). 아래 두 테스트가 그 수정이 되돌아가지 않게 잠근다.


def _extract_js_function_body(template: str, name: str) -> str:
    match = re.search(
        r"function\s+" + re.escape(name) + r"\s*\([^)]*\)\s*\{(.*?)\n\}",
        template,
        re.S,
    )
    assert match, f"템플릿에서 {name}() 함수를 찾지 못했습니다"
    return match.group(1)


def _strip_js_comments(code: str) -> str:
    """JS 주석을 지운 코드만 남긴다.

    이 저장소의 템플릿 수정에는 **왜 고쳤는지를 적은 긴 주석**이 함께 들어가고, 그
    주석이 "수정 전 수식"을 그대로 인용하는 경우가 있다(예: `renderLineChart`의
    `cy = padT+plotH*(1-v/maxV)`). 주석까지 검사하면 "옛 수식이 남아 있으면 실패"
    같은 회귀 가드를 쓸 수 없으므로 코드만 남겨서 본다.
    """
    code = re.sub(r"/\*.*?\*/", "", code, flags=re.S)
    return re.sub(r"^\s*//.*$", "", code, flags=re.M)


def _grade_d_threshold(template: str) -> int:
    """템플릿 `grade()`가 D를 주는 상한(=가장 낮은 임계값)을 템플릿에서 직접 읽는다."""
    body = _extract_js_function_body(template, "grade")
    thresholds = [int(v) for v in re.findall(r"score\s*>=\s*(\d+)", body)]
    assert thresholds, "템플릿 grade()의 점수 임계값을 찾지 못했습니다"
    return min(thresholds)


def test_score_stability_forces_low_score_on_capital_impairment(template_text):
    """[H4] 부채비율이 음수(자본잠식)면 D 등급이 확정되는 가드가 템플릿에 있어야 한다.

    가드가 없거나 임계값 비교(`< 0.5`)보다 뒤에 오면 음수가 "0.5 미만"으로 취급돼
    +30점을 받고, 자본이 전액 잠식된 회사가 A/B(우수·양호)로 인쇄된다.
    """
    body = _extract_js_function_body(template_text, "scoreStability")
    guard = re.search(r"if\s*\(\s*last\.부채비율\s*<\s*0\s*\)\s*return\s+(\d+)\s*;", body)
    assert guard, "자본잠식(부채비율 음수) 가드가 scoreStability()에 없습니다"

    # 반환 점수가 D 등급 구간(템플릿 grade()의 최저 임계값 미만)이어야 한다.
    assert int(guard.group(1)) < _grade_d_threshold(template_text)
    # 가드는 반드시 기존 임계값 비교보다 **먼저** 와야 한다.
    assert body.index(guard.group(0)) < body.index("부채비율 < 0.5")


# 템플릿 JS를 실제로 실행해 등급 문자열·리스크 문구·차트 도형까지 확인하는 end-to-end
# 검증용 최소 DOM 스텁. (jsdom을 새로 설치하지 않기 위해 Node 내장 `vm` + 스텁 document만
# 쓴다.) 출력은 다음 형태다:
#   {파일명: {"grades": {영역: 등급},
#             "risks": [[level, text], ...],
#             "kpi": kpiRow innerHTML,
#             "charts": {컨테이너id: {"nodes": [{tag, attrs, text, html}, ...], "html": ...}}}}
# `charts`의 `nodes`는 컨테이너 아래 SVG/legend 트리를 전부 펼친 것이라, 실제로 그려진
# 도형의 좌표(circle의 cy 등)와 범례 문구를 그대로 검사할 수 있다.
_NODE_RENDER_HARNESS = r"""
const fs = require('fs'), path = require('path'), vm = require('vm');
const dir = process.argv[2];
function makeEl(tag){
  return {
    tagName: tag, _text: '', _html: '', children: [], attrs: {},
    classList: { add(){}, remove(){}, contains(){ return false; } },
    setAttribute(k, v){ this.attrs[k] = v; },
    appendChild(c){ this.children.push(c); return c; },
    get textContent(){ return this._text; },
    set textContent(v){ this._text = String(v); },
    get innerHTML(){ return this._html; },
    set innerHTML(v){ this._html = String(v); },
  };
}
const result = {};
for(const file of fs.readdirSync(dir).filter(f => f.endsWith('.html'))){
  const html = fs.readFileSync(path.join(dir, file), 'utf8');
  const scripts = [...html.matchAll(/<script>([\s\S]*?)<\/script>/g)].map(m => m[1]);
  const els = new Map();
  const document = {
    getElementById(id){ if(!els.has(id)) els.set(id, makeEl(id)); return els.get(id); },
    querySelector(sel){ if(!els.has(sel)) els.set(sel, makeEl(sel)); return els.get(sel); },
    createElement(t){ return makeEl(t); },
    createElementNS(ns, t){ return makeEl(t); },
  };
  const sandbox = { document, console, Math, JSON, Date, Number, String, Array,
                    setTimeout, window: {} };
  sandbox.globalThis = sandbox;
  vm.createContext(sandbox);
  vm.runInContext(scripts[scripts.length - 1], sandbox, { filename: file });
  setTimeout(() => {   // async IIFE의 마이크로태스크가 끝난 뒤 읽는다
    const cards = (els.get('gradeCards') || {})._html || '';
    const grades = {};
    for(const m of cards.matchAll(
        /<div class="label">(.*?)<\/div><div class="grade">(\w)<\/div>/g)){
      grades[m[1]] = m[2];
    }
    const riskHtml = (els.get('riskList') || {})._html || '';
    const risks = [...riskHtml.matchAll(
        /<div class="risk-item risk-(\w+)"><span class="dot"><\/span><span>(.*?)<\/span><\/div>/g)
      ].map(m => [m[1], m[2]]);
    const collect = (el, out) => {
      (el.children || []).forEach(c => {
        out.push({ tag: c.tagName, attrs: c.attrs, text: c._text, html: c._html });
        collect(c, out);
      });
      return out;
    };
    const charts = {};
    for(const id of ['chartStability', 'chartStructure', 'chartRevenue',
                     'chartPeer', 'chartRegionRank']){
      const el = els.get(id);
      charts[id] = el ? { nodes: collect(el, []), html: el._html } : null;
    }
    const kpi = (els.get('kpiRow') || {})._html || '';
    const footnote = (els.get('peerFootnote') || {})._text || '';
    const regionTable = (els.get('#regionTable tbody') || {})._html || '';
    result[file] = { grades, risks, kpi, charts, footnote, regionTable };
    if(Object.keys(result).length === fs.readdirSync(dir).filter(f => f.endsWith('.html')).length){
      console.log(JSON.stringify(result));
    }
  }, 0);
}
"""


def _render_reports_with_node(output_dir: Path, tmp_path: Path) -> dict:
    """생성된 HTML을 Node로 실제 렌더해 {파일명: {"grades":…, "risks":…}}를 돌려준다."""
    harness = tmp_path / "render_report.js"
    harness.write_text(_NODE_RENDER_HARNESS, encoding="utf-8")
    proc = subprocess.run(
        [shutil.which("node"), str(harness), str(output_dir)],
        capture_output=True,
        text=True,
        encoding="utf-8",
        timeout=60,
    )
    assert proc.returncode == 0, proc.stderr
    return json.loads(proc.stdout.strip().splitlines()[-1])


@pytest.mark.skipif(shutil.which("node") is None, reason="Node.js가 없어 템플릿 렌더 검증 생략")
def test_rendered_report_grades_capital_impaired_company_as_d(report_settings, tmp_path):
    """[H4] 생성된 HTML을 실제로 렌더해 자본잠식 회사가 D(주의 필요)로 나오는지 확인.

    같은 요청에 정상(자본총계 양수) 회사를 함께 넣어, 그쪽 등급은 기존과 동일한지
    (회귀 없음) 함께 본다.
    """
    impaired = [  # 자본총계 < 0 → 부채비율 음수
        _snapshot_stub("2024", total_equity=-40, total_liab=60, total_assets=20),
        _snapshot_stub("2025", total_equity=-50, total_liab=70, total_assets=20),
    ]
    items = [
        ReportInput(result=_result_stub(1, "㈜자본잠식", "주소1"), snapshots=impaired),
        ReportInput(
            result=_result_stub(2, "㈜정상", "주소2"),
            snapshots=[_snapshot_stub("2024"), _snapshot_stub("2025")],
        ),
    ]
    outcome = generate_reports(items, today=date(2026, 8, 3))
    assert [f.result_id for f in outcome.files] == [1, 2]

    rendered = _render_reports_with_node(outcome.output_dir, tmp_path)

    # 자본잠식 회사 → 재무안정성 D(주의 필요). 부채비율 -1.4는 예전 로직에서 "0.5 미만"에
    # 걸려 +30점을 받아 B/A로 인쇄됐다.
    assert rendered["㈜자본잠식.html"]["grades"]["재무안정성"] == "D"
    # 정상 회사(부채비율 1.0 / 유동비율 2.0)는 기존과 같은 B.
    assert rendered["㈜정상.html"]["grades"]["재무안정성"] == "B"


# ---------------------------------------------------------------------------
# [H4-2] 자본잠식 회사의 리스크 문구 (템플릿 `buildRiskList`)
# ---------------------------------------------------------------------------
#
# 위 `scoreStability` 수정과 같은 뿌리의 결함이다 — 음수 부채비율이 단순 대소 비교
# `last.부채비율 < first.부채비율`도 통과해, 자본잠식으로 **악화**된 회사에
# "부채비율이 …로 지속 개선되어 재무안정성이 강화되었습니다"라는 초록 문구가 찍혔다
# (개발 DB 실측 81건). 같은 문서의 등급은 D(주의 필요)라 한 보고서 안에서 서술이
# 모순됐다. 2026-08-03에 사용자 승인 하에 템플릿을 두 번째로 고쳤고, 아래 두 테스트가
# 그 수정을 잠근다.

_RISK_IMPROVEMENT_TEXT = "지속 개선되어 재무안정성이 강화"
_RISK_IMPAIRMENT_TEXT = "완전자본잠식"


def test_risk_list_never_calls_capital_impairment_an_improvement(template_text):
    """[H4-2] `buildRiskList()`의 부채비율 개선 판정에 음수 가드가 있어야 한다."""
    body = _extract_js_function_body(template_text, "buildRiskList")

    guard = re.search(r"if\s*\(\s*last\.부채비율\s*<\s*0\s*\)", body)
    assert guard, "자본잠식(부채비율 음수) 가드가 buildRiskList()에 없습니다"

    improvement = re.search(
        r"(?:else\s+)?if\s*\(([^)]*last\.부채비율\s*<\s*first\.부채비율[^)]*)\)", body
    )
    assert improvement, "부채비율 개선 판정 분기를 찾지 못했습니다"
    # 개선 판정은 가드보다 뒤에 있어야 하고, 시작 연도가 음수인 경우(자본잠식 상태에서
    # 시작)도 배제해야 한다 — 음수는 항상 "더 낮은 값"이라 대소 비교가 뒤집힌다.
    assert body.index(guard.group(0)) < body.index(improvement.group(0))
    assert "first.부채비율 >= 0" in improvement.group(1)

    # 가드 분기는 초록(개선)이 아니라 빨강(위험) 항목을 넣어야 한다.
    guard_block = body[body.index(guard.group(0)) : body.index(improvement.group(0))]
    assert "level:'red'" in guard_block
    assert _RISK_IMPAIRMENT_TEXT in guard_block
    assert "level:'green'" not in guard_block


@pytest.mark.skipif(shutil.which("node") is None, reason="Node.js가 없어 템플릿 렌더 검증 생략")
def test_rendered_risk_list_warns_on_capital_impairment_instead_of_improvement(
    report_settings, tmp_path
):
    """[H4-2] 실제 렌더 결과에서 모순 문구가 사라지고 자본잠식 경고가 나오는지 확인.

    실사례(`results.id=6223`, 부채비율 903.1% → -336.8%)와 같은 형태의 회사,
    시작 연도부터 이미 자본잠식이면서 더 악화된 회사, 그리고 정상적으로 부채비율이
    개선된 회사(회귀 확인용)를 함께 넣는다.
    """
    items = [
        ReportInput(  # 903.1% → -336.8% (자본잠식 진입) — 실사례 6223과 같은 형태
            result=_result_stub(1, "㈜자본잠식진입", "주소1"),
            snapshots=[
                _snapshot_stub("2022", total_liab=903, total_equity=100, total_assets=1003),
                _snapshot_stub("2025", total_liab=337, total_equity=-100, total_assets=237),
            ],
        ),
        ReportInput(  # -600% → -1300% (이미 자본잠식 + 악화, 숫자상으론 "감소")
            result=_result_stub(2, "㈜자본잠식악화", "주소2"),
            snapshots=[
                _snapshot_stub("2022", total_liab=600, total_equity=-100, total_assets=500),
                _snapshot_stub("2025", total_liab=1300, total_equity=-100, total_assets=1200),
            ],
        ),
        ReportInput(  # 182.9% → 148.6% (실제 개선) — 기존 초록 문구가 유지돼야 한다
            result=_result_stub(3, "㈜정상개선", "주소3"),
            snapshots=[
                _snapshot_stub("2022", total_liab=1829, total_equity=1000, total_assets=2829),
                _snapshot_stub("2025", total_liab=1486, total_equity=1000, total_assets=2486),
            ],
        ),
    ]
    outcome = generate_reports(items, today=date(2026, 8, 3))
    assert [f.result_id for f in outcome.files] == [1, 2, 3]

    rendered = _render_reports_with_node(outcome.output_dir, tmp_path)

    for filename in ("㈜자본잠식진입.html", "㈜자본잠식악화.html"):
        risks = rendered[filename]["risks"]
        texts = " / ".join(text for _, text in risks)
        assert _RISK_IMPROVEMENT_TEXT not in texts, f"{filename}: 모순된 개선 문구가 남아 있다"
        impairment = [(level, text) for level, text in risks if _RISK_IMPAIRMENT_TEXT in text]
        assert len(impairment) == 1, f"{filename}: 자본잠식 경고가 없거나 중복된다"
        assert impairment[0][0] == "red"
        # 같은 문서의 등급과 서술이 일치해야 한다(등급 D + 빨강 경고).
        assert rendered[filename]["grades"]["재무안정성"] == "D"

    # 정상 개선 회사는 기존과 동일하게 초록 개선 문구가 그대로 나온다.
    normal = rendered["㈜정상개선.html"]["risks"]
    improvement = [(level, text) for level, text in normal if _RISK_IMPROVEMENT_TEXT in text]
    assert len(improvement) == 1
    assert improvement[0][0] == "green"
    assert "182.9%에서 148.6%" in improvement[0][1]
    assert not [text for _, text in normal if _RISK_IMPAIRMENT_TEXT in text]


# ---------------------------------------------------------------------------
# [M-1] 자본구성 도넛 차트 (템플릿 `renderDoughnut`) — 템플릿 무수정 관행의 예외 ③
# ---------------------------------------------------------------------------
#
# 위 H4/H4-2와 같은 뿌리다. 자본총계가 음수면 `total`(=유동부채+비유동부채+자본총계)이
# 부채 합보다 작아져 유동부채 조각의 `frac`이 1을 크게 넘고(원호가 여러 바퀴 겹침),
# 자본총계 조각은 음수 스윕이 되며, 범례에는 "유동부채 903.2% · 자본총계 -864.7%"처럼
# 말이 되지 않는 구성비가 인쇄됐다(실측 result_id=8583). 등급(D)·리스크 문구는 이미
# 자본잠식을 반영하고 있어 차트만 모순되던 상태였다.

_DOUGHNUT_FALLBACK_TEXT = "구성비를 표시할 수 없습니다"


def test_doughnut_chart_guards_negative_slices(template_text):
    """[M-1] 음수 조각이 있으면 도넛(원호·구성비 범례)을 아예 그리지 않아야 한다."""
    body = _strip_js_comments(_extract_js_function_body(template_text, "renderDoughnut"))

    guard = re.search(r"if\s*\(\s*data\.some\([^)]*v\s*<\s*0\s*\)[^{]*\)\s*\{", body)
    assert guard, "음수 조각 가드가 renderDoughnut()에 없습니다"

    guard_at = body.index(guard.group(0))
    # 가드는 원호 계산(frac)과 구성비 범례(pct)보다 **먼저** 오고, 조기 return으로
    # 그 둘에 아예 도달하지 않아야 한다.
    assert guard_at < body.index("frac*Math.PI*2")
    assert guard_at < body.index("pct(data[i]/total)")
    guard_block = body[guard_at : body.index("frac*Math.PI*2")]
    assert "return;" in guard_block
    # 대체 표시는 구성비(%)가 아니라 금액이어야 한다.
    assert "백만원" in guard_block
    assert "pct(" not in guard_block


def test_line_chart_scale_includes_negative_values(template_text):
    """[M-2] 선그래프 y축이 음수까지 포함해 잡히고, 양수 구간은 기존과 동일해야 한다."""
    body = _strip_js_comments(_extract_js_function_body(template_text, "renderLineChart"))

    # 예전 스케일식(0~maxV 고정)은 남아 있으면 안 된다.
    assert "padT+plotH*(1-v/maxV)" not in body
    # 최솟값이 음수면 minV가 음수가 되고, 좌표는 (maxV-v)/span으로 잡는다.
    assert re.search(r"const\s+minV\s*=\s*rawMin\s*<\s*0\s*\?", body), "음수 최솟값 처리가 없습니다"
    assert re.search(r"yOf\s*=\s*v\s*=>\s*padT\s*\+\s*plotH\*\(maxV-v\)/span", body)
    # 0선을 별도로 긋는다(renderGroupedBarChart의 zeroY 패턴과 동일).
    assert "zeroY" in body


def _chart_nodes(rendered_file: dict, container_id: str) -> list[dict]:
    chart = rendered_file["charts"][container_id]
    assert chart is not None, f"{container_id} 컨테이너가 렌더되지 않았습니다"
    return chart["nodes"]


@pytest.mark.skipif(shutil.which("node") is None, reason="Node.js가 없어 템플릿 렌더 검증 생략")
def test_rendered_charts_handle_capital_impairment(report_settings, tmp_path):
    """[M-1/M-2] 자본잠식 회사의 두 차트가 실제 렌더에서 정상적으로 보이는지 확인.

    - 자본구성 도넛: 원호(path)를 그리지 않고 금액 안내로 대체, 범례에 %가 없다.
    - 재무안정성 선그래프: 음수 부채비율 지점이 plot 영역 안에 있어 잘리지 않는다.
    정상(자본총계 양수) 회사도 함께 넣어 기존 렌더가 그대로인지 확인한다.
    """
    items = [
        ReportInput(  # 자본총계 < 0 → 부채비율 음수
            result=_result_stub(1, "㈜자본잠식", "주소1"),
            snapshots=[
                _snapshot_stub("2024", total_liab=60, total_equity=-40, total_assets=20),
                _snapshot_stub("2025", total_liab=70, total_equity=-50, total_assets=20),
            ],
        ),
        ReportInput(
            result=_result_stub(2, "㈜정상", "주소2"),
            snapshots=[_snapshot_stub("2024"), _snapshot_stub("2025")],
        ),
    ]
    outcome = generate_reports(items, today=date(2026, 8, 3))
    rendered = _render_reports_with_node(outcome.output_dir, tmp_path)

    impaired = rendered["㈜자본잠식.html"]
    normal = rendered["㈜정상.html"]

    # --- [M-1] 도넛: 자본잠식이면 원호를 그리지 않고 금액 안내로 대체 ---
    impaired_structure = _chart_nodes(impaired, "chartStructure")
    assert not [n for n in impaired_structure if n["tag"] == "path"], "음수 조각 원호가 그려졌다"
    structure_text = " ".join(
        (n["text"] or "") + " " + (n["html"] or "") for n in impaired_structure
    )
    assert _DOUGHNUT_FALLBACK_TEXT in structure_text
    assert "백만원" in structure_text
    assert "%" not in structure_text, f"무의미한 구성비가 인쇄됐다: {structure_text}"

    # 정상 회사는 기존과 동일하게 도넛 3조각 + 구성비 범례가 그려진다.
    normal_structure = _chart_nodes(normal, "chartStructure")
    assert len([n for n in normal_structure if n["tag"] == "path"]) == 3
    normal_structure_text = " ".join((n["html"] or "") for n in normal_structure)
    assert "%" in normal_structure_text
    assert "-" not in re.sub(r"[^0-9.%\-]", "", normal_structure_text).replace("%", "")

    # --- [M-2] 선그래프: 음수 지점도 plot 영역(padT=14 ~ H-padB=164) 안에 있어야 한다 ---
    for name, chart in (("자본잠식", impaired), ("정상", normal)):
        cys = [
            float(n["attrs"]["cy"])
            for n in _chart_nodes(chart, "chartStability")
            if n["tag"] == "circle"
        ]
        assert cys, f"{name}: 선그래프 데이터 점이 없습니다"
        assert all(14 <= cy <= 164 for cy in cys), f"{name}: 차트 밖으로 잘린 점이 있다 {cys}"

    # 자본잠식 회사에는 0선이 추가로 그려진다(음수 구간이 있다는 표시).
    impaired_lines = [n for n in _chart_nodes(impaired, "chartStability") if n["tag"] == "line"]
    normal_lines = [n for n in _chart_nodes(normal, "chartStability") if n["tag"] == "line"]
    assert len(impaired_lines) == len(normal_lines) + 1


# ---------------------------------------------------------------------------
# [M-3] 음수 분모(매출액 등)가 등급을 반전시키는 문제 — 백엔드에서 막는다
# ---------------------------------------------------------------------------
#
# 실측 result_id=6704: 2021년 매출액 -5,409,049,845 / 영업이익 -1,705,664,600 →
# 영업이익률이 **+31.5%** 로 계산돼 수익성 등급이 A로 반전됐다(KPI에는 "매출액(2021)
# -5,409백만원"이 함께 찍혀 한 장 안에서 모순). 음수 분모는 비율의 부호를 통째로
# 뒤집으므로 결측과 동일하게 그 연도를 제외한다 — **단 `자본총계`는 제외 대상이
# 아니다**(자본잠식은 실재하는 상태라 연도를 버리면 H4가 의도한 "자본잠식 회사도
# 보고서는 생성하되 등급만 D"가 무너진다).


def test_missing_ratio_inputs_flags_negative_denominators_except_equity():
    """[M-3] 매출액/자산총계/유동부채는 음수도 제외, 자본총계 음수는 그대로 싣는다."""
    complete = {key: 10 for key in RATIO_REQUIRED_KEYS}
    complete["year"] = 2021

    assert missing_ratio_inputs(dict(complete, 매출액=-5_409_049_845)) == ["매출액"]
    assert missing_ratio_inputs(dict(complete, 자산총계=-1)) == ["자산총계"]
    assert missing_ratio_inputs(dict(complete, 유동부채=-1)) == ["유동부채"]

    # 자본잠식(자본총계 음수)은 정상 값으로 통과해야 한다 — H4의 전제.
    assert missing_ratio_inputs(dict(complete, 자본총계=-40)) == []
    # 0은 기존대로 모든 분모에서 제외(Infinity 방지).
    assert missing_ratio_inputs(dict(complete, 자본총계=0)) == ["자본총계"]
    # 분모가 아닌 항목의 음수(영업손실 등)는 정상 값이다.
    assert missing_ratio_inputs(dict(complete, 영업이익=-1_705_664_600)) == []
    assert missing_ratio_inputs(dict(complete, 당기순이익=-1)) == []
    assert missing_ratio_inputs(dict(complete, 매출총이익=-1)) == []

    assert "자본총계" not in audit_proposal.RATIO_POSITIVE_DENOMINATOR_KEYS
    assert audit_proposal.RATIO_POSITIVE_DENOMINATOR_KEYS < audit_proposal.RATIO_DENOMINATOR_KEYS


def test_select_financial_rows_drops_negative_revenue_year_but_keeps_capital_impairment():
    """[M-3] 음수 매출액 연도만 빠지고, 자본잠식 연도는 그대로 실린다."""
    snapshots = [
        _snapshot_stub("2020", revenue=4000, operating_income=-400),
        _snapshot_stub("2021", revenue=-5409, operating_income=-1706),  # 제외 대상
        _snapshot_stub("2022", total_liab=70, total_equity=-50, total_assets=20),  # 자본잠식
    ]
    selection = select_financial_rows(snapshots)

    assert [row["year"] for row in selection.rows] == [2020, 2022]
    assert selection.incomplete_years == [("2021", ["매출액"])]
    # 자본잠식 연도는 값 그대로 실려야 한다(H4가 템플릿에서 D 등급을 매기는 근거).
    assert selection.rows[-1]["자본총계"] == -50


def test_collect_warnings_explains_abnormal_denominator_exclusion():
    """[M-3] 음수 때문에 뺀 연도도 "결측"이 아니라 "결측이거나 비정상"으로 안내한다."""
    item = ReportInput(
        result=_result_stub(6704, "㈜매출음수", "주소"),
        snapshots=[
            _snapshot_stub("2020", revenue=4000, operating_income=-400),
            _snapshot_stub("2021", revenue=-5409, operating_income=-1706),
        ],
    )
    joined = " / ".join(
        w.message for w in collect_warnings(item, select_financial_rows(item.snapshots))
    )
    assert "비정상(0/음수)" in joined
    assert "2021년(매출액)" in joined
    assert "생성하지 않았습니다" not in joined  # 남은 연도가 있으므로 생성은 된다


@pytest.mark.skipif(shutil.which("node") is None, reason="Node.js가 없어 템플릿 렌더 검증 생략")
def test_rendered_report_does_not_invert_profitability_on_negative_revenue(
    report_settings, tmp_path
):
    """[M-3] 음수 매출액 연도를 뺀 뒤 수익성 등급이 A로 반전되지 않는지 실제 렌더로 확인.

    같은 요청에 자본잠식 회사를 넣어 **H4의 "자본잠식도 보고서는 생성하되 D"** 가
    이 변경으로 깨지지 않는지 함께 본다(회귀 위험이 가장 큰 지점).
    """
    items = [
        ReportInput(  # 실측 6704과 같은 형태 — 마지막 연도의 매출액이 음수
            result=_result_stub(1, "㈜매출음수", "주소1"),
            snapshots=[
                _snapshot_stub("2020", revenue=4000, operating_income=-400),
                _snapshot_stub("2021", revenue=-5409, operating_income=-1706),
            ],
        ),
        ReportInput(  # 자본잠식(자본총계 음수) — 연도가 제외되지 않고 D로 나와야 한다
            result=_result_stub(2, "㈜자본잠식", "주소2"),
            snapshots=[
                _snapshot_stub("2024", total_liab=60, total_equity=-40, total_assets=20),
                _snapshot_stub("2025", total_liab=70, total_equity=-50, total_assets=20),
            ],
        ),
    ]
    outcome = generate_reports(items, today=date(2026, 8, 3))
    assert [f.result_id for f in outcome.files] == [1, 2]

    # 음수 매출액 연도는 아예 실리지 않는다.
    html = (outcome.output_dir / "㈜매출음수.html").read_text(encoding="utf-8")
    payload = _extract_embedded_json(html)
    assert [row["year"] for row in payload["financials"]] == [2020]

    rendered = _render_reports_with_node(outcome.output_dir, tmp_path)

    negative = rendered["㈜매출음수.html"]
    # 영업이익률 -400/4000 = -10% → 수익성 D. 수정 전에는 2021년(-1706/-5409=+31.5%)이
    # 마지막 연도로 남아 A가 찍혔다.
    assert negative["grades"]["수익성"] == "D"
    assert "-5,409" not in negative["kpi"]  # KPI에도 음수 매출액이 남지 않는다

    # --- H4 회귀 확인: 자본잠식 회사는 여전히 "생성되고" 등급만 D다 ---
    impaired_payload = _extract_embedded_json(
        (outcome.output_dir / "㈜자본잠식.html").read_text(encoding="utf-8")
    )
    assert [row["year"] for row in impaired_payload["financials"]] == [2024, 2025]
    assert impaired_payload["financials"][-1]["자본총계"] == -50
    assert rendered["㈜자본잠식.html"]["grades"]["재무안정성"] == "D"


# ---------------------------------------------------------------------------
# 비교군 (peers / industryAverage / regionGroup) — 2026-08-04
# ---------------------------------------------------------------------------
#
# 소스는 **같은 Job의 다른 results 행**뿐이다(신규 API 호출 0건). 아래 테스트는
#   ① 업종 매칭(소분류 → 중분류 폴백 → 표본 부족 시 빈 값),
#   ② 매출액가중평균(비율끼리 산술평균이 아니다),
#   ③ 대상 회사가 regionGroup에 isTarget으로 들어가는지,
#   ④ 휴면·폐업 추정/재무이력 없는 회사가 후보에서 빠지는지,
#   ⑤ 개수 상한(8 / 대상 포함 16),
#   ⑥ 비율 계산식이 템플릿 `calcRatios()`와 같은지(드리프트 가드)
# 를 잠근다.


def _peer_item(
    result_id: int,
    name: str,
    induty_code: str,
    revenue: int,
    *,
    gross_profit: int = 40,
    years: tuple[str, ...] = ("2023", "2024"),
    snapshots: list | None = None,
    **result_overrides,
) -> ReportInput:
    """비교군 후보 1건(결과 + 최근 연도까지의 스냅샷)을 만든다."""
    if snapshots is None:
        snapshots = [
            _snapshot_stub(year, revenue=revenue, gross_profit=gross_profit) for year in years
        ]
    return ReportInput(
        result=_result_stub(
            result_id, name, f"경남 김해시 {result_id}", induty_code=induty_code,
            **result_overrides,
        ),
        snapshots=snapshots,
    )


def _pool_and_target(items: list[ReportInput], target_index: int = 0):
    """`items` 전체로 풀을 만들고, 그 중 한 건을 대상 회사 후보로 돌려준다."""
    pool = build_peer_pool(items)
    target_item = items[target_index]
    target = build_peer_candidate(
        target_item.result, select_financial_rows(target_item.snapshots)
    )
    assert target is not None
    return pool, target


def test_compute_report_ratios_matches_template_formula(template_text):
    """[드리프트 가드] 비교군 비율식이 템플릿 `calcRatios()`와 글자 그대로 같아야 한다.

    대상 회사의 수치는 템플릿이, 비교군의 수치는 백엔드가 계산하므로 식이 갈리면
    **같은 표 안에서 기준이 다른 숫자**가 나란히 인쇄된다.
    """
    body = _extract_js_function_body(template_text, "calcRatios")
    formulas = {
        "매출총이익률": ("매출총이익", "매출액"),
        "영업이익률": ("영업이익", "매출액"),
        "부채비율": ("부채총계", "자본총계"),
        "자기자본비율": ("자본총계", "자산총계"),
    }
    for name, (numerator, denominator) in formulas.items():
        assert f"{name}: f.{numerator}/f.{denominator}" in body, f"{name} 식이 템플릿과 다릅니다"

    row = {"매출액": 1000, "매출총이익": 250, "영업이익": 100, "부채총계": 300,
           "자본총계": 500, "자산총계": 800}
    assert compute_report_ratios(row) == {
        "매출총이익률": 0.25,
        "영업이익률": 0.1,
        "부채비율": 0.6,
        "자기자본비율": 0.625,
    }
    # 분모가 0이거나 결측이면 계산하지 않는다(ZeroDivisionError로 500이 되면 안 된다).
    assert compute_report_ratios(dict(row, 매출액=0)) is None
    assert compute_report_ratios(dict(row, 자본총계=None)) is None


def test_select_peers_prefers_minor_industry_prefix():
    """소분류(앞 3자리)로 표본이 충분하면 중분류로 넓히지 않는다."""
    items = [
        _peer_item(1, "㈜대상", "25110", 1000),
        _peer_item(2, "㈜같은소분류A", "25119", 1100),
        _peer_item(3, "㈜같은소분류B", "25120", 900),
        _peer_item(4, "㈜같은중분류", "25900", 1000),  # 251이 아니라 259 → 제외돼야 한다
        _peer_item(5, "㈜다른업종", "31000", 1000),
    ]
    pool, target = _pool_and_target(items)

    peers = select_peers(pool, target)
    assert [p.name for p in peers] == ["㈜같은소분류A", "㈜같은소분류B"]
    # 대상 회사 자신은 절대 비교군에 들어가지 않는다.
    assert target.result_id not in [p.result_id for p in peers]


def test_select_peers_falls_back_to_major_industry_prefix():
    """소분류 매칭이 최소 표본(2건) 미만이면 중분류(앞 2자리)로 한 번 넓힌다."""
    items = [
        _peer_item(1, "㈜대상", "25110", 1000),
        _peer_item(2, "㈜같은소분류", "25130", 1000),  # 251 매칭 1건뿐 → 부족
        _peer_item(3, "㈜같은중분류A", "25900", 1000),
        _peer_item(4, "㈜같은중분류B", "25200", 1000),
        _peer_item(5, "㈜다른업종", "31000", 1000),
    ]
    pool, target = _pool_and_target(items)

    peers = select_peers(pool, target)
    assert MIN_COMPARISON_SAMPLE == 2
    # 폴백하면 소분류 매칭 건도 함께 포함된다(중분류가 상위 집합이므로).
    assert sorted(p.name for p in peers) == ["㈜같은소분류", "㈜같은중분류A", "㈜같은중분류B"]
    assert "㈜다른업종" not in [p.name for p in peers]


def test_select_peers_returns_empty_when_both_prefixes_are_short_of_sample():
    """소분류·중분류 모두 표본이 모자라면 빈 값 — 에러가 아니라 정상 케이스다."""
    items = [
        _peer_item(1, "㈜대상", "25110", 1000),
        _peer_item(2, "㈜다른업종A", "31000", 1000),
        _peer_item(3, "㈜다른업종B", "46000", 1000),
    ]
    pool, target = _pool_and_target(items)

    peers = select_peers(pool, target)
    assert peers == []
    assert build_industry_average(peers, target) is None
    # [2026-08-05] 지역 비교군도 같은 업종 매칭을 쓰므로 함께 빈 값이 된다 — 예전에는
    # "같은 Job이면 업종 무관"이라 3행이 실렸다(타업종 회사끼리 매출총이익률 순위를
    # 매기는 표가 인쇄되던 것이 이번 변경의 계기다).
    assert select_region_group(pool, target) == []

    # 업종코드 자체가 비어 있으면(구 Job 등) 매칭을 시도하지 않는다.
    blank = build_peer_candidate(
        _result_stub(9, "㈜업종없음", "주소", induty_code=None),
        select_financial_rows([_snapshot_stub("2024")]),
    )
    assert select_peers(pool, blank) == []


def test_build_industry_average_is_revenue_weighted():
    """비율끼리 산술평균이 아니라 원 금액을 합산해 나눈 매출액가중평균이어야 한다."""
    items = [
        _peer_item(1, "㈜대상", "25110", 500, gross_profit=100),
        _peer_item(2, "㈜소형", "25110", 100, gross_profit=40),   # 40%
        _peer_item(3, "㈜대형", "25110", 900, gross_profit=90),   # 10%
    ]
    pool, target = _pool_and_target(items)
    peers = select_peers(pool, target)
    average = build_industry_average(peers, target)

    assert {p.name for p in peers} == {"㈜소형", "㈜대형"}
    # 가중평균 = (40+90)/(100+900) = 0.13. 산술평균이었다면 0.25다.
    assert average["매출총이익률"] == pytest.approx(0.13)
    assert average["매출총이익률"] != pytest.approx(0.25)
    # 대상 회사(100/500 = 20%)는 평균 계산에 들어가지 않는다.
    assert average["매출총이익률"] == pytest.approx(130 / 1000)
    assert "금속가공제품 제조업" in average["설명"]
    assert "2개사" in average["설명"] and "매출액가중평균" in average["설명"]
    # [M1] 이 표본은 매출 근접순 상한(MAX_PEERS) 절단 **후**의 일부다 — "그 업종에
    # 2개사뿐"으로 읽히지 않게 규모 유사 표본임을 문구에 남긴다(계산값은 무변경).
    assert "매출 규모가 유사한" in average["설명"]
    assert "동일 업종" not in average["설명"]


def test_region_group_includes_target_first_with_is_target_flag():
    items = [
        # 업종 매칭(2026-08-05)이 생겨 이웃도 같은 소분류(251)여야 표에 실린다.
        _peer_item(1, "㈜대상", "25110", 1000),
        _peer_item(2, "㈜이웃A", "25119", 1010),
        _peer_item(3, "㈜이웃B", "25120", 990),
    ]
    pool, target = _pool_and_target(items)

    region = select_region_group(pool, target)
    assert [r["name"] for r in region] == ["㈜대상", "㈜이웃A", "㈜이웃B"]
    assert region[0]["isTarget"] is True
    assert all(r["isTarget"] is False for r in region[1:])
    assert set(region[0]) == {"name", "industry", "매출액", "매출총이익률", "opinion", "isTarget"}
    assert region[0]["industry"] == "금속가공제품 제조업"
    assert region[0]["opinion"] == "적정"
    assert region[0]["매출총이익률"] == pytest.approx(40 / 1000)


def test_region_group_is_empty_when_only_one_other_company():
    items = [
        _peer_item(1, "㈜대상", "25110", 1000),
        _peer_item(2, "㈜유일한이웃", "25110", 1000),
    ]
    pool, target = _pool_and_target(items)

    assert select_region_group(pool, target) == []
    # peers도 같은 최소 표본 기준을 **독립적으로** 적용한다.
    assert select_peers(pool, target) == []


def test_peer_pool_excludes_stale_disclosure_and_companies_without_usable_years():
    items = [
        _peer_item(1, "㈜대상", "25110", 1000),
        _peer_item(2, "㈜휴면추정", "25110", 1000, excluded_by_stale_disclosure=1),
        _peer_item(3, "㈜이력없음", "25110", 1000, snapshots=[]),
        _peer_item(4, "㈜매출결측", "25110", 1000, snapshots=[_snapshot_stub("2024", revenue=None)]),
        _peer_item(5, "㈜정상이웃", "25110", 1000),
        _peer_item(6, "㈜정상이웃2", "25110", 1000),
        # 매출액/총자산 조건 제외 건은 **정상 영업 회사**라 후보로 남는다.
        _peer_item(7, "㈜매출조건제외", "25110", 1000, excluded_by_revenue=1),
    ]
    pool, target = _pool_and_target(items)

    names = {c.name for c in pool.candidates}
    assert names == {"㈜대상", "㈜정상이웃", "㈜정상이웃2", "㈜매출조건제외"}
    assert "㈜휴면추정" not in names
    assert "㈜이력없음" not in names
    assert "㈜매출결측" not in names
    assert "㈜매출조건제외" in {p.name for p in select_peers(pool, target)}


def test_unreliable_opinion_companies_are_excluded_from_peers(template_text):
    """[M3] 의견거절/부적정 회사는 peers 후보에서 통째로 빠진다(한정은 남는다).

    `regionGroup`처럼 값을 null로 두는 방식이 peers에서는 통하지 않기 때문이다 —
    템플릿 `chartPeer` 렌더가 null 체크 없이 `p.매출총이익률*100`을 곱해 JS에서
    `null*100 == 0`, 즉 "0.0%"라는 **틀린 값이 조용히** 인쇄된다(제외가 아니라 왜곡).
    """
    # [드리프트 가드] 템플릿이 여전히 null 체크 없이 곱하고 있음을 확인 — 만약 템플릿이
    # null-safe해지면 이 회피책(후보 제외)을 재검토해야 한다.
    assert "peers.map(p=>+(p.매출총이익률*100).toFixed(1))" in template_text
    assert UNRELIABLE_OPINIONS == {"의견거절", "부적정"}

    items = [
        _peer_item(1, "㈜대상", "25110", 1000),
        _peer_item(2, "㈜의견거절", "25110", 1010, audit_opinion="의견거절"),
        _peer_item(3, "㈜부적정", "25110", 1020, audit_opinion="부적정"),
        _peer_item(4, "㈜한정", "25110", 1030, audit_opinion="한정"),
        _peer_item(5, "㈜적정", "25110", 1040),
    ]
    pool, target = _pool_and_target(items)

    # 후보 풀에는 남는다(regionGroup이 "행은 남기고 값만 감추는" 방식이라 필요하다).
    assert {c.name for c in pool.candidates} == {
        "㈜대상", "㈜의견거절", "㈜부적정", "㈜한정", "㈜적정"
    }

    peers = select_peers(pool, target)
    assert [p.name for p in peers] == ["㈜한정", "㈜적정"]

    # 상한 절단보다 **먼저** 걸러지므로 업종평균 가중평균에도 섞이지 않는다.
    average = build_industry_average(peers, target)
    assert average["매출총이익률"] == pytest.approx((40 + 40) / (1030 + 1040))
    assert "2개사" in average["설명"]


def test_region_group_hides_values_of_unreliable_opinion_companies():
    """[M3] regionGroup은 의견거절/부적정 회사의 **금액/비율만** null로 둔다.

    템플릿 규약(250-252행)대로 순위 차트(`filter(r=>r.매출총이익률!=null)`)에서 빠지고
    표에는 `fmtM`/`pct`가 "-"로 찍는다. 회사명·감사의견은 그대로 남는다.
    """
    items = [
        _peer_item(1, "㈜대상", "25110", 1000),
        _peer_item(2, "㈜의견거절", "25110", 1010, audit_opinion="의견거절"),
        _peer_item(3, "㈜부적정", "25110", 1020, audit_opinion="부적정"),
        _peer_item(4, "㈜한정", "25110", 1030, audit_opinion="한정"),
    ]
    pool, target = _pool_and_target(items)

    region = {row["name"]: row for row in select_region_group(pool, target)}
    assert set(region) == {"㈜대상", "㈜의견거절", "㈜부적정", "㈜한정"}
    for name in ("㈜의견거절", "㈜부적정"):
        assert region[name]["매출액"] is None
        assert region[name]["매출총이익률"] is None
        # 회사명/업종/감사의견은 표에 그대로 인쇄된다(수치만 감춘다).
        assert region[name]["opinion"] in UNRELIABLE_OPINIONS
        assert region[name]["industry"] == "금속가공제품 제조업"
    # 한정의견은 실무상 흔한 의견이라 제외 대상이 아니다 — 정상 값으로 남는다.
    assert region["㈜한정"]["매출액"] == 1030
    assert region["㈜한정"]["매출총이익률"] == pytest.approx(40 / 1030)

    # 대상 회사 자신은 의견거절이어도 감추지 않는다 — 같은 문서의 KPI/차트/등급이
    # 이미 그 수치를 인쇄하므로 지역 표에서만 "-"로 가리면 서술이 어긋나고, 순위
    # 차트에서 대상 막대 강조(targetIdx)까지 사라진다.
    _, disclaimed_target = _pool_and_target(items, target_index=1)
    own_row = select_region_group(pool, disclaimed_target)[0]
    assert own_row["name"] == "㈜의견거절" and own_row["isTarget"] is True
    assert own_row["매출액"] == 1010
    assert own_row["매출총이익률"] == pytest.approx(40 / 1010)


def test_extreme_ratio_candidates_are_excluded_from_comparison_pool():
    """[M2] 비율이 ±100%를 벗어난 후보는 peers/regionGroup **양쪽에서** 빠진다.

    실측 사례(`근하하이테크산업`: 매출 155백만 / 매출원가 3,035백만 → 매출총이익률
    -1857.9%, `parse_status=OK`)가 지역 순위 차트의 축을 통째로 끌고 가 대상 회사
    막대가 7px로 뭉개졌다. 경계값(정확히 ±100%)은 정상으로 본다.
    """
    assert MAX_COMPARISON_ABS_RATIO == 1.0
    items = [
        _peer_item(1, "㈜대상", "25110", 1000),
        # 매출총이익률 -1857.9%(실측 재현) — 155 - 3035 = -2880
        _peer_item(2, "㈜근하하이테크산업", "25110", 155, gross_profit=-2880),
        # 영업이익률 -200%(영업이익만 극단) — 매출총이익률은 정상 범위다.
        _peer_item(
            3, "㈜영업이익극단", "25110", 1000,
            snapshots=[_snapshot_stub("2024", revenue=1000, gross_profit=400,
                                      operating_income=-2000)],
        ),
        # 경계값: 매출총이익률 정확히 +100% / 영업이익률 정확히 -100% → 포함한다.
        _peer_item(
            4, "㈜경계값", "25110", 1000,
            snapshots=[_snapshot_stub("2024", revenue=1000, gross_profit=1000,
                                      operating_income=-1000)],
        ),
        _peer_item(5, "㈜정상이웃", "25110", 1050),
        _peer_item(6, "㈜정상이웃2", "25110", 1060),
    ]
    pool, target = _pool_and_target(items)

    names = {c.name for c in pool.candidates}
    assert "㈜근하하이테크산업" not in names
    assert "㈜영업이익극단" not in names
    assert names == {"㈜대상", "㈜경계값", "㈜정상이웃", "㈜정상이웃2"}

    assert [p.name for p in select_peers(pool, target)] == [
        "㈜경계값", "㈜정상이웃", "㈜정상이웃2"
    ]
    assert [r["name"] for r in select_region_group(pool, target)] == [
        "㈜대상", "㈜경계값", "㈜정상이웃", "㈜정상이웃2"
    ]

    # **대상 회사 자신에게는 적용하지 않는다** — 비교 상대에서만 뺀다.
    _, extreme_target = _pool_and_target(items, target_index=1)
    assert extreme_target.ratios["매출총이익률"] == pytest.approx(-2880 / 155)
    payload = build_report_payload(items[1], peer_pool=pool)
    assert payload["regionGroup"][0]["name"] == "㈜근하하이테크산업"
    assert payload["regionGroup"][0]["isTarget"] is True
    assert payload["regionGroup"][0]["매출총이익률"] == pytest.approx(-2880 / 155)
    # 비교 상대에는 여전히 극단값 후보가 없다(자기 자신도 풀에서 빠져 있다).
    assert {p["name"] for p in payload["peers"]} == {
        "㈜대상", "㈜경계값", "㈜정상이웃", "㈜정상이웃2"
    }


def test_peer_candidate_uses_latest_selected_year_only():
    """후보의 재무값은 그 회사의 `select_financial_rows()` 결과 중 최근 연도 1개다."""
    item = _peer_item(
        1,
        "㈜연도선별",
        "25110",
        0,
        snapshots=[
            _snapshot_stub("2022", revenue=100, gross_profit=10),
            _snapshot_stub("2023", revenue=200, gross_profit=50),
            # 최신 연도지만 파싱 실패라 선별에서 빠진다 → 2023이 최근 연도가 된다.
            _snapshot_stub("2024", revenue=999, gross_profit=999, parse_status=ParseStatus.FAILED),
        ],
    )
    candidate = build_peer_pool([item]).candidates[0]

    assert candidate.year == 2023
    assert candidate.revenue == 200
    assert candidate.ratios["매출총이익률"] == pytest.approx(0.25)


def test_comparison_groups_apply_size_caps_by_revenue_proximity():
    """상한(peers 8 / regionGroup 대상 포함 16)과 "매출 규모가 가까운 순" 정렬."""
    items = [_peer_item(1, "㈜대상", "25110", 1000)]
    # 매출액 1100, 1200, ... 3900 (29개사) — 대상과 가까운 순서가 곧 id 순서다.
    items += [_peer_item(i, f"㈜이웃{i:02d}", "25110", 1000 + 100 * (i - 1)) for i in range(2, 31)]
    pool, target = _pool_and_target(items)

    peers = select_peers(pool, target)
    assert MAX_PEERS == 8
    assert len(peers) == MAX_PEERS
    assert [p.name for p in peers] == [f"㈜이웃{i:02d}" for i in range(2, 10)]

    region = select_region_group(pool, target)
    assert MAX_REGION_GROUP == 16
    assert len(region) == MAX_REGION_GROUP
    assert region[0]["name"] == "㈜대상"
    assert [r["name"] for r in region[1:]] == [f"㈜이웃{i:02d}" for i in range(2, 17)]


def test_build_report_payload_fills_comparison_groups_from_pool():
    """`build_report_payload()`가 풀을 받으면 세 필드를 채우고, 안 받으면 빈 값이다."""
    items = [
        _peer_item(1, "㈜대상", "25110", 1000),
        _peer_item(2, "㈜이웃A", "25110", 1100),
        _peer_item(3, "㈜이웃B", "25110", 900),
    ]
    pool = build_peer_pool(items)

    filled = build_report_payload(items[0], peer_pool=pool)
    assert [p["name"] for p in filled["peers"]] == ["㈜이웃A", "㈜이웃B"]
    assert set(filled["peers"][0]) == {
        "name", "매출액", "매출총이익률", "영업이익률", "부채비율", "자기자본비율"
    }
    assert filled["industryAverage"]["매출총이익률"] == pytest.approx(80 / 2000)
    assert [r["name"] for r in filled["regionGroup"]] == ["㈜대상", "㈜이웃A", "㈜이웃B"]
    # 자유 서술 필드는 이번 범위가 아니다.
    assert filled["opinionSummary"] == ""

    # 풀을 넘기지 않으면 예전과 동일하게 빈 값(구 호출부 호환).
    empty = build_report_payload(items[0])
    assert empty["peers"] == []
    assert empty["industryAverage"] is None
    assert empty["regionGroup"] == []
    assert build_report_payload(items[0], peer_pool=PeerPool())["regionGroup"] == []


@pytest.mark.skipif(shutil.which("node") is None, reason="Node.js가 없어 템플릿 렌더 검증 생략")
def test_rendered_report_draws_comparison_charts(report_settings, tmp_path):
    """비교군을 실제로 렌더해 순위 차트/각주/표가 제대로 그려지는지 확인한다.

    템플릿은 손대지 않았으므로(peers 관련 JS는 이미 null-safe) 검증 대상은
    "우리가 넣은 데이터가 템플릿 계약에 맞는가"다 — 비교군을 채운 뒤에도 렌더가
    중단되지 않고, 대상 회사가 순위 차트에서 강조되는지를 본다.
    """
    pool_items = [
        _peer_item(1, "㈜대상", "25110", 1000, gross_profit=200),
        _peer_item(2, "㈜동종A", "25119", 1100, gross_profit=110),
        _peer_item(3, "㈜동종B", "25120", 900, gross_profit=90),
        _peer_item(4, "㈜타업종", "46000", 1000, gross_profit=500),
    ]
    outcome = generate_reports(
        [pool_items[0]], peer_pool=build_peer_pool(pool_items), today=date(2026, 8, 4)
    )
    rendered = _render_reports_with_node(outcome.output_dir, tmp_path)["㈜대상.html"]

    # 동종업종 그룹 막대: (대상 + 동종 2개사 + 업종평균) 4묶음, 영업이익률은 업종평균만 없다.
    peer_rects = [n for n in _chart_nodes(rendered, "chartPeer") if n["tag"] == "rect"]
    assert len(peer_rects) == 4 + 3
    assert "㈜동종A, ㈜동종B" in rendered["footnote"]
    assert "업종평균" in rendered["footnote"]

    # 지역 순위 가로막대: 동종업종 3개사(대상 + 동종 2개사)만 그려지고 그 중 대상
    # 회사 1건만 강조색(#b8892b)으로 칠해진다. **㈜타업종(46000)은 2026-08-05
    # 업종 필터로 빠진다** — 예전에는 지역만 같으면 실려 4개 막대였다.
    region_rects = [n for n in _chart_nodes(rendered, "chartRegionRank") if n["tag"] == "rect"]
    assert len(region_rects) == 3
    assert len([r for r in region_rects if r["attrs"]["fill"] == "#b8892b"]) == 1

    # 표에도 3행이 들어가고 대상 회사 행에만 target 클래스가 붙는다.
    assert rendered["regionTable"].count("<tr") == 3
    assert "㈜타업종" not in rendered["regionTable"]
    assert rendered["regionTable"].count('class="target"') == 1


def test_generate_reports_embeds_comparison_groups(report_settings):
    """생성된 HTML의 EMBEDDED_DATA에 비교군이 실제로 실린다."""
    pool_items = [
        _peer_item(1, "㈜대상", "25110", 1000),
        _peer_item(2, "㈜이웃A", "25110", 1100),
        _peer_item(3, "㈜이웃B", "25110", 900),
    ]
    outcome = generate_reports(
        [pool_items[0]], peer_pool=build_peer_pool(pool_items), today=date(2026, 8, 4)
    )
    data = _extract_embedded_json(
        (outcome.output_dir / "㈜대상.html").read_text(encoding="utf-8")
    )

    assert [p["name"] for p in data["peers"]] == ["㈜이웃A", "㈜이웃B"]
    assert data["industryAverage"] is not None
    assert [r["name"] for r in data["regionGroup"]] == ["㈜대상", "㈜이웃A", "㈜이웃B"]


# ---------------------------------------------------------------------------
# [L1~L5] 비교군 후속 보완 (2026-08-04, dart-qa 1차 리뷰 잔여 지적)
# ---------------------------------------------------------------------------
#
#   L1 비교 연도가 회사마다 달라도 문서에 표기가 없었다 → peers 각주에 기준 연도
#   L2 비교군 후보의 PARTIAL 연도가 조용히 쓰였다 → 회사당 한 줄 경고
#   L3 `compute_report_ratios()` 분모 가드가 음수를 통과시켰다 → 명시적 가드
#   L5 감사의견이 없으면 지역 표에 빈 태그가 찍혔다 → "미상"
#
# (L4 업종코드 자릿수 비대칭은 구조적 제약이라 문서화만 하고 로직은 그대로 뒀다.)


def test_industry_average_description_states_reference_years(template_text):
    """[L1] 각주 문구에 기준 연도가 들어간다 — 템플릿은 손대지 않는다.

    비교군 후보의 재무값은 각 회사 자신의 "쓸 수 있는 최근 연도"라 대상 회사와 시점이
    다를 수 있다(실측 9.2%). 템플릿 `peerFootnote`가 백엔드가 만든
    `industryAverage.설명`을 그대로 이어 붙여 찍으므로 문자열 조립만으로 해결된다.
    """
    # [드리프트 가드] 템플릿이 여전히 `설명`을 각주에 그대로 찍고 있어야 한다.
    assert "peerFootnote" in template_text
    assert "industryAverage.설명" in template_text

    mixed = [
        _peer_item(1, "㈜대상", "25110", 1000, years=("2023", "2024")),
        _peer_item(2, "㈜동종A", "25110", 1100, years=("2021", "2022")),
        _peer_item(3, "㈜동종B", "25110", 900, years=("2020",)),
    ]
    pool, target = _pool_and_target(mixed)
    peers = select_peers(pool, target)
    description = build_industry_average(peers, target)["설명"]

    assert build_comparison_year_note(peers, target) in description
    assert "대상 2024년" in description
    assert "비교사 2020~2022년" in description
    # 기존 M1 문구(표본 성격 안내)는 그대로 유지된다.
    assert "매출 규모가 유사한" in description and "2개사" in description

    # 연도가 전부 같으면 굳이 대상/비교사를 나눠 쓰지 않는다(각주가 길어지지 않게).
    same = [
        _peer_item(1, "㈜대상", "25110", 1000, years=("2024",)),
        _peer_item(2, "㈜동종A", "25110", 1100, years=("2024",)),
        _peer_item(3, "㈜동종B", "25110", 900, years=("2024",)),
    ]
    same_pool, same_target = _pool_and_target(same)
    same_description = build_industry_average(
        select_peers(same_pool, same_target), same_target
    )["설명"]
    assert same_description.endswith("(2024년 기준)")
    assert "대상" not in same_description


@pytest.mark.skipif(shutil.which("node") is None, reason="Node.js가 없어 템플릿 렌더 검증 생략")
def test_rendered_peer_footnote_and_region_table_carry_year_and_opinion(
    report_settings, tmp_path
):
    """[L1/L5] 실제 렌더 결과에서 각주의 기준 연도와 표의 "미상" 표기를 확인한다."""
    pool_items = [
        _peer_item(1, "㈜대상", "25110", 1000, gross_profit=200, years=("2023", "2024")),
        _peer_item(2, "㈜동종A", "25119", 1100, gross_profit=110, years=("2021",)),
        _peer_item(
            3, "㈜의견없음", "25120", 900, gross_profit=90, years=("2022",),
            audit_opinion=None,
        ),
    ]
    outcome = generate_reports(
        [pool_items[0]], peer_pool=build_peer_pool(pool_items), today=date(2026, 8, 4)
    )
    rendered = _render_reports_with_node(outcome.output_dir, tmp_path)["㈜대상.html"]

    assert "대상 2024년" in rendered["footnote"]
    assert "비교사 2021~2022년" in rendered["footnote"]

    # 감사의견이 없는 회사도 표에 "미상"으로 찍힌다 — 빈 태그(`opinion-tag "></span>`)가
    # 남으면 데이터 없음인지 의견 없음인지 문서만 봐서는 알 수 없다.
    assert ">미상<" in rendered["regionTable"]
    assert 'class="opinion-tag "></span>' not in rendered["regionTable"]


# ---------------------------------------------------------------------------
# [F1] 감사의견 태그 배경색 fallback — 템플릿 무수정 관행의 예외 ④ (2026-08-04)
# ---------------------------------------------------------------------------
#
# `.opinion-tag`는 `color:#fff`만 두고 배경색은 `적정`/`한정`/`의견거절` 3개 클래스
# 규칙에만 있었다. 그 밖의 값 — 기존부터 있던 `부적정`과 위 L5로 새로 채워지는
# `미상` — 은 어느 규칙에도 안 걸려 **배경 없이 흰 글자만** 남았고, 흰색 행이나
# 베이지색 대상 행(`tr.target`) 위에서 인쇄물상 사실상 보이지 않았다(개발 DB 실측
# 143건, 4.7%). `.opinion-tag` 기본 규칙에 회색 fallback 배경 한 줄만 추가해 고쳤다
# (3개 클래스 규칙이 specificity로 그대로 덮어쓰므로 기존 색은 무변경).

_OPINION_TAG_PALETTE = {"적정": "#1e8a5f", "한정": "#c98a10", "의견거절": "#c0392b"}
_OPINION_TAG_MIN_CONTRAST = 3.0  # 흰 글자가 배경 위에서 읽히는 최소 대비
_TARGET_ROW_BACKGROUND = "#f4e6c9"  # tr.target(var(--gold-l)) — 대상 회사 행


def _css_variables(template_text: str) -> dict[str, str]:
    root = re.search(r":root\s*\{(.*?)\}", template_text, re.S)
    assert root, ":root 변수 블록을 찾지 못했습니다"
    return {
        name: value.strip()
        for name, value in re.findall(r"(--[\w-]+)\s*:\s*([^;]+);", root.group(1))
    }


def _resolve_opinion_tag_background(template_text: str, class_attr: str) -> str:
    """`class="opinion-tag X"` 요소에 실제로 적용되는 background를 CSS 규칙에서 푼다.

    브라우저 규칙과 동일하게 (①선택자에 적힌 클래스를 전부 가진 규칙만 적용,
    ②클래스 개수가 많은 쪽(specificity)이 우선, ③같으면 나중에 선언된 쪽이 우선)으로
    고른 뒤 `var(--x)`를 `:root` 값으로 치환한다. 못 찾으면 빈 문자열(= 배경 없음).
    """
    classes = set(class_attr.split())
    variables = _css_variables(template_text)
    best_specificity, background = -1, ""
    for extra, body in re.findall(r"\.opinion-tag([^{,]*)\{([^}]*)\}", template_text):
        required = set(re.findall(r"\.([^.\s]+)", extra))
        if extra.strip() and not required:  # 자손/기타 선택자는 다루지 않는다
            continue
        if not required <= classes:
            continue
        declared = re.search(r"(?:^|;)\s*background\s*:\s*([^;]+)", body)
        if not declared:
            continue
        specificity = 1 + len(required)
        if specificity >= best_specificity:
            best_specificity = specificity
            background = declared.group(1).strip()
    var_ref = re.fullmatch(r"var\((--[\w-]+)\)", background)
    if var_ref:
        assert var_ref.group(1) in variables, f"정의되지 않은 CSS 변수: {background}"
        return variables[var_ref.group(1)]
    return background


def _relative_luminance(hex_color: str) -> float:
    raw = hex_color.lstrip("#")
    channels = [int(raw[i : i + 2], 16) / 255 for i in (0, 2, 4)]
    linear = [c / 12.92 if c <= 0.04045 else ((c + 0.055) / 1.055) ** 2.4 for c in channels]
    return 0.2126 * linear[0] + 0.7152 * linear[1] + 0.0722 * linear[2]


def _contrast_ratio(a: str, b: str) -> float:
    first, second = sorted((_relative_luminance(a), _relative_luminance(b)))
    return (second + 0.05) / (first + 0.05)


@pytest.mark.skipif(shutil.which("node") is None, reason="Node.js가 없어 템플릿 렌더 검증 생략")
def test_rendered_opinion_tag_has_visible_background_for_unknown_opinions(
    report_settings, tmp_path, template_text
):
    """[F1] 지역 비교표의 감사의견 태그가 어떤 값이든 배경색을 갖는지 확인한다.

    실제 렌더 결과에서 태그의 class를 뽑아 템플릿 CSS로 배경색을 풀어 본다 —
    `미상`/`부적정`처럼 전용 클래스 규칙이 없는 값도 기본 규칙의 fallback 배경을
    받아야 하고(흰 글자가 흰 종이에 찍히면 안 된다), 기존 3종은 색이 그대로여야 한다.
    """
    pool_items = [
        _peer_item(1, "㈜대상", "25110", 1000, gross_profit=200),  # 적정(기본값)
        _peer_item(2, "㈜한정", "25110", 1010, audit_opinion="한정"),
        _peer_item(3, "㈜의견거절", "25110", 1020, audit_opinion="의견거절"),
        _peer_item(4, "㈜부적정", "25110", 1030, audit_opinion="부적정"),
        _peer_item(5, "㈜의견없음", "25110", 1040, audit_opinion=None),
    ]
    outcome = generate_reports(
        [pool_items[0]], peer_pool=build_peer_pool(pool_items), today=date(2026, 8, 4)
    )
    rendered = _render_reports_with_node(outcome.output_dir, tmp_path)["㈜대상.html"]

    tags = dict(
        (label, class_attr)
        for class_attr, label in re.findall(
            r'<span class="(opinion-tag[^"]*)">([^<]*)</span>', rendered["regionTable"]
        )
    )
    assert set(tags) == {"적정", "한정", "의견거절", "부적정", UNKNOWN_OPINION_LABEL}

    # ① 기존 3종은 팔레트 색 그대로여야 한다(fallback이 덮어쓰면 회귀다).
    for label, expected in _OPINION_TAG_PALETTE.items():
        assert _resolve_opinion_tag_background(template_text, tags[label]) == expected

    # ② 전용 규칙이 없는 값도 배경이 있어야 하고, 흰 글자가 읽히는 대비여야 한다.
    #    대상 회사 행은 베이지(`tr.target`)라 그 위에서도 태그가 보여야 한다.
    for label in ("부적정", UNKNOWN_OPINION_LABEL):
        background = _resolve_opinion_tag_background(template_text, tags[label])
        assert background, f"'{label}' 태그에 배경색이 없습니다(흰 글자만 남습니다)"
        assert background.lower() not in {"#fff", "#ffffff", "transparent", "none"}
        assert _contrast_ratio(background, "#ffffff") >= _OPINION_TAG_MIN_CONTRAST
        assert _contrast_ratio(background, _TARGET_ROW_BACKGROUND) >= _OPINION_TAG_MIN_CONTRAST


# ---------------------------------------------------------------------------
# [L1] 비교군이 없을 때의 안내 문구 — 템플릿 무수정 관행의 예외 ⑤ (2026-08-05)
# ---------------------------------------------------------------------------
#
# 2026-08-05 업종 필터가 붙으면서 `regionGroup=[]`인 회사가 생겼다(개발 DB 실측
# 후보 1,488건 중 63건, 4.2%). regionGroup이 비면 peers도 반드시 비므로(지역 후보
# 목록이 peers 후보의 상위집합이다) 2페이지의 비교 영역 3개가 통째로 백지로
# 인쇄되는데, peers 쪽에는 이미 각주("동종업종 비교사 데이터 없음")가 있는 반면
# 지역 쪽 두 섹션(순위 차트 · 상세 현황 표)에는 아무 표시가 없었다.
# → 템플릿 **HTML/CSS만** 고쳤다: 안내 문구 div 2개 + `#regionTable tbody:empty`를
#   조건으로 쓰는 CSS 규칙 4줄. **JS 렌더 로직은 무변경**이라 regionGroup이 비어
#   있지 않은 문서(95.8%)는 렌더 결과가 한 글자도 바뀌지 않는다.

_EMPTY_NOTICE_CONDITION = "#page2:has(#regionTable tbody:empty)"
_EMPTY_NOTICE_HIDDEN = ("#regionRankCard", "#regionTable", "#regionOpinionNote")


def _template_style_block(template_text: str) -> str:
    style = re.search(r"<style>(.*?)</style>", template_text, re.S)
    assert style, "<style> 블록을 찾지 못했습니다"
    return re.sub(r"/\*.*?\*/", "", style.group(1), flags=re.S)  # 주석은 선택자가 아니다


def _resolve_display(template_text: str, element: str, *, region_group_empty: bool) -> str:
    """템플릿 CSS만으로 해당 요소의 최종 `display`를 푼다(못 찾으면 빈 문자열 = 기본값).

    이번 예외 수정이 쓰는 조건부 규칙(`#page2:has(#regionTable tbody:empty) …`)만
    해석한다 — 그 조건은 "지역 비교군이 0건이라 tbody가 비었다"와 정확히 같으므로
    `region_group_empty`로 바꿔 평가한다. 조건부 규칙이 기본 규칙보다 뒤에 오고
    specificity도 높아, 마지막에 적용된 선언을 그대로 쓴다.
    """
    display = ""
    for selectors, body in re.findall(r"([^{}]+)\{([^{}]*)\}", _template_style_block(template_text)):
        declared = re.search(r"(?:^|;)\s*display\s*:\s*([^;]+)", body)
        if not declared:
            continue
        for selector in selectors.split(","):
            selector = selector.strip()
            if not selector.endswith(element):
                continue
            condition = selector[: -len(element)].strip()
            if condition and condition != _EMPTY_NOTICE_CONDITION:
                continue  # 이 테스트가 다루지 않는 다른 조건부 규칙
            if condition and not region_group_empty:
                continue  # 조건이 성립하지 않으면 적용되지 않는다
            display = declared.group(1).strip()
    return display


def test_template_hides_empty_comparison_sections_with_notice(template_text):
    """[예외 ⑤] 빈 비교군 안내가 HTML/CSS만으로 들어가 있어야 한다(JS 무변경).

    안내 문구 자체는 항상 HTML에 있고 CSS가 보이고/숨기고를 결정한다 —
    JS를 고치지 않아야 regionGroup이 있는 문서의 렌더 결과가 그대로 유지된다.
    """
    notices = re.findall(r'<div class="footnote empty-note">(※[^<]*)</div>', template_text)
    assert len(notices) == 2, "안내 문구 div 2개(순위 차트 자리 / 상세 현황 표 자리)가 필요합니다"
    for notice in notices:
        # 문구는 섹션 제목("동일 지역·규모 비교군 순위")과 어휘를 맞춘다(2026-08-05 Info-3).
        assert "동일 지역에 비교 가능한 동종업종 회사가 없어" in notice
    # 두 안내는 각각 어느 섹션 자리인지 알 수 있게 문구가 달라야 한다.
    assert notices[0] != notices[1]

    # ① 비교군이 있으면(기존 95.8%) 안내는 숨고 차트·표는 손대지 않는다.
    assert _resolve_display(template_text, ".empty-note", region_group_empty=False) == "none"
    for element in _EMPTY_NOTICE_HIDDEN:
        assert _resolve_display(template_text, element, region_group_empty=False) == ""

    # ② 비교군이 없으면 안내가 뜨고, 빈 차트 카드·머리글만 남은 표·의견거절 각주는 숨는다.
    assert _resolve_display(template_text, ".empty-note", region_group_empty=True) == "block"
    for element in _EMPTY_NOTICE_HIDDEN:
        assert _resolve_display(template_text, element, region_group_empty=True) == "none"

    # ③ JS는 손대지 않았다 — 안내 문구/신설 id를 스크립트가 다루면 안 된다.
    scripts = "".join(re.findall(r"<script>(.*?)</script>", template_text, re.S))
    for token in ("empty-note", "regionRankCard", "regionOpinionNote"):
        assert token not in scripts, f"JS가 {token}을 다루고 있습니다(HTML/CSS만 고쳐야 한다)"


@pytest.mark.skipif(shutil.which("node") is None, reason="Node.js가 없어 템플릿 렌더 검증 생략")
def test_rendered_report_leaves_region_sections_empty_without_industry_peers(
    report_settings, tmp_path, template_text
):
    """[예외 ⑤] 동종업종 비교사가 없는 회사와 있는 회사를 한 번에 렌더해 비교한다.

    앞쪽(비교군 없음)은 표·차트가 비어 CSS 안내가 뜨는 상태여야 하고, 뒤쪽은
    **기존과 똑같이** 행·막대가 그려지고 안내는 숨어 있어야 한다(무회귀).
    """
    pool_items = [
        _peer_item(1, "㈜단독업종", "25110", 1000, gross_profit=200),
        _peer_item(2, "㈜비교대상", "46100", 1000, gross_profit=200),
        _peer_item(3, "㈜동종A", "46110", 1100, gross_profit=110),
        _peer_item(4, "㈜동종B", "46120", 900, gross_profit=90),
    ]
    outcome = generate_reports(
        [pool_items[0], pool_items[1]],
        peer_pool=build_peer_pool(pool_items),
        today=date(2026, 8, 5),
    )
    rendered = _render_reports_with_node(outcome.output_dir, tmp_path)

    # ① 동종업종 후보가 2건 미만이라 regionGroup=[] → 표 0행 · 순위 막대 0개.
    alone = rendered["㈜단독업종.html"]
    assert alone["regionTable"] == ""
    assert [n for n in _chart_nodes(alone, "chartRegionRank") if n["tag"] == "rect"] == []
    # 렌더는 중단되지 않는다(뒤쪽 섹션까지 그대로 채워진다).
    assert alone["grades"] and alone["kpi"]
    assert "동종업종 비교사 데이터 없음" in alone["footnote"]
    # 이때만 CSS 안내가 켜진다.
    assert _resolve_display(template_text, ".empty-note", region_group_empty=True) == "block"

    # ② 동종업종이 있는 회사는 예전과 동일하다 — 표 3행(대상+2), 막대 3개, 안내는 숨김.
    paired = rendered["㈜비교대상.html"]
    assert paired["regionTable"].count("<tr") == 3
    assert paired["regionTable"].count('class="target"') == 1
    assert len([n for n in _chart_nodes(paired, "chartRegionRank") if n["tag"] == "rect"]) == 3
    assert _resolve_display(template_text, ".empty-note", region_group_empty=False) == "none"


def _partial_peer_item(result_id: int, name: str, revenue: int, year: str = "2024"):
    """최근 연도 스냅샷이 PARTIAL인 비교군 후보."""
    return _peer_item(
        result_id,
        name,
        "25110",
        revenue,
        snapshots=[
            _snapshot_stub(year, revenue=revenue, parse_status=ParseStatus.PARTIAL)
        ],
    )


def test_comparison_warning_lists_partial_parsed_peers():
    """[L2] 비교군 후보의 PARTIAL 연도를 회사당 한 줄 경고로 알린다(후보는 뺀다 X)."""
    items = [
        _peer_item(1, "㈜대상", "25110", 1000),
        _partial_peer_item(2, "㈜부분A", 1010),
        _partial_peer_item(3, "㈜부분B", 1020),
        _peer_item(4, "㈜정상이웃", "25110", 1030),
    ]
    pool, _ = _pool_and_target(items)
    # PARTIAL이어도 후보 풀에는 남는다(값이 없는 게 아니라 검수가 덜 된 것이다).
    assert {c.name for c in pool.candidates if c.partial} == {"㈜부분A", "㈜부분B"}

    selection = select_financial_rows(items[0].snapshots)
    warnings = collect_comparison_warnings(items[0], selection, pool)
    assert len(warnings) == 1
    message = warnings[0].message
    assert warnings[0].result_id == 1 and warnings[0].corp_name == "㈜대상"
    assert "비교군에 부분 파싱(PARTIAL)" in message
    # peers와 regionGroup에 같은 회사가 중복으로 들어가도 회사 단위로 한 번만 센다.
    assert "2곳 있습니다" in message
    assert "㈜부분A(2024년)" in message and "㈜부분B(2024년)" in message
    assert "외 " not in message
    assert "㈜정상이웃" not in message

    # 비교군에 PARTIAL이 없으면 경고 자체가 없다.
    clean = [
        _peer_item(1, "㈜대상", "25110", 1000),
        _peer_item(2, "㈜정상A", "25110", 1010),
        _peer_item(3, "㈜정상B", "25110", 1020),
    ]
    clean_pool, _ = _pool_and_target(clean)
    assert (
        collect_comparison_warnings(
            clean[0], select_financial_rows(clean[0].snapshots), clean_pool
        )
        == []
    )
    # 풀을 넘기지 않는 구 호출부에서는 아무 일도 하지 않는다.
    assert collect_comparison_warnings(items[0], selection, None) == []


def test_generate_reports_warns_about_partial_peers_and_truncates_the_list(report_settings):
    """[L2] 생성 경로에도 붙고, 이름이 많으면 "외 N곳"으로 줄인다."""
    items = [_peer_item(1, "㈜대상", "25110", 1000)]
    items += [_partial_peer_item(i, f"㈜부분{i:02d}", 1000 + i) for i in range(2, 8)]  # 6곳

    outcome = generate_reports(
        [items[0]], peer_pool=build_peer_pool(items), today=date(2026, 8, 4)
    )
    messages = [w.message for w in outcome.warnings if "비교군에 부분 파싱" in w.message]
    assert len(messages) == 1
    assert "6곳 있습니다" in messages[0]
    assert messages[0].count("(2024년)") == MAX_WARNING_NAME_SAMPLES
    assert f"외 {6 - MAX_WARNING_NAME_SAMPLES}곳" in messages[0]
    # 경고일 뿐 생성은 막지 않는다.
    assert [f.corp_name for f in outcome.files] == ["㈜대상"]


def test_compute_report_ratios_rejects_negative_denominators_but_allows_capital_impairment():
    """[L3] truthy 검사(`not x`)로는 음수가 통과한다 — 명시적으로 막는다.

    비대칭은 `RATIO_POSITIVE_DENOMINATOR_KEYS`와 같다: 매출액/자산총계는 음수도 제외,
    자본총계는 0만 제외(완전자본잠식은 실재하는 정상 파싱 결과라 계산해야 한다).
    """
    row = {"매출액": 1000, "매출총이익": 250, "영업이익": 100, "부채총계": 300,
           "자본총계": 500, "자산총계": 800}

    assert compute_report_ratios(dict(row, 매출액=-1000)) is None
    assert compute_report_ratios(dict(row, 자산총계=-800)) is None
    assert compute_report_ratios(dict(row, 매출액=0)) is None
    assert compute_report_ratios(dict(row, 자산총계=0)) is None
    assert compute_report_ratios(dict(row, 자본총계=0)) is None

    impaired = compute_report_ratios(dict(row, 자본총계=-500, 부채총계=1300))
    assert impaired is not None
    assert impaired["부채비율"] == pytest.approx(-2.6)
    assert impaired["자기자본비율"] == pytest.approx(-0.625)


def test_region_group_labels_missing_opinion_as_unknown(template_text):
    """[L5] 감사의견이 없으면 지역 표에 "미상"을 넣는다(빈 태그 방지, 템플릿 무수정)."""
    # [드리프트 가드] 템플릿이 opinion 문자열을 클래스와 본문에 그대로 쓴다.
    assert 'class="opinion-tag ${r.opinion}">${r.opinion}</span>' in template_text

    items = [
        _peer_item(1, "㈜대상", "25110", 1000, audit_opinion=None),
        _peer_item(2, "㈜의견없음", "25110", 1010, audit_opinion=""),
        _peer_item(3, "㈜적정", "25110", 1020),
    ]
    pool, target = _pool_and_target(items)
    region = {row["name"]: row for row in select_region_group(pool, target)}

    # 대상 회사에도 같은 규칙을 적용한다(표 안에서 표기가 갈리지 않게).
    assert region["㈜대상"]["opinion"] == UNKNOWN_OPINION_LABEL
    assert region["㈜의견없음"]["opinion"] == UNKNOWN_OPINION_LABEL
    assert region["㈜적정"]["opinion"] == "적정"
    # "미상"은 의견거절/부적정과 달리 신뢰성 문제가 아니므로 수치를 감추지 않는다.
    assert region["㈜의견없음"]["매출액"] == 1010

    # 문서 상단 회사 정보(`company.opinion`)는 무변경 — 보정 범위는 이 표뿐이다.
    assert build_company_payload(items[0].result)["opinion"] == ""


def test_capital_impaired_peer_keeps_explicit_negative_ratios(template_text):
    """[커버리지] 자본총계가 음수인 비교군 후보의 비율 2종이 무엇이 되는지 잠근다.

    지금은 템플릿이 peers의 `부채비율`/`자기자본비율`을 쓰지 않아 무해하지만, 값 자체는
    **음수 그대로**(None이 아니다) 실린다 — 나중에 템플릿이 이 필드를 쓰기 시작하면
    자본잠식 3건과 **같은 유형의 버그**가 조용히 재발한다. 마지막 두 어서션이 그
    카나리아다(템플릿이 이 필드를 쓰기 시작하면 이 테스트가 먼저 깨진다).
    """
    items = [
        _peer_item(1, "㈜대상", "25110", 1000),
        _peer_item(
            2, "㈜자본잠식", "25110", 1010,
            snapshots=[
                _snapshot_stub(
                    "2024", revenue=1010, total_equity=-40, total_liab=60, total_assets=20
                )
            ],
        ),
        _peer_item(3, "㈜정상이웃", "25110", 1020),
    ]
    pool, target = _pool_and_target(items)

    # 자본잠식은 후보에서 빠지지 않는다(`has_extreme_ratios`는 손익 비율만 본다).
    candidate = next(c for c in pool.candidates if c.name == "㈜자본잠식")
    assert candidate.ratios["부채비율"] == pytest.approx(60 / -40)
    assert candidate.ratios["자기자본비율"] == pytest.approx(-40 / 20)

    rows = {row["name"]: row for row in build_peer_rows(select_peers(pool, target))}
    assert rows["㈜자본잠식"]["부채비율"] == pytest.approx(-1.5)
    assert rows["㈜자본잠식"]["자기자본비율"] == pytest.approx(-2.0)
    # 지역 표에 실리는 값(매출액/매출총이익률)은 자본잠식과 무관하게 정상이다.
    assert select_region_group(pool, target)[0]["매출총이익률"] is not None

    # [카나리아] 템플릿은 peers의 이 두 필드를 아직 쓰지 않는다 — 쓰기 시작하면
    # 음수 부채비율이 차트/등급에 어떻게 들어가는지 먼저 검토해야 한다.
    assert "p.부채비율" not in template_text
    assert "p.자기자본비율" not in template_text


# ---------------------------------------------------------------------------
# [2026-08-05] regionGroup 업종 매칭 — peers와 같은 헬퍼를 공유한다
# ---------------------------------------------------------------------------
#
# 사용자가 실제 생성된 문서의 "비교군 상세 현황" 표에서 조선부품/시내버스운송/도장공사/
# 상품중개가 한 장에 나란히 실린 것을 보고 "업종기준으로 필터해야지"라고 지적해 추가된
# 규칙이다. peers에 이미 있던 소분류(3자리) → 중분류(2자리) 1회 폴백을
# `match_by_industry_prefix()`로 추출해 **두 비교군이 같은 함수를 호출**한다.


def test_match_by_industry_prefix_rules():
    """[헬퍼 단위] 소분류 우선 → 표본 부족 시 중분류 1회 폴백 → 그래도 부족하면 빈 값."""
    pool = [
        _peer_item(2, "㈜소분류A", "25119", 1000),
        _peer_item(3, "㈜소분류B", "25120", 1000),
        _peer_item(4, "㈜중분류", "25900", 1000),
        _peer_item(5, "㈜타업종", "46000", 1000),
    ]
    candidates = list(build_peer_pool(pool).candidates)

    # 소분류(251)로 2건 이상 → 중분류(25900)까지 넓히지 않는다.
    assert [c.name for c in match_by_industry_prefix(candidates, "25110")] == [
        "㈜소분류A", "㈜소분류B"
    ]
    # 소분류(259)는 1건뿐 → 중분류(25)로 한 번 넓힌다(소분류 매칭 건도 함께 들어온다).
    assert {c.name for c in match_by_industry_prefix(candidates, "25910")} == {
        "㈜소분류A", "㈜소분류B", "㈜중분류"
    }
    # 대분류(4)까지는 넓히지 않는다 — 46000 한 건뿐이라 빈 값이다.
    assert match_by_industry_prefix(candidates, "46100") == []
    # 업종코드가 없으면(구 Job 등) 매칭 자체를 시도하지 않는다.
    assert match_by_industry_prefix(candidates, "") == []
    assert match_by_industry_prefix(candidates, "  ") == []
    # 코드가 prefix 길이보다 짧으면 그 단계를 건너뛴다(2자리 회사 → 중분류만).
    assert {c.name for c in match_by_industry_prefix(candidates, "25")} == {
        "㈜소분류A", "㈜소분류B", "㈜중분류"
    }


def test_region_group_excludes_other_industries():
    """지역 비교군에서 업종이 다른 회사가 빠진다(이번 변경의 본체)."""
    items = [
        _peer_item(1, "㈜대상", "25110", 1000),
        _peer_item(2, "㈜동종A", "25119", 1010),
        _peer_item(3, "㈜동종B", "25120", 990),
        _peer_item(4, "㈜시내버스", "49231", 1005),
        _peer_item(5, "㈜상품중개", "46102", 995),
    ]
    pool, target = _pool_and_target(items)

    region = select_region_group(pool, target)
    # 매출 차이가 동률(±10)이라 id 오름차순으로 A → B.
    assert [r["name"] for r in region] == ["㈜대상", "㈜동종A", "㈜동종B"]
    assert not {"㈜시내버스", "㈜상품중개"} & {r["name"] for r in region}
    # 기존 규칙은 그대로 — 대상이 맨 앞 isTarget, 나머지는 매출 근접순.
    assert region[0]["isTarget"] is True
    assert all(r["isTarget"] is False for r in region[1:])
    # peers는 무영향(같은 매칭 결과를 공유한다).
    assert [p.name for p in select_peers(pool, target)] == ["㈜동종A", "㈜동종B"]


def test_region_group_falls_back_to_major_industry_prefix():
    """소분류 매칭이 2건 미만이면 regionGroup도 peers와 **같은 조건**으로 중분류 폴백."""
    items = [
        _peer_item(1, "㈜대상", "25110", 1000),
        _peer_item(2, "㈜같은소분류", "25130", 1000),  # 251 매칭 1건뿐 → 부족
        _peer_item(3, "㈜같은중분류A", "25900", 1000),
        _peer_item(4, "㈜같은중분류B", "25200", 1000),
        _peer_item(5, "㈜타업종", "46000", 1000),
    ]
    pool, target = _pool_and_target(items)

    region = {r["name"] for r in select_region_group(pool, target)}
    assert region == {"㈜대상", "㈜같은소분류", "㈜같은중분류A", "㈜같은중분류B"}
    assert "㈜타업종" not in region
    # 폴백 단계까지 peers와 동일해야 한다(같은 헬퍼를 쓰므로 드리프트가 없다).
    assert region - {"㈜대상"} == {p.name for p in select_peers(pool, target)}


def test_region_group_is_empty_when_industry_match_is_short_of_sample():
    """중분류 폴백까지 실패하면 `regionGroup=[]` — peers의 기존 규칙을 그대로 따른다.

    1개사만 놓고 "지역 내 순위"라고 인쇄하면 대외 문서로서 오해를 주기 때문이다
    (템플릿은 빈 배열을 "비교군 없음"으로 안전하게 렌더한다).
    """
    items = [
        _peer_item(1, "㈜대상", "25110", 1000),
        _peer_item(2, "㈜유일한동종", "25119", 1010),  # 251 매칭 1건뿐
        _peer_item(3, "㈜타업종A", "46000", 1000),
        _peer_item(4, "㈜타업종B", "49000", 1000),
        _peer_item(5, "㈜타업종C", "20000", 1000),
    ]
    pool, target = _pool_and_target(items)

    assert select_region_group(pool, target) == []
    assert select_region_candidates(pool, target) == []
    assert select_peers(pool, target) == []

    # 업종코드가 비어 있는 회사도 같은 이유로 빈 값이다(매칭 자체가 불가능).
    blank = build_peer_candidate(
        _result_stub(9, "㈜업종없음", "주소", induty_code=None),
        select_financial_rows([_snapshot_stub("2024")]),
    )
    assert select_region_group(pool, blank) == []


def test_region_group_industry_filter_runs_before_the_size_cap():
    """업종 필터는 매출 근접순 상한 절단 **앞에** 온다.

    뒤에 두면 상한 15칸이 매출만 가까운 타업종으로 채워져 업종 필터가 사실상 무력해진다
    (여기서는 타업종 20곳이 매출이 훨씬 가깝지만 동종 3곳만 남아야 한다).
    """
    items = [_peer_item(1, "㈜대상", "25110", 1000)]
    items += [_peer_item(i, f"㈜타업종{i:02d}", "46000", 1000 + i) for i in range(2, 22)]
    items += [_peer_item(30 + i, f"㈜동종{i}", "25110", 5000 + i) for i in range(3)]
    pool, target = _pool_and_target(items)

    region = select_region_group(pool, target)
    assert [r["name"] for r in region] == ["㈜대상", "㈜동종0", "㈜동종1", "㈜동종2"]
    assert len(region) < MAX_REGION_GROUP


def test_region_group_counts_unreliable_opinion_companies_as_sample():
    """의견거절/부적정은 regionGroup에서 **표본으로는 센다**(peers와 적용 지점이 다름).

    regionGroup은 그 회사들의 행을 남기고 금액/비율만 감추므로(`_region_row()`),
    업종 매칭·최소 표본 판정에서 미리 빼면 표에 실릴 회사와 셈이 어긋난다.
    """
    items = [
        _peer_item(1, "㈜대상", "25110", 1000),
        _peer_item(2, "㈜의견거절", "25110", 1010, audit_opinion="의견거절"),
        _peer_item(3, "㈜부적정", "25110", 1020, audit_opinion="부적정"),
    ]
    pool, target = _pool_and_target(items)

    region = {r["name"]: r for r in select_region_group(pool, target)}
    assert set(region) == {"㈜대상", "㈜의견거절", "㈜부적정"}
    for name in ("㈜의견거절", "㈜부적정"):
        assert region[name]["매출액"] is None and region[name]["매출총이익률"] is None
    # peers 쪽은 두 회사를 후보에서 통째로 빼므로 표본 부족으로 빈 값이다.
    assert select_peers(pool, target) == []


# ---------------------------------------------------------------------------
# [M2] 라벨 엑셀 — 제어문자 방어
# ---------------------------------------------------------------------------


def test_excel_safe_text_strips_control_characters():
    assert excel_safe_text("㈜제어\x01문자\x1f") == "㈜제어문자"
    assert excel_safe_text(None) == ""
    assert excel_safe_text(" 여백 ") == "여백"
    assert excel_safe_text("줄\n바꿈") == "줄\n바꿈"  # 개행은 엑셀이 허용한다


def test_generate_reports_survives_control_characters_in_company_name(report_settings):
    items = [
        ReportInput(
            result=_result_stub(1, "㈜제어\x01문자", "경남 김해시\x07 1"),
            snapshots=[_snapshot_stub("2024")],
        )
    ]
    outcome = generate_reports(items, today=date(2026, 8, 3))

    wb = openpyxl.load_workbook(outcome.output_dir / LABEL_FILENAME)
    assert list(wb.active.iter_rows(values_only=True)) == [
        ("회사명", "주소"),
        ("㈜제어문자", "경남 김해시 1"),
    ]


def test_write_label_workbook_wraps_openpyxl_errors(tmp_path, monkeypatch):
    """openpyxl의 IllegalCharacterError/ValueError도 507용 예외로 통일한다."""
    from openpyxl.utils.exceptions import IllegalCharacterError

    class _BoomSheet:
        title = ""

        def append(self, *args, **kwargs):
            raise IllegalCharacterError("bad cell")

    class _BoomWorkbook:
        active = _BoomSheet()

    monkeypatch.setattr(audit_proposal, "Workbook", lambda: _BoomWorkbook())
    with pytest.raises(ReportGenerationError, match="발송처 목록 엑셀"):
        write_label_workbook([], tmp_path / "x.xlsx")


# ---------------------------------------------------------------------------
# API
# ---------------------------------------------------------------------------


@pytest.fixture
def client_with_db():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine, autoflush=False, autocommit=False)

    def _override_get_db():
        db = factory()
        try:
            yield db
        finally:
            db.close()

    app_main.app.dependency_overrides[get_db] = _override_get_db
    yield TestClient(app_main.app), factory
    app_main.app.dependency_overrides.clear()


def _seed(factory) -> tuple[int, int, int]:
    """Job 1개 + 결과 2건(재무이력 있음/없음)을 넣고 (job_id, id_있음, id_없음)을 반환."""
    db = factory()
    try:
        job = Job(
            created_at="2026-08-03T00:00:00",
            name="보고서 테스트 Job",
            cond_region="{}",
            cond_revenue="{}",
            cond_industry="[]",
            cond_period="{}",
            status=JobStatus.DONE,
            current_step=6,
            progress_done=2,
            progress_total=2,
        )
        db.add(job)
        db.commit()
        db.refresh(job)

        with_history = Result(
            job_id=job.id,
            corp_code="00100001",
            rcept_no="20260601000001",
            corp_name="(주)이력있음",
            address="경상남도 김해시 삼계로 1",
            ceo_name="홍길동",
            induty_code="25",
            induty_name="금속가공제품 제조업",
            fiscal_date="20251231",
            audit_opinion="적정",
            auditor_name="안경회계법인",
            parse_status=ParseStatus.OK,
        )
        without_history = Result(
            job_id=job.id,
            corp_code="00100002",
            rcept_no=None,
            corp_name="(주)이력없음",
            address="경상남도 김해시 분성로 2",
            ceo_name="김철수",
            induty_code="25",
            induty_name="금속가공제품 제조업",
            fiscal_date="20251231",
            parse_status=ParseStatus.FAILED,
        )
        db.add_all([with_history, without_history])
        db.commit()
        db.refresh(with_history)
        db.refresh(without_history)

        db.add_all(
            [
                FinancialSnapshot(
                    result_id=with_history.id,
                    rcept_no="20260601000001",
                    fiscal_year=str(year),
                    current_assets=10,
                    noncurrent_assets=10,
                    total_assets=20,
                    current_liab=5,
                    noncurrent_liab=5,
                    total_liab=10,
                    total_equity=10,
                    revenue=100 + year,
                    cogs=60,
                    gross_profit=40,
                    sga=20,
                    operating_income=20,
                    net_income=10,
                    parse_status=ParseStatus.OK,
                    from_current_period=1,
                )
                for year in (2023, 2024)
            ]
        )
        db.commit()
        return job.id, with_history.id, without_history.id
    finally:
        db.close()


def test_generate_report_endpoint_creates_folder_and_files(
    client_with_db, report_settings, tmp_path
):
    client, factory = client_with_db
    job_id, id_with, id_without = _seed(factory)

    resp = client.post(
        f"/api/jobs/{job_id}/generate-report", json={"ids": [id_without, id_with]}
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()

    output_dir = Path(body["output_dir"])
    assert output_dir.parent == tmp_path / "report"
    assert re.fullmatch(r"\d{4}-\d{2}-\d{2}(_\d+)?", output_dir.name)
    # 재무이력이 없는 회사는 생성되지 않는다(선택 2건 중 1건).
    assert body["generated_count"] == 1
    assert body["label_file"] == LABEL_FILENAME
    assert (output_dir / LABEL_FILENAME).is_file()

    # id 오름차순으로 생성된다(선택 다운로드와 동일한 정렬 계약).
    assert [f["result_id"] for f in body["files"]] == [id_with]
    for entry in body["files"]:
        assert (output_dir / entry["filename"]).is_file()

    # 재무이력이 있는 회사는 실제 연도 데이터가 실린다.
    target = next(f for f in body["files"] if f["result_id"] == id_with)
    data = _extract_embedded_json(
        (output_dir / target["filename"]).read_text(encoding="utf-8")
    )
    assert [row["year"] for row in data["financials"]] == [2023, 2024]
    assert data["financials"][0]["매출액"] == 2123

    # 이력 없는 회사는 skipped + 경고로 알리고 요청 자체는 성공한다.
    assert [s["result_id"] for s in body["skipped"]] == [id_without]
    company_warnings = [w for w in body["warnings"] if w["result_id"] is not None]
    assert [w["result_id"] for w in company_warnings] == [id_without]
    assert "보고서를 생성하지 않았습니다" in company_warnings[0]["message"]


def test_generate_report_endpoint_rejects_foreign_and_invalid_ids(
    client_with_db, report_settings
):
    client, factory = client_with_db
    job_id, id_with, _ = _seed(factory)

    foreign = client.post(
        f"/api/jobs/{job_id}/generate-report", json={"ids": [id_with, 999_999]}
    )
    assert foreign.status_code == 400
    assert "999999" in foreign.json()["detail"]

    overflow = client.post(f"/api/jobs/{job_id}/generate-report", json={"ids": [2**63]})
    assert overflow.status_code == 400

    empty = client.post(f"/api/jobs/{job_id}/generate-report", json={"ids": []})
    assert empty.status_code == 422  # pydantic min_length=1

    missing_job = client.post("/api/jobs/999999/generate-report", json={"ids": [id_with]})
    assert missing_job.status_code == 404


def _seed_peer_job(factory) -> tuple[int, dict[str, int]]:
    """비교군 검증용 Job — 같은 업종/다른 업종/휴면 추정/이력 없음을 섞어 넣는다.

    반환값은 (job_id, {회사명: results.id}).
    """
    db = factory()
    try:
        job = Job(
            created_at="2026-08-04T00:00:00",
            name="비교군 테스트 Job",
            cond_region="{}",
            cond_revenue="{}",
            cond_industry="[]",
            cond_period="{}",
            status=JobStatus.DONE,
            current_step=6,
            progress_done=5,
            progress_total=5,
        )
        db.add(job)
        db.commit()
        db.refresh(job)

        # (회사명, 업종코드, 매출액, 매출총이익, 휴면추정, 이력유무)
        spec = [
            ("(주)대상", "25110", 1000, 200, 0, True),
            ("(주)동종A", "25119", 1100, 110, 0, True),
            ("(주)동종B", "25120", 900, 90, 0, True),
            ("(주)타업종", "46000", 1000, 500, 0, True),
            ("(주)휴면추정", "25110", 1000, 900, 1, True),  # 비교군에서 빠져야 한다
            ("(주)이력없음", "25110", 1000, 900, 0, False),  # 스냅샷 0건 → 후보 아님
        ]
        ids: dict[str, int] = {}
        for name, code, revenue, gross_profit, stale, has_history in spec:
            result = Result(
                job_id=job.id,
                corp_code=f"0010{len(ids):04d}",
                rcept_no="20260601000001",
                corp_name=name,
                address=f"경상남도 김해시 삼계로 {len(ids)}",
                ceo_name="홍길동",
                induty_code=code,
                induty_name="금속가공제품 제조업",
                fiscal_date="20251231",
                audit_opinion="적정",
                auditor_name="안경회계법인",
                parse_status=ParseStatus.OK,
                excluded_by_stale_disclosure=stale,
            )
            db.add(result)
            db.commit()
            db.refresh(result)
            ids[name] = result.id
            if not has_history:
                continue
            db.add_all(
                [
                    FinancialSnapshot(
                        result_id=result.id,
                        rcept_no="20260601000001",
                        fiscal_year=str(year),
                        current_assets=10,
                        noncurrent_assets=10,
                        total_assets=20,
                        current_liab=5,
                        noncurrent_liab=5,
                        total_liab=10,
                        total_equity=10,
                        revenue=revenue,
                        cogs=revenue - gross_profit,
                        gross_profit=gross_profit,
                        sga=20,
                        operating_income=20,
                        net_income=10,
                        parse_status=ParseStatus.OK,
                        from_current_period=1,
                    )
                    for year in (2023, 2024)
                ]
            )
            db.commit()
        return job.id, ids
    finally:
        db.close()


def test_generate_report_endpoint_fills_comparison_groups_from_same_job(
    client_with_db, report_settings
):
    """[2026-08-04] 비교군은 선택 회사가 아니라 **같은 Job 전체**에서 온다.

    대상 1건만 선택해도 같은 Job의 다른 회사가 peers/regionGroup에 실려야 하고,
    휴면·폐업 추정과 재무이력 0건인 회사는 빠져야 한다.
    """
    client, factory = client_with_db
    job_id, ids = _seed_peer_job(factory)

    resp = client.post(
        f"/api/jobs/{job_id}/generate-report", json={"ids": [ids["(주)대상"]]}
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["generated_count"] == 1

    output_dir = Path(body["output_dir"])
    data = _extract_embedded_json(
        (output_dir / body["files"][0]["filename"]).read_text(encoding="utf-8")
    )

    # 소분류(251) 매칭 2건 — 타업종/휴면추정/이력없음은 들어오지 않는다.
    assert [p["name"] for p in data["peers"]] == ["(주)동종A", "(주)동종B"]
    assert data["industryAverage"]["매출총이익률"] == pytest.approx(200 / 2000)
    assert "2개사" in data["industryAverage"]["설명"]

    # 지역 비교군도 **같은 업종 매칭**을 쓴다(2026-08-05) — 대상이 맨 앞 isTarget,
    # 타업종/휴면/이력없음은 제외. 예전에는 업종 무관이라 (주)타업종도 실렸다.
    region = data["regionGroup"]
    assert region[0]["name"] == "(주)대상" and region[0]["isTarget"] is True
    assert sorted(r["name"] for r in region[1:]) == ["(주)동종A", "(주)동종B"]
    assert not [
        r for r in region if r["name"] in ("(주)타업종", "(주)휴면추정", "(주)이력없음")
    ]

    assert data["opinionSummary"] == ""  # 자유 서술 필드는 이번 범위가 아니다


def _seed_many_companies(factory, count: int) -> int:
    """같은 Job에 회사 `count`건(각 2개년 스냅샷)을 넣고 job_id를 돌려준다."""
    db = factory()
    try:
        job = Job(
            created_at="2026-08-04T00:00:00",
            name="N+1 회귀 가드 Job",
            cond_region="{}",
            cond_revenue="{}",
            cond_industry="[]",
            cond_period="{}",
            status=JobStatus.DONE,
            current_step=6,
            progress_done=count,
            progress_total=count,
        )
        db.add(job)
        db.commit()
        db.refresh(job)

        results = [
            Result(
                job_id=job.id,
                corp_code=f"0020{i:04d}",
                rcept_no="20260601000001",
                corp_name=f"(주)회사{i:02d}",
                address=f"경상남도 김해시 삼계로 {i}",
                ceo_name="홍길동",
                induty_code="25110",
                induty_name="금속가공제품 제조업",
                fiscal_date="20251231",
                audit_opinion="적정",
                auditor_name="안경회계법인",
                parse_status=ParseStatus.OK,
                excluded_by_stale_disclosure=0,
            )
            for i in range(count)
        ]
        db.add_all(results)
        db.commit()
        for result in results:
            db.refresh(result)

        db.add_all(
            [
                FinancialSnapshot(
                    result_id=result.id,
                    rcept_no="20260601000001",
                    fiscal_year=str(year),
                    current_assets=10,
                    noncurrent_assets=10,
                    total_assets=20,
                    current_liab=5,
                    noncurrent_liab=5,
                    total_liab=10,
                    total_equity=10,
                    revenue=1000 + index,
                    cogs=960 + index,
                    gross_profit=40,
                    sga=20,
                    operating_income=20,
                    net_income=10,
                    parse_status=ParseStatus.OK,
                    from_current_period=1,
                )
                for index, result in enumerate(results)
                for year in (2023, 2024)
            ]
        )
        db.commit()
        return job.id
    finally:
        db.close()


def test_load_job_peer_pool_runs_two_queries_regardless_of_company_count(client_with_db):
    """[커버리지] 비교군 후보 풀 조회는 회사 수와 무관하게 **쿼리 2건**이어야 한다.

    회사마다 재무이력을 조회하도록 회귀하면(N+1) 수천 건 Job에서 보고서 생성이
    통째로 느려지는데, 기존 엔드포인트 테스트는 회사 6건짜리라 그 회귀가 조용히
    통과한다. 여기서는 실제로 실행된 SQL 문 수를 세어 명시적으로 잠근다.
    """
    _, factory = client_with_db
    job_id = _seed_many_companies(factory, 20)

    db = factory()
    try:
        engine = db.get_bind()
        statements: list[str] = []

        def _record(conn, cursor, statement, parameters, context, executemany):
            statements.append(statement)

        event.listen(engine, "before_cursor_execute", _record)
        try:
            pool = _load_job_peer_pool(db, job_id)
        finally:
            event.remove(engine, "before_cursor_execute", _record)
    finally:
        db.close()

    assert len(pool.candidates) == 20
    assert len(statements) == 2, statements  # 결과 목록 1건 + 스냅샷 전체 1건
    assert sum(1 for s in statements if "financial_snapshots" in s) == 1


def test_generate_report_endpoint_returns_friendly_error_on_io_failure(
    client_with_db, report_settings, monkeypatch
):
    """파일 저장 실패는 500 스택트레이스가 아니라 읽을 수 있는 메시지로 응답한다."""
    client, factory = client_with_db
    job_id, id_with, _ = _seed(factory)

    def _boom(*args, **kwargs):
        raise ReportGenerationError("보고서 폴더를 만들 수 없습니다: 디스크 공간 부족")

    monkeypatch.setattr(audit_proposal, "allocate_output_dir", _boom)

    resp = client.post(f"/api/jobs/{job_id}/generate-report", json={"ids": [id_with]})
    assert resp.status_code == 507
    assert "보고서 폴더를 만들 수 없습니다" in resp.json()["detail"]
