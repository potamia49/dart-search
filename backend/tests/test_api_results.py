"""app/api/results.py 라우터 테스트 (결과 조회 + M4 export).

test_api_jobs.py와 동일한 패턴으로 dependency_override + 인메모리 SQLite를
사용한다. export는 파이프라인을 타지 않으므로 Job/Result를 세션에 직접
삽입해 준비한다.
"""

from __future__ import annotations

import io
from types import SimpleNamespace

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app import main as app_main
from app.core.db import get_db
from app.exporters.excel import RESULT_COLUMN_LABELS
from app.models import Base
from app.models.financial_snapshot import FinancialSnapshot
from app.models.job import Job, JobStatus
from app.models.result import ParseStatus, Result


@pytest.fixture
def client_with_db():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine, autoflush=False, autocommit=False)

    def _override_get_db():
        db = factory()
        try:
            yield db
        finally:
            db.close()

    app_main.app.dependency_overrides[get_db] = _override_get_db

    yield TestClient(app_main.app), factory

    app_main.app.dependency_overrides.clear()


@pytest.fixture
def report_output_to_tmp(tmp_path, monkeypatch):
    """보고서 산출물 폴더만 tmp_path로 돌린다(템플릿은 저장소 실제 파일 그대로).

    tests/test_reports.py의 `report_settings`와 같은 방식 — 이 픽스처가 없으면
    API 테스트가 개발 저장소의 `backend/report/`에 실제 파일을 쌓는다.
    """
    from app import config as app_config
    from app.reports import audit_proposal

    real = app_config.get_settings()
    stub = SimpleNamespace(
        report_output_dir=str(tmp_path / "report"),
        report_template_path=real.report_template_path,
    )
    monkeypatch.setattr(audit_proposal, "get_settings", lambda: stub)
    return stub


def _seed_job_with_results(factory) -> int:
    db = factory()
    try:
        job = Job(
            created_at="2026-07-15T00:00:00",
            name="테스트 Job",
            cond_region='{"sido": "경남", "sigungu": ["김해시"]}',
            cond_revenue="{}",
            cond_industry="[]",
            cond_period='{"bgn_de": "20260101", "end_de": "20260131"}',
            status=JobStatus.DONE,
            current_step=6,
            progress_done=2,
            progress_total=2,
            error_msg=None,
        )
        db.add(job)
        db.commit()
        db.refresh(job)

        db.add_all(
            [
                Result(
                    job_id=job.id,
                    corp_code="00100001",
                    rcept_no="20260601000001",
                    corp_name="㈜성공테스트",
                    address="경상남도 김해시 삼계로 1",
                    phone="055-000-0000",
                    ceo_name="홍길동",
                    induty_code="25",
                    induty_name="금속가공제품 제조업",
                    fiscal_date="20251231",
                    audit_opinion="적정",
                    auditor_name="안경회계법인",
                    auditor_address="경상남도 창원시 중앙대로 1",
                    revenue_cur=10_000_000_000,
                    revenue_prv=9_000_000_000,
                    parse_status=ParseStatus.OK,
                    parse_note=None,
                    excluded_by_revenue=0,
                ),
                Result(
                    job_id=job.id,
                    corp_code="00100002",
                    rcept_no="20260601000002",
                    corp_name="㈜실패테스트",
                    address="경상남도 김해시 분성로 2",
                    phone=None,
                    ceo_name="김철수",
                    induty_code="25",
                    induty_name="금속가공제품 제조업",
                    fiscal_date="20251231",
                    audit_opinion=None,
                    revenue_cur=None,
                    revenue_prv=None,
                    parse_status=ParseStatus.FAILED,
                    parse_note="XML 파싱 실패",
                    excluded_by_revenue=0,
                ),
            ]
        )
        db.commit()
        return job.id
    finally:
        db.close()


def test_list_results_returns_seeded_rows(client_with_db):
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)

    resp = client.get(f"/api/jobs/{job_id}/results")
    assert resp.status_code == 200
    body = resp.json()
    assert body["total"] == 2
    assert len(body["items"]) == 2


def test_list_results_ids_only_returns_all_ids_without_items(client_with_db):
    """`ids_only=true`는 페이징을 무시하고 필터를 통과한 id 전체만 돌려준다.

    화면의 "현재 필터 전체 선택"이 페이지를 순회하지 않고 선택 목록을 만드는 근거다
    (2026-08-03). `page_size=1`을 함께 줘도 잘려서는 안 된다.
    """
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)

    body = client.get(
        f"/api/jobs/{job_id}/results", params={"ids_only": True, "page_size": 1}
    ).json()

    assert body["items"] == []
    assert body["total"] == 2
    assert len(body["ids"]) == 2
    # total은 항상 반환한 ids 개수와 일치하고, page/page_size는 "한 쪽에 전부"를 뜻한다.
    assert body["total"] == len(body["ids"])
    assert (body["page"], body["page_size"]) == (1, 2)


def test_list_results_ids_only_applies_filter_and_sort(client_with_db):
    """`ids_only`도 목록과 똑같은 필터·정렬을 탄다 — 화면 순서 = 선택 순서."""
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)

    all_desc = client.get(
        f"/api/jobs/{job_id}/results",
        params={"ids_only": True, "sort": "corp_name:desc"},
    ).json()
    paged_desc = client.get(
        f"/api/jobs/{job_id}/results", params={"sort": "corp_name:desc"}
    ).json()
    assert all_desc["ids"] == [r["id"] for r in paged_desc["items"]]

    filtered = client.get(
        f"/api/jobs/{job_id}/results", params={"ids_only": True, "q": "성공"}
    ).json()
    assert filtered["total"] == 1
    assert filtered["ids"] == [
        r["id"]
        for r in client.get(f"/api/jobs/{job_id}/results", params={"q": "성공"}).json()["items"]
    ]

    empty = client.get(
        f"/api/jobs/{job_id}/results", params={"ids_only": True, "q": "없는회사"}
    ).json()
    assert empty["ids"] == [] and empty["total"] == 0


def test_list_results_omits_ids_field_without_ids_only(client_with_db):
    """`ids_only`를 주지 않으면 `ids`는 **null**이다(빈 배열이 아님).

    프론트는 이 차이로 "서버가 ids_only를 지원하는지"를 판정한다 — 구버전 백엔드는
    모르는 쿼리 파라미터를 조용히 무시하고 페이지 1쪽만 주기 때문에, `ids`가 null이면
    전체 선택을 실행하면 안 된다(`/export`의 `_selected` 파일명 접미어와 같은 계약).
    """
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)

    body = client.get(f"/api/jobs/{job_id}/results").json()
    assert body["ids"] is None
    assert len(body["items"]) == 2


def test_list_results_sorts_by_column_and_pushes_missing_values_last(client_with_db):
    """매출액 오름차순 정렬 — 값이 없는 행(파싱 실패)은 방향과 무관하게 항상 뒤로."""
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)

    asc = client.get(
        f"/api/jobs/{job_id}/results", params={"sort_by": "revenue_cur", "sort_dir": "asc"}
    ).json()
    assert [r["corp_name"] for r in asc["items"]] == ["㈜성공테스트", "㈜실패테스트"]

    desc = client.get(
        f"/api/jobs/{job_id}/results", params={"sort_by": "revenue_cur", "sort_dir": "desc"}
    ).json()
    # 내림차순이어도 revenue_cur=None인 ㈜실패테스트가 앞으로 오면 안 된다.
    assert [r["corp_name"] for r in desc["items"]] == ["㈜성공테스트", "㈜실패테스트"]


def test_list_results_multi_sort_by_name_then_industry(client_with_db):
    """다중 정렬 — `sort=corp_name:asc,induty_name:asc`면 회사명이 같은 행끼리는
    업종명 오름차순으로 2차 정렬된다(프론트 Shift+클릭 다중 정렬의 서버측 계약)."""
    client, factory = client_with_db
    db = factory()
    try:
        job = Job(
            created_at="2026-07-15T00:00:00",
            name="다중정렬 Job",
            cond_region="{}",
            cond_revenue="{}",
            cond_industry="[]",
            cond_period="{}",
            status=JobStatus.DONE,
            current_step=6,
            progress_done=3,
            progress_total=3,
            error_msg=None,
        )
        db.add(job)
        db.commit()
        db.refresh(job)
        db.add_all(
            [
                Result(job_id=job.id, corp_code="c1", corp_name="가나기업",
                       induty_name="소프트웨어 개발업", parse_status=ParseStatus.OK),
                Result(job_id=job.id, corp_code="c2", corp_name="가나기업",
                       induty_name="금속가공제품 제조업", parse_status=ParseStatus.OK),
                Result(job_id=job.id, corp_code="c3", corp_name="다라기업",
                       induty_name="건설업", parse_status=ParseStatus.OK),
            ]
        )
        db.commit()
        job_id = job.id
    finally:
        db.close()

    body = client.get(
        f"/api/jobs/{job_id}/results",
        params={"sort": "corp_name:asc,induty_name:asc"},
    ).json()
    # 가나기업 두 건은 업종명 오름차순("금속..." < "소프트..."), 그다음 다라기업.
    assert [(r["corp_name"], r["induty_name"]) for r in body["items"]] == [
        ("가나기업", "금속가공제품 제조업"),
        ("가나기업", "소프트웨어 개발업"),
        ("다라기업", "건설업"),
    ]

    # 화이트리스트 밖/형식오류 항목은 무시하고 유효한 기준만 적용한다.
    safe = client.get(
        f"/api/jobs/{job_id}/results",
        params={"sort": "bogus:asc,corp_name:desc"},
    ).json()
    assert [r["corp_name"] for r in safe["items"]][0] == "다라기업"


def test_list_results_rejects_unknown_sort_column(client_with_db):
    """화이트리스트 밖의 컬럼명은 무시하고 기본 정렬로 되돌린다(500이 아니라 200)."""
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)

    resp = client.get(f"/api/jobs/{job_id}/results", params={"sort_by": "id; DROP TABLE results"})
    assert resp.status_code == 200
    assert resp.json()["total"] == 2


def test_list_results_filters_by_keyword_including_auditor(client_with_db):
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)

    by_name = client.get(f"/api/jobs/{job_id}/results", params={"q": "성공"}).json()
    assert [r["corp_name"] for r in by_name["items"]] == ["㈜성공테스트"]

    # 감사인명도 검색 대상이다 — "안경회계법인이 감사한 회사만" 추리는 용도.
    by_auditor = client.get(f"/api/jobs/{job_id}/results", params={"q": "안경회계"}).json()
    assert by_auditor["total"] == 1
    assert by_auditor["items"][0]["auditor_name"] == "안경회계법인"
    assert by_auditor["items"][0]["auditor_address"] == "경상남도 창원시 중앙대로 1"

    assert client.get(f"/api/jobs/{job_id}/results", params={"q": "없는회사"}).json()["total"] == 0


def test_list_results_splits_failed_by_has_disclosure(client_with_db):
    """FAILED 중 "파싱 실패"(rcept_no 있음)와 "감사보고서 없음"(rcept_no 없음)을
    `has_disclosure`로 구분할 수 있어야 한다(2026-07-20 추가)."""
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)

    db = factory()
    try:
        db.add(
            Result(
                job_id=job_id,
                corp_code="00100003",
                rcept_no=None,
                corp_name="㈜공시없음테스트",
                parse_status=ParseStatus.FAILED,
                parse_note="최근 감사보고서 공시를 찾을 수 없음(Phase 1 추정치만 존재)",
                excluded_by_revenue=0,
            )
        )
        db.commit()
    finally:
        db.close()

    all_failed = client.get(f"/api/jobs/{job_id}/results", params={"parse_status": "FAILED"})
    assert all_failed.json()["total"] == 2

    to_review = client.get(
        f"/api/jobs/{job_id}/results",
        params={"parse_status": "FAILED", "has_disclosure": True},
    )
    assert [r["corp_name"] for r in to_review.json()["items"]] == ["㈜실패테스트"]

    no_disclosure = client.get(
        f"/api/jobs/{job_id}/results",
        params={"parse_status": "FAILED", "has_disclosure": False},
    )
    assert [r["corp_name"] for r in no_disclosure.json()["items"]] == ["㈜공시없음테스트"]


def test_list_results_filters_by_excluded_by_stale_disclosure(client_with_db):
    """"최근 1년 이내 DART 공시 없음" 배제 플래그도 다른 excluded_by_*와 동일한
    tri-state 패턴(값을 안 주면 필터 없음/true/false)으로 필터할 수 있어야 한다
    (2026-07-21 추가, 실사례 "주식회사 유진")."""
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)

    db = factory()
    try:
        db.add(
            Result(
                job_id=job_id,
                corp_code="00100004",
                rcept_no="20240101000004",
                corp_name="㈜유진",
                parse_status=ParseStatus.OK,
                excluded_by_revenue=0,
                latest_disclosure_date="20240101",
                excluded_by_stale_disclosure=1,
            )
        )
        db.commit()
    finally:
        db.close()

    # 값을 안 주면(기존 excluded_by_revenue/assets와 동일) 필터하지 않는다.
    unfiltered = client.get(f"/api/jobs/{job_id}/results")
    assert unfiltered.json()["total"] == 3

    stale_only = client.get(
        f"/api/jobs/{job_id}/results", params={"excluded_by_stale_disclosure": True}
    )
    assert [r["corp_name"] for r in stale_only.json()["items"]] == ["㈜유진"]
    assert stale_only.json()["items"][0]["latest_disclosure_date"] == "20240101"

    not_stale = client.get(
        f"/api/jobs/{job_id}/results", params={"excluded_by_stale_disclosure": False}
    )
    assert {r["corp_name"] for r in not_stale.json()["items"]} == {"㈜성공테스트", "㈜실패테스트"}


def test_list_results_filters_by_auditor_changed(client_with_db):
    """연도별 감사인 변동 여부(2026-07-26)도 tri-state 필터다 — true/false/미지정.

    판정 불가(NULL: 감사인 이름을 확보한 연도가 1개 이하, 이 기능 도입 이전 Job)는
    true/false 어느 쪽에도 잡히지 않는다."""
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)

    db = factory()
    try:
        db.add_all(
            [
                Result(
                    job_id=job_id,
                    corp_code="00100005",
                    corp_name="㈜감사인교체",
                    parse_status=ParseStatus.OK,
                    auditor_name="나중회계법인",
                    auditor_changed=1,
                ),
                Result(
                    job_id=job_id,
                    corp_code="00100006",
                    corp_name="㈜감사인유지",
                    parse_status=ParseStatus.OK,
                    auditor_name="계속회계법인",
                    auditor_changed=0,
                ),
            ]
        )
        db.commit()
    finally:
        db.close()

    # 값을 안 주면 필터하지 않는다(기존 excluded_by_* 와 동일).
    assert client.get(f"/api/jobs/{job_id}/results").json()["total"] == 4

    changed = client.get(f"/api/jobs/{job_id}/results", params={"auditor_changed": True}).json()
    assert [r["corp_name"] for r in changed["items"]] == ["㈜감사인교체"]
    assert changed["items"][0]["auditor_changed"] == 1

    unchanged = client.get(
        f"/api/jobs/{job_id}/results", params={"auditor_changed": False}
    ).json()
    assert [r["corp_name"] for r in unchanged["items"]] == ["㈜감사인유지"]

    # 기존(도입 이전) 행은 NULL로 남아 판정 불가로 노출된다.
    seeded = client.get(f"/api/jobs/{job_id}/results", params={"q": "성공"}).json()
    assert seeded["items"][0]["auditor_changed"] is None


def test_list_results_not_found_returns_404(client_with_db):
    client, _factory = client_with_db
    resp = client.get("/api/jobs/9999/results")
    assert resp.status_code == 404


def test_set_result_excluded_toggles_flag(client_with_db):
    """CandidatesView "선택 취소" — phase=CANDIDATES(기본값)에서는 자유롭게 토글 가능."""
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)

    db = factory()
    try:
        result_id = db.execute(select(Result.id).where(Result.job_id == job_id)).scalars().first()
    finally:
        db.close()

    resp = client.patch(f"/api/jobs/{job_id}/results/{result_id}/exclude", json={"excluded": True})
    assert resp.status_code == 200
    assert resp.json()["excluded_manually"] == 1

    resp = client.patch(f"/api/jobs/{job_id}/results/{result_id}/exclude", json={"excluded": False})
    assert resp.status_code == 200
    assert resp.json()["excluded_manually"] == 0


def test_set_result_excluded_rejects_when_phase_financials(client_with_db):
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)

    db = factory()
    try:
        job = db.get(Job, job_id)
        job.phase = "FINANCIALS"
        db.commit()
        result_id = db.execute(select(Result.id).where(Result.job_id == job_id)).scalars().first()
    finally:
        db.close()

    resp = client.patch(f"/api/jobs/{job_id}/results/{result_id}/exclude", json={"excluded": True})
    assert resp.status_code == 400


def test_set_result_excluded_not_found_returns_404(client_with_db):
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)

    resp = client.patch(f"/api/jobs/{job_id}/results/999999/exclude", json={"excluded": True})
    assert resp.status_code == 404

    resp = client.patch("/api/jobs/9999/results/1/exclude", json={"excluded": True})
    assert resp.status_code == 404


def test_export_xlsx_returns_valid_workbook_with_korean_headers(client_with_db):
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)

    resp = client.get(f"/api/jobs/{job_id}/export", params={"format": "xlsx"})
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in resp.headers["content-disposition"]

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["results"]
    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "회사명" in header_row
    assert "매출액(당기)" in header_row
    # 데이터 행 2개(헤더 제외) 확인
    assert ws.max_row == 3


def test_export_csv_has_utf8_bom_and_korean_headers(client_with_db):
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)

    resp = client.get(f"/api/jobs/{job_id}/export", params={"format": "csv"})
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("text/csv")
    assert "utf-8-sig" in resp.headers["content-type"]
    assert "attachment" in resp.headers["content-disposition"]

    # utf-8-sig BOM이 실제로 포함되어 있는지 확인
    assert resp.content.startswith(b"\xef\xbb\xbf")
    text = resp.content.decode("utf-8-sig")
    assert "회사명" in text.splitlines()[0]
    assert "㈜성공테스트" in text


def test_export_invalid_format_returns_400(client_with_db):
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)

    resp = client.get(f"/api/jobs/{job_id}/export", params={"format": "pdf"})
    assert resp.status_code == 400


def test_export_not_found_returns_404(client_with_db):
    client, _factory = client_with_db
    resp = client.get("/api/jobs/9999/export", params={"format": "xlsx"})
    assert resp.status_code == 404


def test_export_filters_by_parse_status(client_with_db):
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)

    resp = client.get(
        f"/api/jobs/{job_id}/export",
        params={"format": "csv", "parse_status": "OK"},
    )
    assert resp.status_code == 200
    text = resp.content.decode("utf-8-sig")
    assert "㈜성공테스트" in text
    assert "㈜실패테스트" not in text


# ---------------------------------------------------------------------------
# GET /api/jobs/{id}/export?ids=&include_history= — 다중 선택 다운로드(§4-11, M9)
# ---------------------------------------------------------------------------


def test_export_ids_selects_only_given_rows_and_ignores_other_filters(client_with_db):
    """`ids`가 오면 parse_status/q/sort 등 다른 필터는 전부 무시하고 그 id만 내보낸다.

    사용자는 화면에서 필터로 찾아 체크한 뒤라, 다운로드 시점에 필터를 다시
    태우면 "체크했는데 파일에 없다"가 된다(§4-11).
    """
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)
    failed_id = _get_result_id(factory, job_id, "00100002")  # parse_status=FAILED 행

    resp = client.get(
        f"/api/jobs/{job_id}/export",
        params={
            "format": "csv",
            "ids": str(failed_id),
            # 아래 필터들은 이 행을 배제하지만 ids가 있으므로 전부 무시돼야 한다.
            "parse_status": "OK",
            "q": "존재하지않는회사명",
            "sort": "revenue_cur:desc",
        },
    )
    assert resp.status_code == 200
    text = resp.content.decode("utf-8-sig")
    assert "㈜실패테스트" in text
    assert "㈜성공테스트" not in text


def test_export_ids_uses_long_account_format(client_with_db):
    """선택 다운로드는 회사 1건이 당기 계정과목 24행으로 풀리는 long 포맷이다
    (2026-07-28, 2026-08-05 세부계정 5항목 추가로 19 → 24). 값이 없는 계정과목도
    금액만 빈 채로 행이 남는다."""
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)
    ok_id = _get_result_id(factory, job_id, "00100001")

    resp = client.get(
        f"/api/jobs/{job_id}/export", params={"format": "xlsx", "ids": str(ok_id)}
    )
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["results"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    assert header == [
        "결과ID",
        "Job ID",
        "고유번호",
        "접수번호",
        "회사명",
        "주소",
        "전화번호(미수집)",
        "대표자명",
        "업종코드",
        "업종명",
        "결산기준일",
        "감사의견",
        "감사인",
        "감사인주소",
        "감사인변동여부",
        "계정과목명",
        "금액",
        "파싱상태",  # 계정과목명/금액보다도 뒤, 맨 마지막 컬럼(2026-07-28 사용자 확정)
    ]
    assert ws.max_row == 25  # 헤더 + 계정과목 24행

    account_col, amount_col = header.index("계정과목명") + 1, header.index("금액") + 1
    amounts = {
        ws.cell(row=r, column=account_col).value: ws.cell(row=r, column=amount_col).value
        for r in range(2, 26)
    }
    assert len(amounts) == 24
    assert amounts["매출액"] == 10_000_000_000
    assert amounts["자산총계"] is None  # 결측도 행은 남고 금액만 빈 값
    assert "매출액(전기)" not in amounts  # 전기 항목은 싣지 않는다
    # 기본정보는 모든 계정과목 행에 반복된다(파싱상태 포함).
    corp_col = header.index("회사명") + 1
    status_col = header.index("파싱상태") + 1
    assert {ws.cell(row=r, column=corp_col).value for r in range(2, 26)} == {"㈜성공테스트"}
    assert {ws.cell(row=r, column=status_col).value for r in range(2, 26)} == {"OK"}


def test_export_ids_csv_uses_long_account_format(client_with_db):
    """CSV(기본정보만) 선택 다운로드도 같은 long 포맷을 쓴다."""
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)
    ok_id = _get_result_id(factory, job_id, "00100001")

    resp = client.get(
        f"/api/jobs/{job_id}/export", params={"format": "csv", "ids": str(ok_id)}
    )
    assert resp.status_code == 200
    lines = resp.content.decode("utf-8-sig").splitlines()
    assert lines[0].endswith("계정과목명,금액,파싱상태")
    assert len(lines) == 25
    assert any(line.endswith("매출액,10000000000,OK") for line in lines[1:])


@pytest.mark.parametrize("format", ["xlsx", "csv"])
def test_export_ids_filename_has_selected_suffix(client_with_db, format):
    """`ids`를 준 선택 다운로드의 파일명에는 반드시 `_selected`가 들어간다.

    프론트엔드(`frontend/src/api/results.ts::exportResults`)가 이 접미어 유무를
    **하드 fail-safe 조건**으로 쓴다 — `ids`를 보냈는데 응답 파일명에 `_selected`가
    없으면 "서버가 구버전이라 ids를 무시하고 전체를 내려줬다"로 보고 저장을 막는다.
    따라서 이 문자열을 바꾸면 프론트의 모든 선택 다운로드가 차단되므로, 반대
    방향(`ids` 없으면 접미어 없음, `test_export_without_new_params_is_unchanged`)과
    함께 양방향으로 잠가 둔다(dart-design-review 2026-07-28).
    """
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)
    ok_id = _get_result_id(factory, job_id, "00100001")

    resp = client.get(
        f"/api/jobs/{job_id}/export", params={"format": format, "ids": str(ok_id)}
    )
    assert resp.status_code == 200
    assert "_selected" in resp.headers["content-disposition"]

    # 빈 선택(0건)도 `ids`를 인식한 응답이므로 접미어가 있어야 한다 —
    # 없으면 프론트가 "구버전 서버"로 오판한다.
    resp = client.get(f"/api/jobs/{job_id}/export", params={"format": format, "ids": ""})
    assert resp.status_code == 200
    assert "_selected" in resp.headers["content-disposition"]


def test_export_ids_with_history_filename_has_both_suffixes(client_with_db):
    """`include_history`가 붙어도 `_selected`가 사라지면 안 된다(프론트 가드 근거)."""
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)
    ok_id = _get_result_id(factory, job_id, "00100001")

    resp = client.get(
        f"/api/jobs/{job_id}/export",
        params={"format": "xlsx", "ids": str(ok_id), "include_history": "true"},
    )
    assert resp.status_code == 200
    disposition = resp.headers["content-disposition"]
    assert "_selected" in disposition and "_with_history" in disposition


def test_export_ids_rejects_result_from_another_job(client_with_db):
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)
    other_job_id = _seed_job_with_results(factory)
    own_id = _get_result_id(factory, job_id, "00100001")
    foreign_id = _get_result_id(factory, other_job_id, "00100001")

    resp = client.get(
        f"/api/jobs/{job_id}/export",
        params={"format": "xlsx", "ids": f"{own_id},{foreign_id}"},
    )
    assert resp.status_code == 400
    assert str(foreign_id) in resp.json()["detail"]


def test_export_ids_rejects_non_integer_token(client_with_db):
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)

    resp = client.get(f"/api/jobs/{job_id}/export", params={"format": "xlsx", "ids": "1,abc"})
    assert resp.status_code == 400


@pytest.mark.parametrize("bad_id", ["1180591620717411303424", str(2**63), "0", "-5"])
def test_export_ids_rejects_out_of_range_integer(client_with_db, bad_id):
    """SQLite INTEGER(int64) 범위를 벗어난 정수는 400이어야 한다 — 그대로
    바인딩하면 `OverflowError`로 500이 났다(dart-qa 2026-07-28 실측)."""
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)

    resp = client.get(f"/api/jobs/{job_id}/export", params={"format": "xlsx", "ids": bad_id})
    assert resp.status_code == 400


def test_export_empty_ids_returns_header_only_file(client_with_db):
    """빈 `ids`(선택 0건)는 에러가 아니라 헤더만 있는 빈 파일이다."""
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)

    resp = client.get(f"/api/jobs/{job_id}/export", params={"format": "xlsx", "ids": ""})
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert wb["results"].max_row == 1

    resp = client.get(
        f"/api/jobs/{job_id}/export",
        params={"format": "xlsx", "ids": "", "include_history": "true"},
    )
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["results", "financial_history"]
    assert wb["results"].max_row == 1
    assert wb["financial_history"].max_row == 1


def test_export_include_history_with_csv_returns_400(client_with_db):
    """csv는 다중 시트를 표현할 수 없으므로 조합 자체를 거부한다(§4-11)."""
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)

    resp = client.get(
        f"/api/jobs/{job_id}/export",
        params={"format": "csv", "include_history": "true"},
    )
    assert resp.status_code == 400


def test_export_include_history_writes_two_sheets_with_joined_corp_name(client_with_db):
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)
    ok_id = _get_result_id(factory, job_id, "00100001")
    failed_id = _get_result_id(factory, job_id, "00100002")

    db = factory()
    try:
        db.add_all(
            [
                FinancialSnapshot(
                    result_id=ok_id,
                    rcept_no="R2",
                    fiscal_year="2025",
                    total_assets=6_000_000_000,
                    revenue=10_000_000_000,
                    auditor_name="안경회계법인",
                    from_current_period=1,
                ),
                FinancialSnapshot(
                    result_id=ok_id,
                    rcept_no="R1",
                    fiscal_year="2024",
                    total_assets=5_000_000_000,
                    revenue=8_000_000_000,
                    auditor_name="이전회계법인",
                    from_current_period=1,
                ),
                # 선택하지 않은 회사의 이력은 파일에 실리면 안 된다.
                FinancialSnapshot(
                    result_id=failed_id, rcept_no="R9", fiscal_year="2025", revenue=1
                ),
            ]
        )
        db.commit()
    finally:
        db.close()

    resp = client.get(
        f"/api/jobs/{job_id}/export",
        params={"format": "xlsx", "ids": str(ok_id), "include_history": "true"},
    )
    assert resp.status_code == 200
    assert "attachment" in resp.headers["content-disposition"]

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["results", "financial_history"]

    ws_results = wb["results"]
    # 기본정보 시트는 long 포맷 — 헤더 + 선택한 1개사 × 계정과목 24행(2026-07-28).
    assert ws_results.max_row == 25
    results_header = [c.value for c in next(ws_results.iter_rows(min_row=1, max_row=1))]
    assert results_header[-3:] == ["계정과목명", "금액", "파싱상태"]
    corp_col = results_header.index("회사명") + 1
    assert ws_results.cell(row=2, column=corp_col).value == "㈜성공테스트"

    ws_history = wb["financial_history"]
    header = [c.value for c in next(ws_history.iter_rows(min_row=1, max_row=1))]
    # 2026-07-29부터 재무이력 시트도 long 포맷(계정과목 세로) — 감사인/파싱상태 없음.
    assert header == ["결과ID", "회사명", "회계연도", "접수번호", "재무제표명", "계정과목", "금액"]
    assert ws_history.max_row == 49  # 헤더 + 2개 연도 × 계정과목 24행
    rows = [
        [ws_history.cell(row=r, column=c + 1).value for c in range(len(header))]
        for r in range(2, 50)
    ]
    name_i, year_i, stmt_i, account_i, amount_i = (
        header.index("회사명"),
        header.index("회계연도"),
        header.index("재무제표명"),
        header.index("계정과목"),
        header.index("금액"),
    )
    # result_id -> fiscal_year 오름차순으로 정렬되고, 회사명은 results에서 조인된다.
    assert [str(r[year_i]) for r in rows] == ["2024"] * 24 + ["2025"] * 24
    assert {r[name_i] for r in rows} == {"㈜성공테스트"}
    # 2026-08-05 세부계정 5항목이 각 표 끝에 붙어 7/8/4 → 9/9/6.
    assert [r[stmt_i] for r in rows[:24]] == (
        ["재무상태표"] * 9 + ["손익계산서"] * 9 + ["현금흐름표"] * 6
    )
    assets = [r[amount_i] for r in rows if r[account_i] == "자산총계"]
    assert assets == [5_000_000_000, 6_000_000_000]


def test_export_include_history_without_ids_covers_filtered_rows(client_with_db):
    """`include_history`는 `ids` 없이(=필터 기준 전체) 써도 동작한다.

    이때 기본정보 시트는 **기존 wide 포맷 그대로**여야 한다 — 포맷을 가르는 기준은
    `ids` 유무 하나뿐이고 `include_history`는 관여하지 않는다(dart-qa 2026-07-28).
    """
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)
    ok_id = _get_result_id(factory, job_id, "00100001")

    db = factory()
    try:
        db.add(
            FinancialSnapshot(
                result_id=ok_id, rcept_no="R1", fiscal_year="2024", revenue=8_000_000_000
            )
        )
        db.commit()
    finally:
        db.close()

    resp = client.get(
        f"/api/jobs/{job_id}/export",
        params={"format": "xlsx", "include_history": "true", "parse_status": "OK"},
    )
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    # `ids`가 없으므로 기본정보 시트는 wide 포맷(OK 필터 통과 1개사 = 1행 + 헤더).
    assert wb["results"].max_row == 2
    header = [c.value for c in next(wb["results"].iter_rows(min_row=1, max_row=1))]
    assert header == list(RESULT_COLUMN_LABELS.values())
    assert "매출액(전기)" in header and "계정과목명" not in header
    assert wb["financial_history"].max_row == 25  # 헤더 + 스냅샷 1건 × 계정과목 24행


def test_export_without_new_params_is_unchanged(client_with_db):
    """`ids`/`include_history`를 주지 않으면 기존 전체 내보내기와 100% 동일하다."""
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)

    resp = client.get(f"/api/jobs/{job_id}/export", params={"format": "xlsx"})
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["results"]
    assert wb["results"].max_row == 3
    # 필터 전체 내보내기는 기존 wide 포맷 그대로다(선택 다운로드 long 포맷 아님).
    header = [c.value for c in next(wb["results"].iter_rows(min_row=1, max_row=1))]
    assert "매출액(당기)" in header and "매출액(전기)" in header
    assert "계정과목명" not in header
    assert "dart_search_job" in resp.headers["content-disposition"]
    assert "_selected" not in resp.headers["content-disposition"]


# ---------------------------------------------------------------------------
# POST /api/jobs/{id}/results/selection-summary — 선택 요약 (§4-12-A, 2026-08-03)
# ---------------------------------------------------------------------------


def _seed_flagged_results(factory, job_id: int) -> dict[str, int]:
    """플래그 조합이 다른 결과 5건을 추가하고 {회사명: id}를 돌려준다.

    `㈜공시없음`은 Phase 2 B1이 감사보고서 공시를 못 찾았을 때 실제로 쓰는 조합
    (`rcept_no=None` + `parse_status=FAILED` + `excluded_by_stale_disclosure=1`,
    `app/core/pipeline.py`)을 그대로 재현한다 — 이 조합을 "검수 필요(FAILED)"로
    세면 안 된다는 것이 selection-summary의 핵심 계약이다.
    """
    db = factory()
    try:
        rows = [
            Result(
                job_id=job_id,
                corp_code="00100010",
                rcept_no="20260601000010",
                corp_name="㈜휴면",
                parse_status=ParseStatus.OK,
                latest_disclosure_date="20240101",
                excluded_by_stale_disclosure=1,
            ),
            Result(
                job_id=job_id,
                corp_code="00100011",
                rcept_no="20260601000011",
                corp_name="㈜매출제외",
                parse_status=ParseStatus.OK,
                excluded_by_revenue=1,
            ),
            Result(
                job_id=job_id,
                corp_code="00100012",
                rcept_no="20260601000012",
                corp_name="㈜둘다제외",
                parse_status=ParseStatus.PARTIAL,
                excluded_by_revenue=1,
                excluded_by_assets=1,
            ),
            Result(
                job_id=job_id,
                corp_code="00100013",
                rcept_no="20260601000013",
                corp_name="㈜정상",
                parse_status=ParseStatus.OK,
            ),
            Result(
                job_id=job_id,
                corp_code="00100014",
                rcept_no=None,
                corp_name="㈜공시없음",
                parse_status=ParseStatus.FAILED,
                parse_note="최근 감사보고서 공시를 찾을 수 없음(참고값만 존재)",
                latest_disclosure_date=None,
                excluded_by_stale_disclosure=1,
            ),
        ]
        db.add_all(rows)
        db.commit()
        return {row.corp_name: row.id for row in rows}
    finally:
        db.close()


def _seed_snapshot_for(
    factory, result_id: int, fiscal_year: str = "2024", complete: bool = False
) -> None:
    """재무 이력 1건을 심는다 — `no_history` 집계가 0건인 회사만 세는지 확인용.

    `complete=True`면 보고서 생성에 필요한 비율 계산 항목까지 모두 채운다
    (그렇지 않으면 이력이 있어도 `select_financial_rows()`가 그 연도를 버려
    생성이 건너뛰어진다 — `no_history`가 어디까지나 **상한**인 이유다).
    """
    values: dict[str, int] = {"revenue": 1_000}
    if complete:
        values = {
            "current_assets": 5_000,
            "noncurrent_assets": 5_000,
            "total_assets": 10_000,
            "current_liab": 2_000,
            "noncurrent_liab": 1_000,
            "total_liab": 3_000,
            "total_equity": 7_000,
            "revenue": 20_000,
            "cogs": 12_000,
            "gross_profit": 8_000,
            "sga": 5_000,
            "operating_income": 3_000,
            "net_income": 2_000,
        }
    db = factory()
    try:
        db.add(
            FinancialSnapshot(
                result_id=result_id,
                rcept_no="20260601000001",
                fiscal_year=fiscal_year,
                parse_status=ParseStatus.OK,
                from_current_period=1,
                **values,
            )
        )
        db.commit()
    finally:
        db.close()


def test_selection_summary_counts_flags_for_selected_ids(client_with_db):
    """확인 모달용 집계 — 선택한 id들만 대상으로 위험 신호 건수를 센다.

    한 회사가 매출액·총자산 두 조건에 동시에 걸릴 수 있어(두 필터는 독립 판정)
    각 건수의 합이 total을 넘을 수 있다 — 화면이 더해서 쓰면 안 되는 근거.
    """
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)  # ㈜성공테스트(OK) + ㈜실패테스트(FAILED)
    ids = _seed_flagged_results(factory, job_id)
    failed_id = _get_result_id(factory, job_id, "00100002")

    selected = [
        ids["㈜휴면"],
        ids["㈜매출제외"],
        ids["㈜둘다제외"],
        ids["㈜정상"],
        ids["㈜공시없음"],
        failed_id,
    ]
    resp = client.post(
        f"/api/jobs/{job_id}/results/selection-summary", json={"ids": selected}
    )
    assert resp.status_code == 200, resp.text
    assert resp.json() == {
        "total": 6,
        # ㈜휴면 + ㈜공시없음 (파이프라인이 공시 미발견 건에도 이 플래그를 세운다)
        "stale_disclosure": 2,
        "excluded_revenue": 2,
        "excluded_assets": 1,
        # ㈜실패테스트만 — ㈜공시없음도 parse_status=FAILED지만 rcept_no가 없어
        # "검수 필요"가 아니다(화면의 두 탭 구분과 동일).
        "failed": 1,
        "no_disclosure": 1,
        # 이 테스트는 스냅샷을 심지 않으므로 전부 재무 이력 0건이다.
        "no_history": 6,
    }


def test_selection_summary_separates_failed_from_missing_disclosure(client_with_db):
    """[dart-qa 2026-08-03] `parse_status=FAILED`만으로 "검수 필요"를 세면 안 된다.

    Phase 2 B1이 공시를 못 찾은 건에 FAILED를 쓰기 때문에, 실측 개발 DB에서는
    FAILED 2,215건 중 2,214건이 `rcept_no IS NULL`(=검수 대상 아님)이었다.
    화면의 "파싱 실패(검수 필요)" 탭(`parse_status=FAILED` + `has_disclosure=true`)
    과 `failed` 필드가 같은 답을 줘야 한다.
    """
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)
    ids = _seed_flagged_results(factory, job_id)
    real_failed_id = _get_result_id(factory, job_id, "00100002")  # rcept_no 있음

    only_missing = client.post(
        f"/api/jobs/{job_id}/results/selection-summary",
        json={"ids": [ids["㈜공시없음"]]},
    ).json()
    assert only_missing["failed"] == 0
    assert only_missing["no_disclosure"] == 1

    only_real = client.post(
        f"/api/jobs/{job_id}/results/selection-summary", json={"ids": [real_failed_id]}
    ).json()
    assert only_real["failed"] == 1
    assert only_real["no_disclosure"] == 0

    # 같은 선택을 목록 조회의 탭 파라미터로 걸러도 답이 일치해야 한다.
    tab = client.get(
        f"/api/jobs/{job_id}/results",
        params={"parse_status": "FAILED", "has_disclosure": True},
    ).json()
    assert [r["id"] for r in tab["items"]] == [real_failed_id]


def test_selection_summary_counts_companies_without_financial_history(
    client_with_db, report_output_to_tmp
):
    """[dart-qa 2026-08-03] `no_history` — 보고서 생성이 건너뛸 회사 수.

    `generate_reports()`는 재무 이력이 0건인 회사를 생성하지 않는다(§4-12).
    확인 모달이 "N건 생성"이라고 물으려면 그 N이 `total`이 아니라
    `total - no_history`(=**최대** 생성 가능 건수)여야 한다(실측 Job 27:
    4,383건 중 3,149건이 스냅샷 0건이라 실제 산출물은 최대 1,234건).

    이력이 **있어도** 전 연도가 파싱 실패/결측이면 추가로 건너뛰므로,
    `total - no_history`는 정확값이 아니라 상한이다 — 아래에서 실제 생성과
    비교해 그 부등식이 성립하는지까지 확인한다.
    """
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)
    ids = _seed_flagged_results(factory, job_id)
    _seed_snapshot_for(factory, ids["㈜정상"], complete=True)
    _seed_snapshot_for(factory, ids["㈜매출제외"], fiscal_year="2023")  # 결측 연도

    selected = [ids["㈜정상"], ids["㈜매출제외"], ids["㈜휴면"], ids["㈜공시없음"]]
    body = client.post(
        f"/api/jobs/{job_id}/results/selection-summary", json={"ids": selected}
    ).json()
    assert body["total"] == 4
    assert body["no_history"] == 2  # ㈜휴면 / ㈜공시없음만 (스냅샷 0건)

    generated = client.post(
        f"/api/jobs/{job_id}/generate-report", json={"ids": selected}
    )
    assert generated.status_code == 200, generated.text
    payload = generated.json()
    # 스냅샷 0건 2건은 반드시 건너뛴다 + 결측 연도뿐인 ㈜매출제외도 건너뛴다.
    assert payload["generated_count"] == 1
    assert {s["result_id"] for s in payload["skipped"]} == {
        ids["㈜휴면"],
        ids["㈜공시없음"],
        ids["㈜매출제외"],
    }
    # 화면이 쓸 "최대 N건 생성" 문구의 계약 — 실제 생성량은 이 상한을 넘지 않는다.
    assert payload["generated_count"] <= body["total"] - body["no_history"]


def test_selection_summary_scopes_to_given_ids_and_dedupes(client_with_db):
    """선택하지 않은 행은 세지 않고, 중복 id는 1건으로만 센다(생성 시와 동일 정규화)."""
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)
    ids = _seed_flagged_results(factory, job_id)

    body = client.post(
        f"/api/jobs/{job_id}/results/selection-summary",
        json={"ids": [ids["㈜휴면"], ids["㈜휴면"]]},
    ).json()
    assert body == {
        "total": 1,
        "stale_disclosure": 1,
        "excluded_revenue": 0,
        "excluded_assets": 0,
        "failed": 0,
        "no_disclosure": 0,
        "no_history": 1,
    }

    clean = client.post(
        f"/api/jobs/{job_id}/results/selection-summary", json={"ids": [ids["㈜정상"]]}
    ).json()
    assert clean == {
        "total": 1,
        "stale_disclosure": 0,
        "excluded_revenue": 0,
        "excluded_assets": 0,
        "failed": 0,
        "no_disclosure": 0,
        "no_history": 1,
    }


def test_selection_summary_shares_id_validation_with_generate_report(client_with_db):
    """입력 검증 계약이 `POST /generate-report`와 동일해야 한다.

    요약이 200인데 생성이 400이면 확인 모달이 거짓 안내를 하게 된다 — 타 Job id
    400(같은 메시지), 범위 밖 정수 400, 빈 목록 422, 없는 Job 404까지 같다.
    """
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)
    other_job_id = _seed_job_with_results(factory)
    own_id = _get_result_id(factory, job_id, "00100001")
    foreign_id = _get_result_id(factory, other_job_id, "00100001")

    foreign = client.post(
        f"/api/jobs/{job_id}/results/selection-summary",
        json={"ids": [own_id, foreign_id]},
    )
    assert foreign.status_code == 400
    assert str(foreign_id) in foreign.json()["detail"]
    # 생성 엔드포인트도 같은 메시지로 거부한다(두 경로가 한 헬퍼를 공유한다).
    generate = client.post(
        f"/api/jobs/{job_id}/generate-report", json={"ids": [own_id, foreign_id]}
    )
    assert generate.status_code == 400
    assert generate.json()["detail"] == foreign.json()["detail"]

    overflow = client.post(
        f"/api/jobs/{job_id}/results/selection-summary", json={"ids": [2**63]}
    )
    assert overflow.status_code == 400

    empty = client.post(f"/api/jobs/{job_id}/results/selection-summary", json={"ids": []})
    assert empty.status_code == 422  # pydantic min_length=1

    missing_job = client.post(
        "/api/jobs/999999/results/selection-summary", json={"ids": [own_id]}
    )
    assert missing_job.status_code == 404


# ---------------------------------------------------------------------------
# GET /api/jobs/{id}/results/{result_id}/history — STEP 7(2026-07-15 추가)
# ---------------------------------------------------------------------------


def _get_result_id(factory, job_id: int, corp_code: str) -> int:
    db = factory()
    try:
        result = db.execute(
            select(Result).where(Result.job_id == job_id, Result.corp_code == corp_code)
        ).scalar_one()
        return result.id
    finally:
        db.close()


def test_get_result_history_returns_oldest_first(client_with_db):
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)
    result_id = _get_result_id(factory, job_id, "00100001")

    db = factory()
    try:
        db.add_all(
            [
                FinancialSnapshot(result_id=result_id, rcept_no="R2", fiscal_year="2025", revenue=10_000),
                FinancialSnapshot(result_id=result_id, rcept_no="R1", fiscal_year="2023", revenue=8_000),
                FinancialSnapshot(result_id=result_id, rcept_no="R1", fiscal_year="2024", revenue=9_000),
            ]
        )
        db.commit()
    finally:
        db.close()

    resp = client.get(f"/api/jobs/{job_id}/results/{result_id}/history")
    assert resp.status_code == 200
    body = resp.json()
    assert [row["fiscal_year"] for row in body] == ["2023", "2024", "2025"]  # 오래된 -> 최신 순
    assert body[0]["revenue"] == 8_000


def test_get_result_history_includes_per_year_auditor_name(client_with_db):
    """연도별 감사인 이름(2026-07-26)이 이력 응답에 실린다 — 도입 이전에 수집된
    행은 null이라 화면이 "판정 불가"로 다룰 수 있어야 한다."""
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)
    result_id = _get_result_id(factory, job_id, "00100001")

    db = factory()
    try:
        db.add_all(
            [
                FinancialSnapshot(
                    result_id=result_id,
                    rcept_no="R2",
                    fiscal_year="2025",
                    from_current_period=1,
                    auditor_name="나중회계법인",
                ),
                FinancialSnapshot(
                    result_id=result_id,
                    rcept_no="R1",
                    fiscal_year="2024",
                    from_current_period=1,
                    auditor_name="이전회계법인",
                ),
                # 전기 열 유래(자기 공시를 못 연 연도)는 감사인을 채우지 않는다.
                FinancialSnapshot(
                    result_id=result_id, rcept_no="R1", fiscal_year="2023", from_current_period=0
                ),
            ]
        )
        db.commit()
    finally:
        db.close()

    body = client.get(f"/api/jobs/{job_id}/results/{result_id}/history").json()
    assert [row["auditor_name"] for row in body] == [None, "이전회계법인", "나중회계법인"]
    # 이름이 없는 2023년은 판정 불가, 2024년은 비교할 이전 연도가 없어 판정 불가,
    # 2025년만 "직전 연도 대비 변경"으로 판정된다.
    assert [row["auditor_changed_from_prev"] for row in body] == [None, None, True]


def test_get_result_history_auditor_change_flag_uses_backend_normalization(client_with_db):
    """연도별 "직전 대비 변경" 판정은 **서버**가 `_auditor_key()`로 계산한다
    (2026-07-26) — 프론트가 공백 제거만 하는 자체 비교를 두면 목록 컬럼
    (`auditor_changed`)과 상세 뱃지의 답이 갈린다(dart-qa 실측 3건)."""
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)
    result_id = _get_result_id(factory, job_id, "00100001")

    db = factory()
    try:
        # 셋 다 같은 감사인이다: 표기 순서 차이 + 서명자(대표이사 교체) 차이뿐.
        db.add_all(
            [
                FinancialSnapshot(
                    result_id=result_id, rcept_no="R1", fiscal_year="2023",
                    from_current_period=1, auditor_name="회계법인 원지",
                ),
                FinancialSnapshot(
                    result_id=result_id, rcept_no="R2", fiscal_year="2024",
                    from_current_period=1, auditor_name="원지회계법인",
                ),
                FinancialSnapshot(
                    result_id=result_id, rcept_no="R3", fiscal_year="2025",
                    from_current_period=1, auditor_name="원지회계법인 대표이사김철수",
                ),
            ]
        )
        db.commit()
    finally:
        db.close()

    body = client.get(f"/api/jobs/{job_id}/results/{result_id}/history").json()
    assert [row["auditor_changed_from_prev"] for row in body] == [None, False, False]


def test_get_result_history_empty_when_no_snapshots(client_with_db):
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)
    result_id = _get_result_id(factory, job_id, "00100001")

    resp = client.get(f"/api/jobs/{job_id}/results/{result_id}/history")
    assert resp.status_code == 200
    assert resp.json() == []


def test_get_result_history_job_not_found_returns_404(client_with_db):
    client, _factory = client_with_db
    resp = client.get("/api/jobs/9999/results/1/history")
    assert resp.status_code == 404


def test_get_result_history_result_not_found_returns_404(client_with_db):
    client, factory = client_with_db
    job_id = _seed_job_with_results(factory)

    resp = client.get(f"/api/jobs/{job_id}/results/9999/history")
    assert resp.status_code == 404


def test_get_result_history_rejects_result_from_other_job(client_with_db):
    """result_id는 존재하지만 다른 job에 속하면 404 (job_id-result_id 불일치)."""
    client, factory = client_with_db
    job_id_1 = _seed_job_with_results(factory)
    job_id_2 = _seed_job_with_results(factory)
    result_id_in_job1 = _get_result_id(factory, job_id_1, "00100001")

    resp = client.get(f"/api/jobs/{job_id_2}/results/{result_id_in_job1}/history")
    assert resp.status_code == 404


# ---------------------------------------------------------------------------
# §4-8 원문 섹션 열람 API (document-sections)
# ---------------------------------------------------------------------------

import shutil  # noqa: E402
from pathlib import Path  # noqa: E402
from types import SimpleNamespace  # noqa: E402

_FIXTURES_DIR = Path(__file__).parent / "fixtures"


def _seed_result_with_rcept(factory, rcept_no: str) -> tuple[int, int]:
    """rcept_no를 가진 결과 1건을 seed하고 (job_id, result_id)를 반환."""
    db = factory()
    try:
        job = Job(
            created_at="2026-07-19T00:00:00",
            name="원문열람 테스트",
            cond_region="{}",
            cond_revenue="{}",
            cond_industry="[]",
            cond_period="{}",
            status=JobStatus.DONE,
        )
        db.add(job)
        db.commit()
        db.refresh(job)
        result = Result(job_id=job.id, corp_code="00100001", rcept_no=rcept_no, corp_name="㈜원문")
        db.add(result)
        db.commit()
        db.refresh(result)
        return job.id, result.id
    finally:
        db.close()


def _point_cache_at_tmp(monkeypatch, tmp_path, rcept_no: str, fixture_id: str) -> None:
    """DOCUMENT_CACHE_DIR을 tmp로 돌리고 fixture XML을 {tmp}/{rcept_no}/에 복사."""
    target = tmp_path / rcept_no
    target.mkdir(parents=True, exist_ok=True)
    shutil.copy(_FIXTURES_DIR / fixture_id / f"{fixture_id}_00760.xml", target / "document.xml")
    monkeypatch.setattr(
        "app.api.results.get_settings",
        lambda: SimpleNamespace(document_cache_dir=str(tmp_path)),
    )


def test_document_section_returns_assembled_html(client_with_db, monkeypatch, tmp_path):
    client, factory = client_with_db
    job_id, result_id = _seed_result_with_rcept(factory, "20260630000641")
    _point_cache_at_tmp(monkeypatch, tmp_path, "20260630000641", "20260630000641")

    resp = client.get(f"/api/jobs/{job_id}/results/{result_id}/document-sections/cf")
    assert resp.status_code == 200
    body = resp.json()
    assert body["available"] is True
    assert body["notice"] is None
    assert "<table>" in body["html"]
    assert "현" in body["html"]  # 현금흐름표 제목/내용


@pytest.mark.parametrize(
    "fixture_id, expected_phrase",
    [
        # 신서식 — <TITLE>독립된 감사인의 감사보고서</TITLE>, 적정의견
        ("20260630000641", "공정하게"),
        # 2012년 구서식 — <TITLE>외부감사인의 감사보고서</TITLE>, "적정하게" 문구
        ("20120110000138", "적정하게"),
    ],
)
def test_document_section_audit_covers_both_report_title_formats(
    client_with_db, monkeypatch, tmp_path, fixture_id, expected_phrase
):
    """감사의견 탭(section=audit)은 신서식("독립된 감사인의...")과 2012년
    구서식("외부감사인의...")을 공통 부분문자열 "감사보고서"로 모두 잡는다."""
    client, factory = client_with_db
    job_id, result_id = _seed_result_with_rcept(factory, fixture_id)
    _point_cache_at_tmp(monkeypatch, tmp_path, fixture_id, fixture_id)

    resp = client.get(f"/api/jobs/{job_id}/results/{result_id}/document-sections/audit")
    assert resp.status_code == 200
    body = resp.json()
    assert body["available"] is True
    assert "감사보고서" in body["html"]
    assert expected_phrase in body["html"]


def test_document_section_renders_te_data_cells(client_with_db, monkeypatch, tmp_path):
    """재무제표 데이터 셀은 TD가 아니라 <TE> 태그다 — 이를 셀로 처리하지 않으면
    계정과목/금액이 전부 빈 <tr></tr>로 렌더된다(§4-8 회귀). 실제 금액 값과
    계정과목이 HTML에 담기는지, 빈 행이 없는지 검증한다."""
    import re

    client, factory = client_with_db
    job_id, result_id = _seed_result_with_rcept(factory, "20260630000641")
    _point_cache_at_tmp(monkeypatch, tmp_path, "20260630000641", "20260630000641")

    resp = client.get(f"/api/jobs/{job_id}/results/{result_id}/document-sections/bs")
    assert resp.status_code == 200
    html = resp.json()["html"]
    # 금액 셀(1,234,567 형태)이 실제로 담겨 있어야 한다.
    assert len(re.findall(r"[0-9]{1,3}(?:,[0-9]{3})+", html)) > 10
    # 데이터 행이 빈 <tr></tr>로 렌더되면 안 된다.
    assert "<tr></tr>" not in html
    assert "자산총계" in html.replace(" ", "")


def test_account_detail_returns_children_per_summary_field(client_with_db, monkeypatch, tmp_path):
    """요약 대분류(유동자산 등)별 세부계정이 계층/값과 함께 반환되는지 검증."""
    client, factory = client_with_db
    job_id, result_id = _seed_result_with_rcept(factory, "20260630000641")
    _point_cache_at_tmp(monkeypatch, tmp_path, "20260630000641", "20260630000641")

    resp = client.get(f"/api/jobs/{job_id}/results/{result_id}/account-detail")
    assert resp.status_code == 200
    body = resp.json()
    assert body["rcept_no"] == "20260630000641"
    assert body["fiscal_year_cur"] == "2026"  # 이력 표의 당기/전기 열 판정 근거

    rows = body["accounts"]["current_assets"]
    assert len(rows) > 5
    # 세부계정은 라벨/레벨/당기·전기 값을 갖는다.
    assert all(row["level"] >= 1 for row in rows)
    assert any(row["cur"] is not None for row in rows)
    # 총계 항목은 하위가 형제 대분류라 children이 비어 있다(토글 비활성 대상).
    assert body["accounts"]["total_assets"] == []


def test_account_detail_returns_auditor_of_that_document(client_with_db, monkeypatch, tmp_path):
    """감사의견 **다음 행**에 표시할 감사인 정보(2026-07-26)를 그 연도 원문에서
    직접 뽑아 함께 내려준다 — 로컬 캐시만 읽으므로 추가 API 호출이 0건이고,
    이 기능 도입 이전에 수집된 기존 Job의 원문에서도 값이 나온다.

    이 fixture는 서명란이 없는 서식이라 이름만 확보되고 주소는 null이 정상이다
    (test_auditor.py의 동일 rcept 케이스와 같은 기대값)."""
    client, factory = client_with_db
    job_id, result_id = _seed_result_with_rcept(factory, "20260630000641")
    _point_cache_at_tmp(monkeypatch, tmp_path, "20260630000641", "20260630000641")

    body = client.get(f"/api/jobs/{job_id}/results/{result_id}/account-detail").json()
    assert body["auditor_name"] == "정진세림회계법인"
    assert body["auditor_address"] is None


def test_account_detail_rejects_foreign_rcept_no(client_with_db, monkeypatch, tmp_path):
    client, factory = client_with_db
    job_id, result_id = _seed_result_with_rcept(factory, "20260630000641")
    _point_cache_at_tmp(monkeypatch, tmp_path, "20260630000641", "20260630000641")

    resp = client.get(
        f"/api/jobs/{job_id}/results/{result_id}/account-detail?rcept_no=19990101000001"
    )
    assert resp.status_code == 404


def test_account_detail_returns_cash_flow_children_and_audit_opinion(
    client_with_db, monkeypatch, tmp_path
):
    """현금흐름표 3항목(영업/투자/재무활동)도 재무상태표·손익계산서와 동일하게
    세부계정이 반환되고("기말의현금"은 총계라 children이 비어 있는 게 정상), 감사의견도
    함께 내려간다(재무상태표 위 안내 행에 쓴다)."""
    client, factory = client_with_db
    job_id, result_id = _seed_result_with_rcept(factory, "20260630000641")
    _point_cache_at_tmp(monkeypatch, tmp_path, "20260630000641", "20260630000641")

    resp = client.get(f"/api/jobs/{job_id}/results/{result_id}/account-detail")
    assert resp.status_code == 200
    body = resp.json()

    assert body["audit_opinion"] == "적정"

    operating_rows = body["accounts"]["cf_operating"]
    assert len(operating_rows) > 3
    assert all(row["level"] >= 1 for row in operating_rows)
    assert any(row["cur"] is not None for row in operating_rows)

    investing_rows = body["accounts"]["cf_investing"]
    assert len(investing_rows) > 3

    financing_rows = body["accounts"]["cf_financing"]
    assert len(financing_rows) > 3

    # 기말의현금은 그 자체가 총계라 하위 대분류가 없다(자산총계 등과 동일 패턴).
    assert body["accounts"].get("cf_ending_cash", []) == []


def test_account_detail_returns_non_operating_children(client_with_db, monkeypatch, tmp_path):
    """손익계산서 세부계정 펼치기에서 영업외수익/영업외비용 대분류(L0)와 그 하위
    세부계정(이자수익/이자비용/외환차익 등)이 유실 없이 반환되는지 검증(2026-07-22).

    이 두 항목은 표준 13항목에 없어 이전에는 _collect_table이 L0에서 current_field를
    닫아 항목 자체와 children이 통째로 스킵됐다 — valid_fields 확장으로 복구된다.
    """
    client, factory = client_with_db
    job_id, result_id = _seed_result_with_rcept(factory, "20260630000641")
    _point_cache_at_tmp(monkeypatch, tmp_path, "20260630000641", "20260630000641")

    resp = client.get(f"/api/jobs/{job_id}/results/{result_id}/account-detail")
    assert resp.status_code == 200
    accounts = resp.json()["accounts"]

    income_rows = accounts["non_operating_income"]
    assert len(income_rows) > 3
    assert all(row["level"] >= 1 for row in income_rows)
    assert any(row["cur"] is not None for row in income_rows)

    expense_rows = accounts["non_operating_expense"]
    assert len(expense_rows) > 3
    assert any(row["cur"] is not None for row in expense_rows)


def test_document_section_and_account_detail_support_ifrs_attach_format(
    client_with_db, monkeypatch, tmp_path
):
    """IFRS "(첨부)재무제표" 서식 원문(이래CS rcept 20260401004343)도 원문 열람과
    세부계정 펼치기가 동작해야 한다(2026-07-27 사용자 실측 신고 재현).

    이 서식은 재무제표별 `<TITLE>`이 없고 데이터 표가 ACLASS="NORMAL"이라, 두
    UI 전용 엔드포인트가 각자 <TABLE-GROUP><TITLE>+FINANCE 조합만 인식하던 동안
    `document-sections/bs`는 available=false("원문을 찾을 수 없습니다"),
    `account-detail`은 accounts={}로 비어 있었다 — 정작 파이프라인 파서는 같은
    원문을 2026-07-22부터 정상 파싱해 재무이력 표에는 값이 보이던 상태였다.
    """
    client, factory = client_with_db
    job_id, result_id = _seed_result_with_rcept(factory, "20260401004343")
    _point_cache_at_tmp(monkeypatch, tmp_path, "20260401004343", "20260401004343")

    resp = client.get(f"/api/jobs/{job_id}/results/{result_id}/document-sections/bs")
    assert resp.status_code == 200
    body = resp.json()
    assert body["available"] is True
    assert body["notice"] is None
    assert "재 무 상 태 표" in body["html"]
    assert "자산총계" in body["html"].replace(" ", "")

    resp = client.get(f"/api/jobs/{job_id}/results/{result_id}/account-detail")
    assert resp.status_code == 200
    body = resp.json()
    assert body["fiscal_year_cur"] == "2025"
    rows = body["accounts"]["current_assets"]
    assert len(rows) == 8
    assert all(row["level"] >= 1 for row in rows)
    # 세부계정 합계 == 요약 대분류(유동자산) 값 — 계층 판정 자체검증.
    assert sum(row["cur"] for row in rows) == 141_442_144_183
    # 감사의견/감사인은 서식과 무관한 별도 로직이라 기존대로 함께 내려간다.
    assert body["audit_opinion"] == "적정"
    assert body["auditor_name"]


def test_document_section_invalid_section_returns_400(client_with_db, monkeypatch, tmp_path):
    client, factory = client_with_db
    job_id, result_id = _seed_result_with_rcept(factory, "20260630000641")
    _point_cache_at_tmp(monkeypatch, tmp_path, "20260630000641", "20260630000641")

    resp = client.get(f"/api/jobs/{job_id}/results/{result_id}/document-sections/xxx")
    assert resp.status_code == 400


def test_document_section_cache_missing_returns_404(client_with_db, monkeypatch, tmp_path):
    client, factory = client_with_db
    job_id, result_id = _seed_result_with_rcept(factory, "20260630000641")
    # 캐시 디렉터리를 비운 채(파일 복사 없이) 조회 → 404
    monkeypatch.setattr(
        "app.api.results.get_settings",
        lambda: SimpleNamespace(document_cache_dir=str(tmp_path)),
    )
    resp = client.get(f"/api/jobs/{job_id}/results/{result_id}/document-sections/cf")
    assert resp.status_code == 404


def test_document_section_absent_section_returns_notice(client_with_db, monkeypatch, tmp_path):
    """재무제표 미첨부(의견거절 계열) 원문의 cf는 에러가 아니라 available=false + 안내."""
    client, factory = client_with_db
    job_id, result_id = _seed_result_with_rcept(factory, "20260630001111")
    _point_cache_at_tmp(monkeypatch, tmp_path, "20260630001111", "20260630001111")

    resp = client.get(f"/api/jobs/{job_id}/results/{result_id}/document-sections/cf")
    assert resp.status_code == 200
    body = resp.json()
    assert body["available"] is False
    assert body["notice"]  # 안내 문구 존재
    assert body["html"] == ""


def test_document_section_rejects_foreign_rcept_no(client_with_db, monkeypatch, tmp_path):
    """?rcept_no=가 이 결과에 속하지 않으면 404 (history 공시가 아닌 임의 값 거부)."""
    client, factory = client_with_db
    job_id, result_id = _seed_result_with_rcept(factory, "20260630000641")
    _point_cache_at_tmp(monkeypatch, tmp_path, "20260630000641", "20260630000641")

    resp = client.get(
        f"/api/jobs/{job_id}/results/{result_id}/document-sections/cf",
        params={"rcept_no": "99999999999999"},
    )
    assert resp.status_code == 404


def test_document_section_allows_history_rcept_no(client_with_db, monkeypatch, tmp_path):
    """?rcept_no=가 이 결과의 financial_snapshots 공시면 허용된다."""
    client, factory = client_with_db
    job_id, result_id = _seed_result_with_rcept(factory, "20260630000641")
    # 이력 공시로 다른 rcept_no를 등록하고 그 원문을 캐시에 둔다.
    db = factory()
    try:
        db.add(
            FinancialSnapshot(
                result_id=result_id, rcept_no="20260630000665", fiscal_year="2024", revenue=1
            )
        )
        db.commit()
    finally:
        db.close()
    tgt = tmp_path / "20260630000665"
    tgt.mkdir(parents=True)
    shutil.copy(
        _FIXTURES_DIR / "20260630000665" / "20260630000665_00760.xml", tgt / "document.xml"
    )
    monkeypatch.setattr(
        "app.api.results.get_settings",
        lambda: SimpleNamespace(document_cache_dir=str(tmp_path)),
    )

    resp = client.get(
        f"/api/jobs/{job_id}/results/{result_id}/document-sections/cf",
        params={"rcept_no": "20260630000665"},
    )
    assert resp.status_code == 200
    assert resp.json()["available"] is True
