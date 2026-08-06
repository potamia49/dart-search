"""app/exporters/excel.py 단위 테스트 — DB/HTTP 없이 export_results() 자체를 검증."""

from __future__ import annotations

import io

import openpyxl
import pytest

from app.exporters.excel import (
    FINANCIAL_SNAPSHOT_ACCOUNTS_BY_STATEMENT,
    FINANCIAL_SNAPSHOT_ACCOUNT_COLUMNS,
    FINANCIAL_SNAPSHOT_ACCOUNT_LABELS,
    FINANCIAL_SNAPSHOT_COLUMN_LABELS,
    FINANCIAL_SNAPSHOT_STATEMENT_BY_ACCOUNT,
    RESULT_COLUMN_LABELS,
    SELECTION_ACCOUNT_COLUMNS,
    SELECTION_ACCOUNT_LABELS,
    SELECTION_EXPORT_COLUMN_LABELS,
    export_results,
    export_results_with_history,
    export_selection_results,
    results_to_dataframe,
    results_to_selection_dataframe,
    snapshots_to_dataframe,
)
from app.exporters.excel import SelectionAccountDetail, _write_xlsx
from app.models.financial_snapshot import FinancialSnapshot
from app.models.result import ParseStatus, Result
from app.parsers.account_detail import _BS_IS_VALID_FIELDS, AccountRow
from app.parsers.base import (
    CF_FINANCIAL_FIELDS,
    DETAIL_FINANCIAL_FIELDS,
    DIRECT_FINANCIAL_FIELDS,
    NON_OPERATING_FINANCIAL_FIELDS,
)


def _sample_result() -> Result:
    return Result(
        id=1,
        job_id=1,
        corp_code="00100001",
        rcept_no="20260601000001",
        corp_name="㈜테스트",
        address="경상남도 김해시 삼계로 1",
        phone="055-000-0000",
        ceo_name="홍길동",
        induty_code="25",
        induty_name="금속가공제품 제조업",
        fiscal_date="20251231",
        audit_opinion="적정",
        revenue_cur=10_000_000_000,
        revenue_prv=9_000_000_000,
        parse_status=ParseStatus.OK,
        parse_note=None,
        excluded_by_revenue=0,
    )


def test_results_to_dataframe_keeps_db_field_names():
    df = results_to_dataframe([_sample_result()])
    assert list(df.columns) == list(RESULT_COLUMN_LABELS.keys())
    assert df.loc[0, "corp_name"] == "㈜테스트"
    assert df.loc[0, "revenue_cur"] == 10_000_000_000


def test_export_results_xlsx_uses_korean_headers():
    content = export_results([_sample_result()], "xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["results"]
    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header_row == list(RESULT_COLUMN_LABELS.values())


def test_export_results_csv_has_bom_and_korean_headers():
    content = export_results([_sample_result()], "csv")
    assert content.startswith(b"\xef\xbb\xbf")
    text = content.decode("utf-8-sig")
    first_line = text.splitlines()[0]
    assert "회사명" in first_line
    assert "corp_name" not in first_line


def test_export_results_empty_list_still_returns_header_only_file():
    content = export_results([], "xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["results"]
    assert ws.max_row == 1  # 헤더만


def test_export_results_invalid_format_raises():
    with pytest.raises(ValueError):
        export_results([_sample_result()], "pdf")  # type: ignore[arg-type]


# ---------------------------------------------------------------------------
# 선택 다운로드 long 포맷(§4-11, 2026-07-28) — results_to_selection_dataframe /
# export_selection_results
# ---------------------------------------------------------------------------


def test_selection_account_columns_cover_every_cur_field():
    """선택 다운로드 계정과목 목록이 wide 포맷의 `_cur` 필드 전체와 정확히 일치해야 한다.

    `SELECTION_ACCOUNT_COLUMNS`는 하드코딩 목록이라, 앞으로 누군가 새 `_cur`/`_prv`
    필드 쌍을 추가하면(CF 4항목·영업외손익 2항목이 그렇게 추가됐다) wide 전체
    내보내기에는 나오지만 선택 다운로드에서는 **에러 없이 조용히** 빠진다 — 그
    드리프트를 여기서 잡는다(dart-qa 2026-07-28 지적).
    """
    assert {c for c in RESULT_COLUMN_LABELS if c.endswith("_cur")} == set(
        SELECTION_ACCOUNT_COLUMNS
    )
    # 계정과목명은 wide 라벨에서 "(당기)"를 떼어 파생시킨다 — 접미어가 없는 라벨이
    # 섞이면 `removesuffix`가 조용히 통과해 "매출액(당기)" 같은 이름이 그대로 나간다.
    for col in SELECTION_ACCOUNT_COLUMNS:
        assert RESULT_COLUMN_LABELS[col].endswith("(당기)")
        assert "(당기)" not in SELECTION_ACCOUNT_LABELS[col]


def test_selection_dataframe_melts_one_company_into_account_rows():
    """회사 1건이 당기 계정과목 24행으로 풀리고 기본정보는 각 행에 반복된다."""
    df = results_to_selection_dataframe([_sample_result()])

    assert list(df.columns) == list(SELECTION_EXPORT_COLUMN_LABELS.keys())
    assert len(df) == len(SELECTION_ACCOUNT_COLUMNS) == 24  # 2026-08-05 세부계정 5항목 추가
    # 기본정보는 모든 행에 동일하게 반복
    assert set(df["corp_name"]) == {"㈜테스트"}
    assert set(df["corp_code"]) == {"00100001"}
    # 계정과목명은 RESULT_COLUMN_LABELS에서 "(당기)"만 뗀 이름
    assert df.loc[0, "account_name"] == "유동자산"
    assert list(df["account_name"])[:3] == ["유동자산", "비유동자산", "자산총계"]
    revenue_row = df[df["account_name"] == "매출액"].iloc[0]
    assert revenue_row["amount"] == 10_000_000_000


def test_selection_dataframe_keeps_rows_for_missing_amounts():
    """값이 None인 계정과목도 행은 남기고 금액만 비운다(어떤 항목이 결측인지 보여야 함)."""
    df = results_to_selection_dataframe([_sample_result()])

    assets_row = df[df["account_name"] == "자산총계"].iloc[0]
    assert assets_row["amount"] is None  # NaN(float)이 아니라 빈 값이어야 한다
    assert len(df[df["amount"].isna()]) == 23  # 매출액 1건만 값이 있다
    # 금액 컬럼이 float64로 승격되면 10000000000.0처럼 소수점이 붙는다.
    assert df["amount"].dtype == object


def test_selection_dataframe_excludes_previous_period_columns():
    """전기(_prv)는 싣지 않는다(사용자 확정 — 전기는 전년도 당기와 같다)."""
    df = results_to_selection_dataframe([_sample_result()])
    assert not any(c.endswith("_prv") for c in df.columns)
    assert not any("(전기)" in str(v) for v in df["account_name"])
    assert 9_000_000_000 not in list(df["amount"])  # revenue_prv


def test_selection_dataframe_multiple_companies_keep_input_order():
    first = _sample_result()
    second = _sample_result()
    second.id = 2
    second.corp_name = "㈜둘째"
    df = results_to_selection_dataframe([first, second])

    assert len(df) == 48
    assert list(df["corp_name"])[:24] == ["㈜테스트"] * 24
    assert list(df["corp_name"])[24:] == ["㈜둘째"] * 24


def test_export_selection_results_xlsx_uses_new_headers():
    content = export_selection_results([_sample_result()], "xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["results"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    assert header == list(SELECTION_EXPORT_COLUMN_LABELS.values())
    assert header[:5] == ["결과ID", "Job ID", "고유번호", "접수번호", "회사명"]
    # 파싱상태는 계정과목명/금액보다도 뒤(2026-07-28), 세부계정비고는 그보다도 뒤다
    # (§4-15, 2026-08-06 — 회사 단위 상태 컬럼 둘을 맨 뒤에 나란히 둔다).
    assert header[-6:] == [
        "상위계정",
        "계정과목명",
        "계정단계",
        "금액",
        "파싱상태",
        "세부계정비고",
    ]
    assert ws.cell(row=2, column=header.index("파싱상태") + 1).value == "OK"
    assert ws.max_row == 25  # 헤더 + 계정과목 24행
    # 결측 금액은 빈 셀
    amount_col = header.index("금액") + 1
    account_col = header.index("계정과목명") + 1
    amounts = {
        ws.cell(row=r, column=account_col).value: ws.cell(row=r, column=amount_col).value
        for r in range(2, 26)
    }
    assert amounts["매출액"] == 10_000_000_000
    assert amounts["자산총계"] is None


def test_export_selection_results_csv_has_bom_and_blank_amount():
    content = export_selection_results([_sample_result()], "csv")
    assert content.startswith(b"\xef\xbb\xbf")
    lines = content.decode("utf-8-sig").splitlines()

    assert lines[0].endswith("계정과목명,계정단계,금액,파싱상태,세부계정비고")
    assert len(lines) == 25
    assert "10000000000.0" not in content.decode("utf-8-sig")
    # 세부계정을 넘기지 않았으므로 상위계정/세부계정비고는 빈 값, 계정단계는 0(요약).
    assert any(line.endswith(",요약,,매출액,0,10000000000,OK,") for line in lines[1:])
    assert any(line.endswith(",요약,,자산총계,0,,OK,") for line in lines[1:])  # 결측은 빈 값


def test_export_selection_results_empty_list_returns_header_only():
    content = export_selection_results([], "xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(content))
    assert wb["results"].max_row == 1


def test_export_selection_results_invalid_format_raises():
    with pytest.raises(ValueError):
        export_selection_results([_sample_result()], "pdf")  # type: ignore[arg-type]


# ---------------------------------------------------------------------------
# 원문 세부계정(§4-15, 2026-08-06) — 선택 다운로드 ① 시트에 원문 계정과목 전체
# ---------------------------------------------------------------------------


def _sample_detail() -> SelectionAccountDetail:
    """원문에서 뽑은 세부계정 — 라벨/레벨은 회사마다 다른 원문 그대로다."""
    return SelectionAccountDetail(
        accounts={
            "current_assets": [
                AccountRow(label="(1) 당좌자산", level=1, cur=3_000_000_000.0, prv=1.0),
                AccountRow(label="1.현금및현금등가물", level=2, cur=1_200_000_000.0, prv=2.0),
                AccountRow(label="2.매출채권", level=2, cur=None, prv=None),
            ],
            "revenue": [
                AccountRow(label="가. 제품매출", level=1, cur=7_000_000_000.0, prv=3.0),
            ],
        }
    )


def test_selection_dataframe_appends_raw_detail_rows_under_each_summary():
    """세부계정 행이 그 대분류(요약) 행 **바로 아래**에 원문 라벨 그대로 붙는다."""
    df = results_to_selection_dataframe([_sample_result()], {1: _sample_detail()})

    # 요약 24행 + 세부 4행
    assert len(df) == 28
    assert list(df["row_kind"]).count("요약") == 24
    assert list(df["row_kind"]).count("세부") == 4

    labels = list(df["account_name"])
    assert labels[:5] == [
        "유동자산",  # 요약
        "(1) 당좌자산",  # 이하 원문 그대로(표준화하지 않는다)
        "1.현금및현금등가물",
        "2.매출채권",
        "비유동자산",  # 다음 요약 항목
    ]
    detail = df[df["row_kind"] == "세부"]
    # 상위계정은 항상 **대분류(요약) 라벨**이다(직속 부모가 아니다 — 깊이는 계정단계).
    assert set(detail[detail["parent_account"] == "유동자산"]["account_level"]) == {1, 2}
    assert list(df[df["account_name"] == "가. 제품매출"]["parent_account"]) == ["매출액"]
    # 재무제표명은 그 대분류가 속한 표를 따른다(② 시트와 같은 분류 체계).
    assert list(df[df["account_name"] == "가. 제품매출"]["statement_name"]) == ["손익계산서"]
    assert set(detail[detail["parent_account"] == "유동자산"]["statement_name"]) == {"재무상태표"}


def test_selection_detail_rows_carry_current_period_amount_as_integer():
    """세부계정 금액은 당기(cur)만 싣고, float은 정수로 되돌린다(csv `1234.0` 방지)."""
    df = results_to_selection_dataframe([_sample_result()], {1: _sample_detail()})

    row = df[df["account_name"] == "1.현금및현금등가물"].iloc[0]
    assert row["amount"] == 1_200_000_000
    assert isinstance(row["amount"], int)
    # 전기(prv)는 어디에도 실리지 않는다(요약 항목과 같은 방침).
    assert 2.0 not in list(df["amount"]) and 2 not in list(df["amount"])
    # 값이 없는 세부계정도 행은 남고 금액만 빈다.
    assert df[df["account_name"] == "2.매출채권"].iloc[0]["amount"] is None


def test_selection_detail_label_keeps_raw_text_but_drops_control_chars():
    """라벨은 원문 그대로(각주·개행 포함) 두되 xlsx가 거부하는 제어문자만 없앤다."""
    detail = SelectionAccountDetail(
        accounts={
            "current_assets": [
                AccountRow(label="1.현금및현금성자산3>\n(주1)\x0b", level=1, cur=1.0, prv=None)
            ]
        }
    )
    df = results_to_selection_dataframe([_sample_result()], {1: detail})
    label = list(df[df["row_kind"] == "세부"]["account_name"])[0]
    assert label == "1.현금및현금성자산3>\n(주1)"  # 개행/각주 표기는 그대로
    # 제어문자가 남으면 openpyxl이 IllegalCharacterError로 다운로드 전체를 실패시킨다.
    export_selection_results([_sample_result()], "xlsx", {1: detail})


def test_selection_detail_notice_repeats_on_every_row_of_that_company():
    """세부계정을 못 얻은 회사는 사유가 그 회사 **모든 행**에 반복된다."""
    df = results_to_selection_dataframe(
        [_sample_result()], {1: SelectionAccountDetail(notice="원문 없음(감사보고서 미공시)")}
    )
    assert len(df) == 24  # 세부계정 행은 없다
    assert set(df["detail_notice"]) == {"원문 없음(감사보고서 미공시)"}
    assert set(df["row_kind"]) == {"요약"}


def test_selection_dataframe_without_details_keeps_previous_row_shape():
    """세부계정을 넘기지 않으면 2026-08-06 이전과 같은 요약 24행 그대로다."""
    df = results_to_selection_dataframe([_sample_result()])
    assert len(df) == 24
    assert set(df["row_kind"]) == {"요약"}
    assert df["parent_account"].isna().all()
    assert df["detail_notice"].isna().all()
    assert set(df["account_level"]) == {0}


def test_selection_detail_only_applies_to_matching_result_id():
    """세부계정 매핑은 result.id로만 붙는다 — 다른 회사에 남의 계정이 섞이면 안 된다."""
    first = _sample_result()
    second = _sample_result()
    second.id = 2
    second.corp_name = "㈜둘째"
    df = results_to_selection_dataframe([first, second], {1: _sample_detail()})

    assert len(df) == 28 + 24
    assert set(df[df["corp_name"] == "㈜둘째"]["row_kind"]) == {"요약"}
    assert list(df[df["account_name"] == "가. 제품매출"]["corp_name"]) == ["㈜테스트"]


def test_export_selection_results_csv_writes_detail_rows():
    content = export_selection_results([_sample_result()], "csv", {1: _sample_detail()})
    lines = content.decode("utf-8-sig").splitlines()

    assert len(lines) == 29  # 헤더 + 요약 24 + 세부 4
    assert any(
        line.endswith("재무상태표,세부,유동자산,1.현금및현금등가물,2,1200000000,OK,")
        for line in lines[1:]
    )
    assert "1200000000.0" not in content.decode("utf-8-sig")


def test_export_results_with_history_selection_sheet_carries_detail_rows():
    """2시트 xlsx의 ① 시트에도 세부계정이 실리고 ② 시트는 무변경이다."""
    content = export_results_with_history(
        [_sample_result()],
        _sample_snapshots(),
        {1: "㈜테스트"},
        use_selection_format=True,
        detail_by_result_id={1: _sample_detail()},
    )
    wb = openpyxl.load_workbook(io.BytesIO(content))
    assert wb["results"].max_row == 29  # 헤더 + 요약 24 + 세부 4
    # ② 시트는 연도별 요약 24항목 그대로다(세부계정 대상 아님).
    assert wb["financial_history"].max_row == 49


def test_full_export_still_uses_wide_format():
    """필터 전체 내보내기는 long 포맷 교체와 무관하게 기존 wide 포맷 그대로다."""
    content = export_results([_sample_result()], "xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["results"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    assert header == list(RESULT_COLUMN_LABELS.values())
    assert "매출액(전기)" in header
    assert "계정과목명" not in header
    assert ws.max_row == 2  # 회사 1건 = 1행


# ---------------------------------------------------------------------------
# 다중 선택 다운로드(§4-11, M9) — snapshots_to_dataframe / export_results_with_history
# ---------------------------------------------------------------------------


def _sample_snapshots() -> list[FinancialSnapshot]:
    return [
        FinancialSnapshot(
            id=11,
            result_id=1,
            rcept_no="20250401000001",
            fiscal_year="2024",
            total_assets=5_000_000_000,
            revenue=8_000_000_000,
            cf_operating=1_000_000,
            non_operating_income=2_000_000,
            auditor_name="안경회계법인",
            parse_status=ParseStatus.OK,
            from_current_period=1,
        ),
        FinancialSnapshot(
            id=12,
            result_id=1,
            rcept_no="20260401000001",
            fiscal_year="2025",
            total_assets=6_000_000_000,
            revenue=10_000_000_000,
            auditor_name="다른회계법인",
            parse_status=ParseStatus.OK,
            from_current_period=1,
        ),
    ]


def test_snapshots_to_dataframe_joins_corp_name_by_result_id():
    """스냅샷 1건이 계정과목 24행으로 풀리고 식별 컬럼은 각 행에 반복된다(2026-07-29 long)."""
    df = snapshots_to_dataframe(_sample_snapshots(), {1: "㈜테스트"})

    assert list(df.columns) == list(FINANCIAL_SNAPSHOT_COLUMN_LABELS.keys())
    assert len(df) == 2 * len(FINANCIAL_SNAPSHOT_ACCOUNT_COLUMNS) == 48
    assert set(df["corp_name"]) == {"㈜테스트"}
    # 입력 순서(result_id → fiscal_year 오름차순) 보존: 앞 24행이 2024, 뒤 24행이 2025
    assert list(df["fiscal_year"])[:24] == ["2024"] * 24
    assert list(df["fiscal_year"])[24:] == ["2025"] * 24
    assert list(df["rcept_no"])[:24] == ["20250401000001"] * 24

    y2024 = df[df["fiscal_year"] == "2024"]
    assert y2024[y2024["account_name"] == "자산총계"].iloc[0]["amount"] == 5_000_000_000
    assert y2024[y2024["account_name"] == "영업활동현금흐름"].iloc[0]["amount"] == 1_000_000
    assert y2024[y2024["account_name"] == "영업외수익"].iloc[0]["amount"] == 2_000_000


def test_snapshots_to_dataframe_column_order_and_statement_classification():
    """컬럼 순서(7개)와 재무제표명 분류·행 순서를 사용자 확정안대로 잠근다(2026-07-29)."""
    df = snapshots_to_dataframe(_sample_snapshots()[:1], {1: "㈜테스트"})

    assert list(FINANCIAL_SNAPSHOT_COLUMN_LABELS.values()) == [
        "결과ID",
        "회사명",
        "회계연도",
        "접수번호",
        "재무제표명",
        "계정과목",
        "금액",
    ]
    # 감사인/파싱상태는 완전히 제거됐다.
    assert "감사인" not in FINANCIAL_SNAPSHOT_COLUMN_LABELS.values()
    assert "파싱상태" not in FINANCIAL_SNAPSHOT_COLUMN_LABELS.values()
    assert "auditor_name" not in df.columns and "parse_status" not in df.columns

    # 재무제표명 → 계정과목 순서: 재무상태표 9 → 손익계산서 9(영업외손익 포함) →
    # 현금흐름표 6 (2026-08-05 세부계정 5항목이 각 표 끝에 붙어 7/8/4 → 9/9/6).
    assert list(df["statement_name"]) == ["재무상태표"] * 9 + ["손익계산서"] * 9 + ["현금흐름표"] * 6
    assert list(df["account_name"]) == [
        "유동자산",
        "비유동자산",
        "자산총계",
        "유동부채",
        "비유동부채",
        "부채총계",
        "자본총계",
        "현금및현금성자산(순액)",
        "매출채권(순액)",
        "매출액",
        "매출원가",
        "매출총이익",
        "판매비와관리비",
        "영업이익",
        "당기순이익",
        "영업외수익",
        "영업외비용",
        "이자비용",
        "영업활동현금흐름",
        "투자활동현금흐름",
        "재무활동현금흐름",
        "기말의현금",
        "감가상각비(현금흐름표)",
        "무형자산상각비",
    ]


def test_snapshot_account_columns_match_selection_account_columns():
    """스냅샷 계정과목 24개가 `results`의 `_cur` 항목 전체와 1:1로 대응해야 한다.

    새 `_cur`/`_prv` 쌍이 추가되면 `financial_snapshots`에도 접미어 없는 같은
    필드가 생기는데(CF 4항목·영업외손익 2항목이 그랬다), 이 시트의 하드코딩 목록에
    반영하지 않으면 **에러 없이 조용히** 빠진다 — 그 드리프트를 여기서 잡는다.
    """
    assert {c.removesuffix("_cur") for c in SELECTION_ACCOUNT_COLUMNS} == set(
        FINANCIAL_SNAPSHOT_ACCOUNT_COLUMNS
    )
    # 라벨도 두 시트가 같아야 한다("매출액" 등 — 한쪽만 바뀌면 대조가 어려워진다).
    for col in SELECTION_ACCOUNT_COLUMNS:
        assert SELECTION_ACCOUNT_LABELS[col] == FINANCIAL_SNAPSHOT_ACCOUNT_LABELS[
            col.removesuffix("_cur")
        ]
    # 모든 계정과목이 정확히 하나의 재무제표에 속한다.
    assert set(FINANCIAL_SNAPSHOT_STATEMENT_BY_ACCOUNT) == set(FINANCIAL_SNAPSHOT_ACCOUNT_COLUMNS)
    assert set(FINANCIAL_SNAPSHOT_STATEMENT_BY_ACCOUNT.values()) == {
        "재무상태표",
        "손익계산서",
        "현금흐름표",
    }


def test_account_detail_parser_fields_are_all_known_selection_accounts():
    """세부계정 파서가 만들 수 있는 필드 키가 ① 시트 요약 24항목 안에 전부 있어야 한다.

    `results_to_selection_dataframe()`은 `children_by_field.get(col.removesuffix("_cur"),
    ())`로 세부계정을 붙이므로, 파서 쪽에 요약 목록에 없는 필드 키가 새로 생기면
    그 대분류의 세부계정이 **에러 없이 조용히** 파일에서 빠진다(§4-15, dart-qa
    2026-08-06 Low-2). 위 두 드리프트 가드와 같은 성격의 검사다.
    """
    produced = set(_BS_IS_VALID_FIELDS) | set(CF_FINANCIAL_FIELDS)
    assert produced == set(DIRECT_FINANCIAL_FIELDS) | set(NON_OPERATING_FINANCIAL_FIELDS) | set(
        CF_FINANCIAL_FIELDS
    )
    assert produced <= {c.removesuffix("_cur") for c in SELECTION_ACCOUNT_COLUMNS}


def test_detail_account_fields_are_exported_in_both_sheets():
    """세부계정 5항목(2026-08-05)이 두 시트 모두에 원천 재무제표 기준으로 실린다.

    파서(`DETAIL_FINANCIAL_FIELDS`)가 뽑은 5항목이 DB 컬럼까지는 왔는데 export
    목록에 빠지면 **에러 없이 조용히** 안 나간다 — 필드 자체의 존재와 소속
    재무제표(원천 고정)를 함께 잠근다. 라벨의 출처 표기("(순액)"/"(현금흐름표)")는
    같은 이름의 다른 숫자와 혼동하지 않기 위한 것이라 문구까지 검증한다.
    """
    assert DETAIL_FINANCIAL_FIELDS == (
        "cash_and_equivalents",
        "trade_receivables",
        "interest_expense",
        "depreciation",
        "amortization",
    )
    # ① 시트: 당기(_cur) 컬럼으로 전부 실린다.
    for field in DETAIL_FINANCIAL_FIELDS:
        assert f"{field}_cur" in SELECTION_ACCOUNT_COLUMNS
        # wide 포맷에도 당기/전기 쌍으로 있다(전체 내보내기 경로).
        assert f"{field}_cur" in RESULT_COLUMN_LABELS
        assert f"{field}_prv" in RESULT_COLUMN_LABELS
    # ② 시트: 각 필드는 파서가 고정한 원천 재무제표에 속한다.
    assert FINANCIAL_SNAPSHOT_STATEMENT_BY_ACCOUNT["cash_and_equivalents"] == "재무상태표"
    assert FINANCIAL_SNAPSHOT_STATEMENT_BY_ACCOUNT["trade_receivables"] == "재무상태표"
    assert FINANCIAL_SNAPSHOT_STATEMENT_BY_ACCOUNT["interest_expense"] == "손익계산서"
    assert FINANCIAL_SNAPSHOT_STATEMENT_BY_ACCOUNT["depreciation"] == "현금흐름표"
    assert FINANCIAL_SNAPSHOT_STATEMENT_BY_ACCOUNT["amortization"] == "현금흐름표"
    # 라벨은 출처를 밝힌다(매출채권 총액/판관비 감가상각비와 혼동 방지).
    assert FINANCIAL_SNAPSHOT_ACCOUNT_LABELS["cash_and_equivalents"] == "현금및현금성자산(순액)"
    assert FINANCIAL_SNAPSHOT_ACCOUNT_LABELS["trade_receivables"] == "매출채권(순액)"
    assert FINANCIAL_SNAPSHOT_ACCOUNT_LABELS["depreciation"] == "감가상각비(현금흐름표)"
    # 각 표의 항목 수(7/8/4 → 9/9/6)
    assert [len(a) for a in FINANCIAL_SNAPSHOT_ACCOUNTS_BY_STATEMENT.values()] == [9, 9, 6]


def test_detail_account_values_reach_both_sheets():
    """세부계정 값이 실제로 두 시트의 "금액" 칸까지 흘러간다(결측이면 빈 값)."""
    result = _sample_result()
    result.cash_and_equivalents_cur = 1_500_000_000
    result.trade_receivables_cur = 2_500_000_000
    result.depreciation_cur = 300_000_000
    df = results_to_selection_dataframe([result])
    assert df[df["account_name"] == "현금및현금성자산(순액)"].iloc[0]["amount"] == 1_500_000_000
    assert df[df["account_name"] == "매출채권(순액)"].iloc[0]["amount"] == 2_500_000_000
    assert df[df["account_name"] == "감가상각비(현금흐름표)"].iloc[0]["amount"] == 300_000_000
    # best-effort라 결측이 정상 — 행은 남고 금액만 빈다.
    assert df[df["account_name"] == "무형자산상각비"].iloc[0]["amount"] is None
    assert df[df["account_name"] == "이자비용"].iloc[0]["amount"] is None

    snapshot = _sample_snapshots()[0]
    snapshot.interest_expense = 120_000_000
    snapshot.amortization = 45_000_000
    history = snapshots_to_dataframe([snapshot], {1: "㈜테스트"})
    assert history[history["account_name"] == "이자비용"].iloc[0]["amount"] == 120_000_000
    assert history[history["account_name"] == "무형자산상각비"].iloc[0]["amount"] == 45_000_000
    assert history[history["account_name"] == "현금및현금성자산(순액)"].iloc[0]["amount"] is None


def test_snapshots_to_dataframe_keeps_rows_for_missing_amounts():
    """값이 None인 계정과목도 행은 남기고 금액만 비운다(기본정보 시트와 동일 방침)."""
    df = snapshots_to_dataframe(_sample_snapshots()[:1], {1: "㈜테스트"})

    assert len(df) == 24
    assert df[df["account_name"] == "유동자산"].iloc[0]["amount"] is None
    # 값이 있는 것은 4개(자산총계/매출액/영업활동현금흐름/영업외수익)
    assert len(df[df["amount"].isna()]) == 20
    # float64로 승격되면 5000000000.0처럼 소수점이 붙는다.
    assert df["amount"].dtype == object


def test_snapshots_to_dataframe_unknown_result_id_leaves_corp_name_blank():
    """매핑에 없는 result_id는 회사명만 비고 파일 생성 자체는 실패하지 않는다."""
    df = snapshots_to_dataframe(_sample_snapshots(), {})
    assert df["corp_name"].isna().all()
    assert len(df) == 48


def test_export_results_with_history_writes_two_sheets():
    content = export_results_with_history(
        [_sample_result()], _sample_snapshots(), {1: "㈜테스트"}
    )
    wb = openpyxl.load_workbook(io.BytesIO(content))
    assert wb.sheetnames == ["results", "financial_history"]

    # 기본정보 시트는 기본값(=필터 전체 내보내기)에서 기존 wide 포맷이다 —
    # long 포맷은 `use_selection_format=True`(=`ids` 지정)일 때만 쓴다.
    results_ws = wb["results"]
    results_header = [c.value for c in next(results_ws.iter_rows(min_row=1, max_row=1))]
    assert results_header == list(RESULT_COLUMN_LABELS.values())
    assert results_ws.max_row == 2  # 헤더 + 회사 1건 = 1행

    history_ws = wb["financial_history"]
    history_header = [c.value for c in next(history_ws.iter_rows(min_row=1, max_row=1))]
    assert history_header == list(FINANCIAL_SNAPSHOT_COLUMN_LABELS.values())
    assert history_ws.max_row == 49  # 헤더 + 스냅샷 2건 × 계정과목 24행(2026-07-29 long)
    assert history_ws.cell(row=2, column=history_header.index("회사명") + 1).value == "㈜테스트"
    year_cell = history_ws.cell(row=2, column=history_header.index("회계연도") + 1).value
    assert str(year_cell) == "2024"

    stmt_col = history_header.index("재무제표명") + 1
    account_col = history_header.index("계정과목") + 1
    amount_col = history_header.index("금액") + 1
    assert history_ws.cell(row=2, column=stmt_col).value == "재무상태표"
    assert history_ws.cell(row=2, column=account_col).value == "유동자산"
    assert history_ws.cell(row=2, column=amount_col).value is None  # 결측은 빈 셀
    # 2024 스냅샷의 자산총계(재무상태표 3번째 행 = 시트 4행)
    assert history_ws.cell(row=4, column=account_col).value == "자산총계"
    assert history_ws.cell(row=4, column=amount_col).value == 5_000_000_000
    # 첫 스냅샷(2024)이 2~25행, 두 번째 스냅샷(2025)은 26행부터 시작한다.
    year_col = history_header.index("회계연도") + 1
    assert str(history_ws.cell(row=25, column=year_col).value) == "2024"
    assert str(history_ws.cell(row=26, column=year_col).value) == "2025"


def test_export_results_with_history_uses_long_format_only_when_asked():
    """기본정보 시트 포맷은 `use_selection_format`(=호출부의 `ids` 유무)만 따른다.

    `include_history`가 포맷을 바꾸면 `ids` 없이 이력만 요청한 전체 내보내기가
    조용히 long 포맷이 된다(dart-qa 2026-07-28 지적) — 두 경우를 함께 잠근다.
    """
    long_wb = openpyxl.load_workbook(
        io.BytesIO(
            export_results_with_history(
                [_sample_result()],
                _sample_snapshots(),
                {1: "㈜테스트"},
                use_selection_format=True,
            )
        )
    )
    long_ws = long_wb["results"]
    long_header = [c.value for c in next(long_ws.iter_rows(min_row=1, max_row=1))]
    assert long_header == list(SELECTION_EXPORT_COLUMN_LABELS.values())
    assert long_ws.max_row == 25  # 헤더 + 회사 1건 × 계정과목 24행
    # 시트 ②는 포맷과 무관하게 동일하다(스냅샷 2건 × 계정과목 24행 + 헤더).
    assert long_wb["financial_history"].max_row == 49

    wide_wb = openpyxl.load_workbook(
        io.BytesIO(
            export_results_with_history(
                [_sample_result()],
                _sample_snapshots(),
                {1: "㈜테스트"},
                use_selection_format=False,
            )
        )
    )
    wide_ws = wide_wb["results"]
    wide_header = [c.value for c in next(wide_ws.iter_rows(min_row=1, max_row=1))]
    assert wide_header == list(RESULT_COLUMN_LABELS.values())
    assert "매출액(전기)" in wide_header and "계정과목명" not in wide_header
    assert wide_ws.max_row == 2
    assert wide_wb["financial_history"].max_row == 49


def test_export_results_with_history_empty_inputs_return_header_only_sheets():
    content = export_results_with_history([], [], {})
    wb = openpyxl.load_workbook(io.BytesIO(content))
    assert wb.sheetnames == ["results", "financial_history"]
    assert wb["results"].max_row == 1
    assert wb["financial_history"].max_row == 1


# ---------------------------------------------------------------------------
# xlsx 스트리밍 직렬화(2026-08-06, dart-qa Medium-2) — pandas 출력과의 동등성
# ---------------------------------------------------------------------------


def _describe_workbook(content: bytes) -> list[tuple]:
    """xlsx 바이트에서 시트 구성·셀 값·헤더 서식만 뽑아 비교 가능한 형태로."""
    wb = openpyxl.load_workbook(io.BytesIO(content))
    described: list[tuple] = [tuple(wb.sheetnames)]
    for ws in wb.worksheets:
        described.append((ws.title, ws.max_row, ws.max_column))
        for row in ws.iter_rows():
            for cell in row:
                described.append(
                    (
                        ws.title,
                        cell.coordinate,
                        cell.value,
                        type(cell.value).__name__,
                        cell.font.bold,
                        cell.alignment.horizontal,
                        cell.alignment.vertical,
                        cell.border.left.style,
                        cell.border.right.style,
                        cell.border.top.style,
                        cell.border.bottom.style,
                        cell.number_format,
                    )
                )
    return described


def test_write_xlsx_output_matches_pandas_to_excel():
    """스트리밍 writer(`_write_xlsx`)가 기존 `DataFrame.to_excel`과 같은 파일을 만든다.

    §4-15로 ① 시트 행 수가 4배 가까이 늘면서 pandas/openpyxl 기본 모드의 피크
    메모리(전 셀을 `Cell` 객체로 보관)가 3.3GB까지 뛰어 write-only 스트리밍으로
    바꿨는데(dart-qa 2026-08-06 Medium-2), **출력물이 달라지면 안 된다** — 셀 값과
    헤더 서식(굵게/얇은 테두리/가운데·위 정렬)을 pandas 결과와 직접 대조해 잠근다.
    pandas가 기본 헤더 서식을 바꾸면 이 테스트가 먼저 깨진다.

    (알려진 차이 2가지는 비교 대상이 아니다 — 빈 값 셀을 pandas는 빈 inlineStr로
    쓰고 write-only는 아예 쓰지 않으며, `<dimension>` 힌트가 빠진다. 둘 다 재읽기
    값·엑셀 표시가 동일하다.)
    """
    import pandas as pd

    first, second = _sample_result(), _sample_result()
    second.id = 2
    second.corp_name = "㈜둘째"
    second.revenue_cur = None  # 결측 → pandas는 NaN, 파일에서는 빈 칸
    second.total_assets_cur = -5_000_000_000  # 음수(자본잠식 등 실재값)
    second.parse_status = ParseStatus.FAILED
    results = [first, second]
    detail = {1: _sample_detail(), 2: SelectionAccountDetail(notice="원문 캐시 없음(재수집 필요)")}
    snapshots = [
        FinancialSnapshot(result_id=1, fiscal_year="2025", rcept_no="R2", revenue=1_000),
        FinancialSnapshot(result_id=2, fiscal_year="2024", rcept_no=None, revenue=None),
    ]

    sheets = [
        (
            "results",
            results_to_selection_dataframe(results, detail).rename(
                columns=SELECTION_EXPORT_COLUMN_LABELS
            ),
        ),
        (
            "wide",
            results_to_dataframe(results).rename(columns=RESULT_COLUMN_LABELS),
        ),
        (
            "financial_history",
            snapshots_to_dataframe(snapshots, {1: "㈜테스트", 2: None}).rename(
                columns=FINANCIAL_SNAPSHOT_COLUMN_LABELS
            ),
        ),
    ]

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for name, df in sheets:
            df.to_excel(writer, sheet_name=name, index=False)

    assert _describe_workbook(_write_xlsx(sheets)) == _describe_workbook(buffer.getvalue())


def test_write_xlsx_rejects_sheet_over_excel_row_limit(monkeypatch):
    """엑셀 시트 행 상한을 넘으면 손상된 파일을 만들지 않고 즉시 실패한다.

    pandas는 이 검사를 자체적으로 했지만(`This sheet is too large!`) openpyxl
    write-only 모드에는 없어, 같은 성격의 방어를 exporter에 둔다. 사용자에게 보이는
    안내는 API 계층(`export_job_results`)의 400이고 이쪽은 최후의 방어선이다.
    """
    import app.exporters.excel as excel_module

    monkeypatch.setattr(excel_module, "XLSX_MAX_ROWS", 3)  # 헤더 포함 3행까지
    df = results_to_dataframe([_sample_result()] * 3).rename(columns=RESULT_COLUMN_LABELS)
    with pytest.raises(ValueError, match="최대 행 수"):
        _write_xlsx([("results", df)])


def _isolate_openpyxl_tempdir(tmp_path, monkeypatch):
    """openpyxl write-only가 만드는 임시파일을 테스트 전용 폴더로 몰아 넣는다.

    openpyxl은 `NamedTemporaryFile`로 시트를 스풀링하고 그 함수는
    `tempfile.gettempdir()`(=`tempfile.tempdir` 전역)을 따르므로, 이 폴더가 비어
    있는지로 "임시파일이 남았는가"를 정확히 판정할 수 있다(실제 `%TEMP%`를 훑으면
    다른 프로세스의 파일과 섞인다).
    """
    import tempfile

    monkeypatch.setattr(tempfile, "tempdir", str(tmp_path))
    return tmp_path


def test_write_xlsx_cleans_temp_files_when_row_write_fails(tmp_path, monkeypatch):
    """직렬화 도중 예외가 나도 `%TEMP%`에 임시파일을 남기지 않는다(dart-qa 2026-08-06 Low).

    write-only 모드는 행을 임시파일로 흘려보내고 `save()`가 끝나야 그것을 지운다 —
    `ws.append()`에서 예외가 나면 `save()`에 도달하지 못해 파일이 **영구히** 남았다
    (Windows에서는 핸들이 열린 채라 openpyxl의 atexit 정리마저 `PermissionError`로
    실패한다). 최대 규모 다운로드 1건의 시트 XML이 압축 전 약 604MB라, 사무실 exe에서
    재시도할 때마다 그만큼 쌓인다.

    재현은 dart-qa 시나리오 그대로 제어문자 → `IllegalCharacterError`로 한다. 지금은
    `_xlsx_cell_value()`가 제어문자를 미리 떼므로(같은 날 함께 반영), 그 방어를 끈
    상태에서 openpyxl이 실제로 거부하게 만든다 — 이 테스트가 잠그는 것은 "제어문자를
    거른다"가 아니라 **"어떤 이유로든 쓰기가 실패해도 임시파일이 정리된다"** 이다.
    """
    from openpyxl.utils.exceptions import IllegalCharacterError

    import app.exporters.excel as excel_module

    _isolate_openpyxl_tempdir(tmp_path, monkeypatch)
    monkeypatch.setattr(excel_module, "_xlsx_cell_value", lambda value: value)

    import pandas as pd

    df = pd.DataFrame({"회사명": ["정상 회사", "제어문자\x01포함"]})
    with pytest.raises(IllegalCharacterError):
        _write_xlsx([("results", df)])

    assert list(tmp_path.iterdir()) == []


def test_write_xlsx_cleans_temp_files_on_success(tmp_path, monkeypatch):
    """정상 종료 경로에서도 임시파일이 남지 않는다(정리 함수의 중복 호출이 무해한지 확인)."""
    _isolate_openpyxl_tempdir(tmp_path, monkeypatch)

    df = results_to_dataframe([_sample_result()]).rename(columns=RESULT_COLUMN_LABELS)
    content = _write_xlsx([("results", df), ("financial_history", df)])

    assert content[:2] == b"PK"
    assert list(tmp_path.iterdir()) == []


def test_xlsx_export_strips_control_chars_from_base_columns():
    """기본정보 컬럼(회사명/주소/감사인명 등)의 제어문자도 제거한다 — 선행 이슈 보정.

    세부계정 라벨만 `_detail_label()`로 막고 있어, DART 원문 유래 값에 제어문자가
    섞이면 `IllegalCharacterError`로 **다운로드 전체가 실패**했다(pandas `to_excel`
    경로에서도 같았던 선행 이슈. 개발 DB 5,204행 실측 검출 0건이지만 방어는 둔다).
    """
    result = _sample_result()
    result.corp_name = "㈜테\x01스트"
    result.address = "경상남도 김해시\x0b삼계로 1"
    result.auditor_name = "회계법인\x1f한빛"

    wide = openpyxl.load_workbook(io.BytesIO(export_results([result], "xlsx")))["results"]
    wide_row = {
        header.value: cell.value
        for header, cell in zip(wide[1], wide[2], strict=False)
    }
    assert wide_row["회사명"] == "㈜테스트"
    assert wide_row["주소"] == "경상남도 김해시삼계로 1"
    assert wide_row["감사인"] == "회계법인한빛"

    selection = openpyxl.load_workbook(
        io.BytesIO(export_selection_results([result], "xlsx"))
    )["results"]
    selection_row = {
        header.value: cell.value
        for header, cell in zip(selection[1], selection[2], strict=False)
    }
    assert selection_row["회사명"] == "㈜테스트"
    assert selection_row["주소"] == "경상남도 김해시삼계로 1"
    assert selection_row["감사인"] == "회계법인한빛"
