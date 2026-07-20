"""M8 1단계 — DART 기업개황(`dsae001`) 전역 인덱스 크롤러. 상세개발계획.md §4-10.

`fsc_corp_index`(금융위 `getCorpOutline_V2` 전수 크롤, 10~16시간)를 대체한다.
DART 전자공시 웹의 "기업개황 > 업종별" 화면이 쓰는 엔드포인트 3개를 그대로
사용하며, 실측 기준 **전체 구축에 약 23분**이 걸린다.

    selectCorpTree.do  업종 트리(대 21 / 중 77 / 소 234 / 세 501 / 세세 1,205)
    downloadExcel.do   조건에 해당하는 전체 회사를 xlsx로 반환 (행 수 상한 없음)
    search.ax          같은 조건의 목록 HTML — corp_code를 직접 포함

## 왜 이 구조인가 (§4-10-A 실측)

- `businessCode`에 **중분류(2자리)** 를 넣으면 하위 세부업종이 전부 포함되므로
  leaf 1,205개가 아니라 **77개**만 돌면 전국이 커버된다(실측 118,266행 / 94.6초).
- 엑셀에는 `corp_code`가 없지만 `search.ax`가 같은 조건·같은 정렬로 45건씩
  페이징하며 준다. 두 응답을 **위치(index)로 결합**한다 — 중분류 `47` 전량
  (2,402행)으로 검증한 결과 순서 무결성 100%, 정규화 이름 일치 99.38%
  (불일치분은 `CJ프레시웨이` vs `씨제이프레시웨이`처럼 공시회사명 표기 차이).
- 위치 결합은 순서가 어긋나면 **전 행이 조용히 오염**되므로,
  `merge_by_position()`이 행 수 일치와 이름 일치율을 검증해 임계값 미만이면
  예외를 던진다. 부분 적재된 인덱스로 Job이 돌면 "후보가 원래 적었나 보다"로
  오인되기 때문에, 조용한 실패를 절대 허용하지 않는다(§4-10-G 열린 질문 1).

`robots.txt`가 막는 경로(`/dsaf001/main.do`, `/report/*`, `/pdf/download/`,
`/dsae001/selectPopup.ax`)는 사용하지 않는다. 요청 간격을 두고 갱신 주기는
연 1~2회로 제한한다.
"""

from __future__ import annotations

import asyncio
import io
import logging
import re
from datetime import datetime, timedelta
from typing import Any

import httpx
from openpyxl import load_workbook
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session, sessionmaker

from app.config import Settings, get_settings
from app.core.dart_client import DartApiError, QuotaExceededError
from app.core.db import get_session_factory
from app.core.filters import (
    cond_sido_list,
    industry_matches,
    normalize_corp_name,
    parse_address,
    region_matches,
)
from app.core.industry_data import INDUSTRIES
from app.models.corp_cache import CacheMeta
from app.models.dart_corp_index import DartCorpIndex

logger = logging.getLogger(__name__)

_BASE = "https://dart.fss.or.kr"
_TREE_URL = f"{_BASE}/dsae001/selectCorpTree.do"
_EXCEL_URL = f"{_BASE}/dsae001/downloadExcel.do"
_SEARCH_URL = f"{_BASE}/dsae001/search.ax"
_REFERER = f"{_BASE}/dsae001/main.do"

# 요청 간 최소 간격 — 공식 API가 아닌 웹 화면 엔드포인트이므로 보수적으로 둔다.
_REQUEST_INTERVAL_SEC = 0.3
_MAX_RETRIES = 3
_RETRY_BACKOFF_SEC = 2.0

# `search.ax` 한 페이지에 담기는 행 수(고정, maxResults 파라미터는 무시된다).
_SEARCH_PAGE_SIZE = 45

# 위치 결합 검증 임계값. 실측상 정상 업종은 포함 관계까지 인정하면 거의 100%이고
# 정렬이 어긋나면 0%에 가까우므로, 두 분포 사이에 넉넉한 여유를 둔 값을 쓴다.
# 처음에 0.95로 잡았다가 소규모 업종에서 두 번 오탐이 났다(65: 9/99, 84: 2/29) —
# **고정 비율은 표본이 작을수록 변동이 커서** 2,402행 표본 하나로 정한 값이
# 29행짜리 업종에는 맞지 않았다. 절대 허용 건수 하한을 함께 둔다.
_MERGE_NAME_MATCH_THRESHOLD = 0.70
_MERGE_MIN_ALLOWED_MISMATCHES = 3
# 축약형으로 인정할 최소 길이 비율 (`_names_align` 참고)
_MERGE_MIN_NAME_LENGTH_RATIO = 0.5

# 상장 시장 법인구분 — 이 회사들은 감사보고서를 별도 공시(pblntf_ty="F")하지 않고
# 사업보고서에 첨부하므로 Phase 2에서 전부 FAILED가 된다. 후보 단계에서 제외한다.
# 실측 분포: 유가증권시장 833 / 코스닥시장 1,818 / 코넥스시장 107 / 기타법인 115,510.
LISTED_CORP_CLS = ("유가증권시장", "코스닥시장", "코넥스시장")

_META_KEY_LAST_INDUSTRY = "dart_index_last_industry"
_META_KEY_UPDATED_AT = "dart_index_updated_at"
# 마지막으로 **전체** 동명 그룹 교정을 마친 시각. 크롤 완료 시각보다 오래됐거나
# 없으면 교정이 밀린 상태다 — `get_dart_index_status()`가 이를 드러낸다.
_META_KEY_RECONCILED_AT = "dart_index_reconciled_at"

# 엑셀 14열 고정 순서 (§4-10-A). 헤더 문자열 대신 위치로 읽되, 열 수로 검증한다.
_EXCEL_COLUMNS = (
    "corp_name",
    "eng_name",
    "disclosure_name",
    "stock_code",
    "ceo_name",
    "corp_cls",
    "jurir_no",
    "bizr_no",
    "address",
    "homepage",
    "ir_homepage",
    "induty_name",
    "est_date",
    "acc_month",
)

_SELECT_KEY_RE = re.compile(r"select\('(\d{8})'\)[^>]*>\s*([^<]+)")
_TOTAL_PAGE_RE = re.compile(r"\[\d+/(\d+)\]")


class DartIndexCrawlError(RuntimeError):
    """크롤 중 데이터 정합성이 깨졌을 때 — 부분 적재로 넘어가지 않고 중단시킨다."""


# ---------------------------------------------------------------------------
# HTTP
# ---------------------------------------------------------------------------


async def _post(client: httpx.AsyncClient, url: str, data: dict[str, str]) -> httpx.Response:
    """지수 백오프 재시도가 붙은 POST.

    `FscCorpInfoClient._get_with_retry`와 같은 정책이지만, 그쪽은 data.go.kr
    전용 쿼터 판정 로직과 얽혀 있어 재사용하지 않고 여기에 따로 둔다
    (2026-07-16 재시도 로직 도입 때와 같은 판단 — 과설계 방지).
    """
    last_exc: Exception | None = None
    for attempt in range(_MAX_RETRIES):
        try:
            resp = await client.post(url, data=data, headers={"Referer": _REFERER})
            if resp.status_code >= 500:
                raise httpx.HTTPStatusError(
                    f"HTTP {resp.status_code}", request=resp.request, response=resp
                )
            resp.raise_for_status()
            return resp
        except (httpx.TimeoutException, httpx.TransportError, httpx.HTTPStatusError) as exc:
            last_exc = exc
            if attempt < _MAX_RETRIES - 1:
                await asyncio.sleep(_RETRY_BACKOFF_SEC * (2**attempt))
    raise DartIndexCrawlError(f"{url} 요청 실패: {last_exc}") from last_exc


def _search_form(business_code: str, page: int = 1) -> dict[str, str]:
    """`searchForm`이 그대로 전송하는 필드 집합(§4-10-A).

    `corpTypeAll="all"`은 유가/코스닥/코넥스/기타를 모두 포함한다는 뜻이다
    (상장사는 적재 후 `stock_code`로 걸러내므로 여기서 제외하지 않는다 —
    `corp_cls` 분포 자체가 인덱스 검증에 쓰인다).
    """
    return {
        "currentPage": str(page),
        "maxResults": "",
        "maxLinks": "",
        "sort": "",
        "series": "",
        "gubun": "",
        "selectKey": "",
        "searchIndex": "",
        "textCrpCik": "",
        "autoSearch": "true",
        "businessCode": business_code,
        "bsnRgsNo": "",
        "crpRgsNo": "",
        "corpTypeAll": "all",
        "textCrpNm": "",
        "searchType": "1",
    }


# ---------------------------------------------------------------------------
# 업종 트리
# ---------------------------------------------------------------------------


async def fetch_industry_tree(client: httpx.AsyncClient) -> list[dict[str, Any]]:
    resp = await _post(client, _TREE_URL, {})
    return resp.json()


def mid_level_codes(tree: list[dict[str, Any]]) -> list[str]:
    """중분류(2자리 숫자) 코드 — 크롤 단위. 실측 77개이며 전국을 커버한다."""
    return sorted({n["ID"] for n in tree if len(str(n["ID"])) == 2 and str(n["ID"]).isdigit()})


def _normalize_induty_name(name: str | None) -> str:
    """업종명 비교용 정규화 — 공백/괄호/세미콜론 차이를 흡수한다.

    같은 KSIC 10차인데도 표기가 갈린다(`섬유제품 제조업(의복 제외)` vs
    `섬유제품 제조업; 의복제외`).
    """
    return "".join(str(name or "").split()).replace(";", "").replace("(", "").replace(")", "")


def build_induty_code_lookup(tree: list[dict[str, Any]]) -> dict[str, str]:
    """업종명 -> 트리 코드(2~5자리) 역매핑.

    회사마다 부여 깊이가 다르므로(2자리 5.18% / 3자리 20.35% / 4자리 15.72% /
    5자리 58.75%) 모든 레벨을 넣는다. 실측 표본 43,560행에서 미매칭 0%,
    이름이 여러 코드에 걸치는 경우 0종이었다(§4-10-G 열린 질문 2).
    """
    lookup: dict[str, str] = {}
    for node in tree:
        code = str(node["ID"])
        if not code.isdigit():
            continue
        key = _normalize_induty_name(node.get("TEXT"))
        # 이름 충돌이 없음을 실측했으나, 만약 생기면 더 얕은(짧은) 코드를 남긴다 —
        # prefix 매칭에서 누락이 덜 생기는 쪽이 안전하다.
        if key not in lookup or len(code) < len(lookup[key]):
            lookup[key] = code
    return lookup


# ---------------------------------------------------------------------------
# 엑셀 / corp_code 수집
# ---------------------------------------------------------------------------


def parse_industry_excel(content: bytes) -> list[dict[str, Any]]:
    """`downloadExcel.do` 응답(xlsx)을 행 dict 리스트로 파싱한다.

    헤더 문자열이 아니라 **위치**로 읽는다(열 순서는 고정). 열 수가 다르면
    DART가 서식을 바꾼 것이므로 즉시 실패시킨다.
    """
    # ⚠ `read_only=True`를 쓰면 안 된다 — DART가 생성한 xlsx는 시트 크기
    # 메타데이터가 잘못돼 있어(실측: 실제 14열 2,101행인데 `max_col=1, max_row=1`)
    # read_only 모드가 그 값을 믿고 첫 열만 읽는다. 일반 모드는 정상이다.
    wb = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    try:
        sheet = wb[wb.sheetnames[0]]
        rows_iter = sheet.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            return []
        if len(header) != len(_EXCEL_COLUMNS):
            raise DartIndexCrawlError(
                f"기업개황 엑셀 열 수가 {len(_EXCEL_COLUMNS)}이 아님({len(header)}) — 서식 변경 의심"
            )
        rows: list[dict[str, Any]] = []
        for raw in rows_iter:
            if raw is None or all(v is None or str(v).strip() == "" for v in raw):
                continue
            row = {
                key: (str(value).strip() if value is not None and str(value).strip() else None)
                for key, value in zip(_EXCEL_COLUMNS, raw)
            }
            if row.get("corp_name"):
                rows.append(row)
        return rows
    finally:
        wb.close()


async def fetch_industry_excel(client: httpx.AsyncClient, code: str) -> list[dict[str, Any]]:
    resp = await _post(client, _EXCEL_URL, _search_form(code))
    return parse_industry_excel(resp.content)


def parse_search_page(html: str) -> tuple[list[tuple[str, str]], int]:
    """목록 HTML에서 `[(corp_code, 표시명)]`과 전체 페이지 수를 뽑는다."""
    pairs = [(code, name.strip()) for code, name in _SELECT_KEY_RE.findall(html)]
    match = _TOTAL_PAGE_RE.search(html)
    total_pages = int(match.group(1)) if match else 1
    return pairs, total_pages


async def fetch_industry_corp_codes(
    client: httpx.AsyncClient, code: str
) -> list[tuple[str, str]]:
    """`search.ax`를 끝까지 페이징해 `corp_code` 순서 리스트를 만든다."""
    collected: list[tuple[str, str]] = []
    page = 1
    total_pages = 1
    while page <= total_pages:
        resp = await _post(client, _SEARCH_URL, _search_form(code, page))
        pairs, total_pages = parse_search_page(resp.text)
        if not pairs:
            break
        collected.extend(pairs)
        page += 1
        if page <= total_pages:
            await asyncio.sleep(_REQUEST_INTERVAL_SEC)
    return collected


# ---------------------------------------------------------------------------
# 위치 결합 + 검증
# ---------------------------------------------------------------------------


def _names_align(excel_name: str | None, listed_name: str) -> bool:
    """두 이름이 같은 회사를 가리키는지 — **부분 문자열 관계까지 허용**한다.

    `search.ax`는 회사명이 아니라 **공시회사명(축약형)** 을 보여주며, 축약은
    앞뒤 어느 쪽에서도 일어난다. 실측 예:

        (뒤가 잘림, 중분류 65 보험)   동양생명보험(주) ↔ 동양생명
                                    현대해상화재보험(주) ↔ 현대해상
                                    캑터스…사모투자 합자회사 ↔ 캑터스…사모투자
        (앞이 잘림, 중분류 84 공공)   재단법인 전북연구원 ↔ 전북연구원
                                    재단법인 한마음평화연구재단 ↔ 한마음평화연구재단

    처음엔 완전 일치만 인정해 보험업(9/99)에서, 다음엔 접두사까지만 허용해
    공공행정(2/29)에서 각각 크롤이 멈췄다 — 둘 다 순서는 완벽했고 오탐이었다.
    양방향 포함 관계로 두 유형을 모두 흡수한다.

    `normalize_corp_name`은 `㈜`(단일 문자)/`합자회사`/`재단법인`을 떼지 않지만,
    그 공유 함수를 바꾸면 기존 이름 매칭 동작까지 흔들리므로 여기서 흡수한다.
    정렬이 실제로 어긋나면 포함 관계도 성립하지 않는다(셔플 시 거의 0%).
    """
    left = normalize_corp_name(excel_name or "")
    right = normalize_corp_name(listed_name or "")
    if not left or not right:
        return False
    if left == right:
        return True
    shorter, longer = (left, right) if len(left) <= len(right) else (right, left)
    if shorter not in longer:
        return False
    # 포함 관계만으로는 우연한 일치를 걸러내지 못한다 — 짧은 이름은 무관한 긴
    # 이름 안에 들어가기 쉽다(예: "회사0" ⊂ "전혀다른회사0"). 축약형은 원본의
    # 절반 이상을 차지한다는 실측(최소 사례가 "현대해상"/"현대해상화재보험" =
    # 정확히 0.5)에 근거해 길이 비율 하한을 함께 건다.
    return len(shorter) / len(longer) >= _MERGE_MIN_NAME_LENGTH_RATIO


def merge_by_position(
    excel_rows: list[dict[str, Any]],
    code_pairs: list[tuple[str, str]],
    *,
    business_code: str,
    induty_lookup: dict[str, str] | None = None,
) -> list[dict[str, Any]]:
    """엑셀 행과 `corp_code`를 위치로 결합하고 정합성을 검증한다.

    검증에 실패하면 `DartIndexCrawlError`를 던진다 — 위치 결합은 순서가
    어긋나면 전 행이 조용히 오염되므로, 의심스러우면 적재하지 않는 편이 낫다.
    """
    if len(excel_rows) != len(code_pairs):
        raise DartIndexCrawlError(
            f"업종 {business_code}: 엑셀 {len(excel_rows)}행 vs 목록 {len(code_pairs)}행 불일치"
        )
    if not excel_rows:
        return []

    matched = sum(
        1
        for row, (_, listed_name) in zip(excel_rows, code_pairs)
        if _names_align(row.get("corp_name"), listed_name)
    )
    mismatched = len(excel_rows) - matched
    allowed = max(
        _MERGE_MIN_ALLOWED_MISMATCHES,
        int(len(excel_rows) * (1 - _MERGE_NAME_MATCH_THRESHOLD)),
    )
    if mismatched > allowed:
        raise DartIndexCrawlError(
            f"업종 {business_code}: 위치 결합 불일치 {mismatched}/{len(excel_rows)}건 "
            f"(허용 {allowed}건 초과) — 정렬/서식 변경 의심"
        )

    lookup = induty_lookup or {}
    now = datetime.now().isoformat(timespec="seconds")
    merged: list[dict[str, Any]] = []
    for row, (corp_code, _) in zip(excel_rows, code_pairs):
        address = row.get("address")
        sido, sigungu = parse_address(address or "")
        jurir_no = (row.get("jurir_no") or "").replace("-", "").strip() or None
        induty_name = row.get("induty_name")
        merged.append(
            {
                "corp_code": corp_code,
                "corp_name": row.get("corp_name"),
                "corp_name_norm": normalize_corp_name(row.get("corp_name") or ""),
                "eng_name": row.get("eng_name"),
                "disclosure_name": row.get("disclosure_name"),
                "stock_code": row.get("stock_code"),
                "corp_cls": row.get("corp_cls"),
                "ceo_name": row.get("ceo_name"),
                "jurir_no": jurir_no,
                "bizr_no": (row.get("bizr_no") or "").replace("-", "").strip() or None,
                "address": address,
                "sido": sido,
                "sigungu": sigungu,
                "homepage": row.get("homepage"),
                # 업종명 역매핑 실패 시 크롤 단위인 중분류 코드로 폴백한다
                # (필터가 조용히 빈 결과를 내는 것보다 낫다).
                "induty_code": lookup.get(_normalize_induty_name(induty_name), business_code),
                "induty_name": induty_name,
                "crawl_induty_code": business_code,
                "est_date": row.get("est_date"),
                "acc_month": row.get("acc_month"),
                "updated_at": now,
            }
        )
    return merged


# ---------------------------------------------------------------------------
# 적재
# ---------------------------------------------------------------------------


def upsert_dart_corp_index(db: Session, items: list[dict[str, Any]]) -> tuple[int, int]:
    """업종 1개 분량을 한 세션/한 커밋으로 upsert한다.

    건별 커밋은 A1 크롤에서 성능 병목으로 확인됐으므로(2026-07-16) 처음부터
    배치로 처리한다. PK가 `corp_code`(부분 인덱스가 아닌 일반 PK)라
    "부분 인덱스가 조회에 안 쓰이는" A1의 함정은 여기서는 발생하지 않는다.
    """
    if not items:
        return 0, 0
    # 같은 회사가 여러 업종에 중복 등장할 수 있으므로 배치 내에서 먼저 병합한다.
    deduped: dict[str, dict[str, Any]] = {}
    for item in items:
        deduped[item["corp_code"]] = item

    codes = list(deduped)
    existing = {
        row.corp_code: row
        for row in db.execute(
            select(DartCorpIndex).where(DartCorpIndex.corp_code.in_(codes))
        ).scalars()
    }
    inserted = updated = 0
    for corp_code, item in deduped.items():
        row = existing.get(corp_code)
        if row is None:
            db.add(DartCorpIndex(**item))
            inserted += 1
        else:
            for key, value in item.items():
                if key != "corp_code":
                    setattr(row, key, value)
            updated += 1
    db.commit()
    return inserted, updated


# ---------------------------------------------------------------------------
# 동명 회사 교차 교정 (2026-07-20 — M8 6단계 검증 중 발견한 버그의 수정)
# ---------------------------------------------------------------------------

# `merge_by_position()`이 순서 어긋남을 감지하는 유일한 신호는 **회사명 비교**다
# (`_names_align`). 그래서 같은 이름을 가진 회사끼리 자리가 바뀌면 이름 검사를
# 그대로 통과하고 주소·업종이 조용히 교차된다 — 공시목록 페이지와 업종 엑셀이
# 동명 회사에 대해 서로 다른 타이브레이크 순서를 쓰기 때문이다.
#
# 실측(2026-07-20, 표본 70건):
#   - 위험군(동일 정규화명 + 동일 크롤 업종) 40건 → 불일치 42.5%
#   - 대조군(이름 유일) 30건                      → 불일치 0.0%
# 인덱스 118,268행 중 위험군은 4,366행(3.69%)뿐이므로, 전체를 다시 받지 않고
# 이 그룹만 `company.json`으로 교정한다.
#
# **덮어쓰기가 아니라 순열 교정을 한다.** 교차는 그룹 내부의 순열이라 그룹이
# 보유한 속성 집합 자체는 보존된다 — 그래서 `jurir_no`(법인등록번호, 두 소스가
# 같은 13자리 무하이픈 형식으로 준다)를 키로 제자리를 찾아주면 주소뿐 아니라
# 업종명·대표자·설립일까지 전 필드가 한 번에 복원된다. company.json 값으로
# 덮어쓰기만 하면 `induty_name`처럼 company.json이 주지 않는 필드가 교차된 채
# 남는다. `jurir_no`로 짝을 못 찾은 경우에만 company.json 값으로 폴백한다.

# 순열 교정 시 corp_code를 제외하고 통째로 옮기는 필드 — 전부 엑셀 유래다.
_PERMUTED_FIELDS = (
    "corp_name",
    "corp_name_norm",
    "eng_name",
    "disclosure_name",
    "stock_code",
    "corp_cls",
    "ceo_name",
    "jurir_no",
    "bizr_no",
    "address",
    "sido",
    "sigungu",
    "homepage",
    "induty_code",
    "induty_name",
    "est_date",
    "acc_month",
)


def find_ambiguous_corp_codes(db: Session) -> list[list[str]]:
    """속성 교차가 일어날 수 있는 그룹(동일 정규화명 + 동일 크롤 업종)을 돌려준다.

    같은 업종 배치 안에서만 위치 결합이 일어나므로, 이름이 같아도 크롤 업종이
    다르면 서로 섞일 수 없다 — 그룹 키에 `crawl_induty_code`를 포함하는 이유다.
    """
    rows = db.execute(
        select(
            DartCorpIndex.corp_code,
            DartCorpIndex.corp_name_norm,
            DartCorpIndex.crawl_induty_code,
        )
    ).all()
    groups: dict[tuple[str, str], list[str]] = {}
    for corp_code, name_norm, crawl_code in rows:
        if not name_norm or not corp_code:
            continue
        groups.setdefault((name_norm, crawl_code or ""), []).append(corp_code)
    return [codes for codes in groups.values() if len(codes) > 1]


def _normalize_jurir_no(value: str | None) -> str | None:
    return (value or "").replace("-", "").strip() or None


def _apply_company_json(row: DartCorpIndex, company: dict[str, Any]) -> None:
    """`jurir_no`로 짝을 못 찾았을 때의 폴백 — company.json이 주는 필드만 덮어쓴다.

    `induty_name`은 company.json에 없으므로 지운다. 코드는 정본으로 바꿔 놓고
    이름만 교차된 값으로 남겨두면 화면에 서로 어긋난 업종이 표시된다.
    """
    address = (company.get("adres") or "").strip() or None
    sido, sigungu = parse_address(address or "")
    row.address = address
    row.sido = sido
    row.sigungu = sigungu
    row.ceo_name = company.get("ceo_nm") or None
    row.jurir_no = _normalize_jurir_no(company.get("jurir_no"))
    row.bizr_no = _normalize_jurir_no(company.get("bizr_no"))
    row.induty_code = company.get("induty_code") or None
    row.induty_name = None
    row.est_date = company.get("est_dt") or None
    row.acc_month = company.get("acc_mt") or None


async def reconcile_ambiguous_rows(
    dart_client: Any,
    session_factory: sessionmaker[Session],
    *,
    max_groups: int | None = None,
) -> dict[str, int]:
    """동명 그룹의 교차된 속성을 `company.json` 기준으로 제자리에 돌려놓는다.

    `QuotaExceededError`는 그대로 올려보내 호출부가 Job과 같은 방식으로 멈출 수
    있게 한다 — 이미 교정한 그룹은 커밋돼 있으므로 다시 호출하면 이어서 진행된다
    (교정이 끝난 그룹은 재실행해도 결과가 같은 멱등 연산이다).
    """
    with session_factory() as db:
        groups = find_ambiguous_corp_codes(db)
    if max_groups is not None:
        groups = groups[:max_groups]

    stats = {"groups": 0, "checked": 0, "repaired": 0, "fallback": 0, "failed": 0}
    for codes in groups:
        companies: dict[str, dict[str, Any]] = {}
        for corp_code in codes:
            try:
                companies[corp_code] = await dart_client.get_company(corp_code)
            except QuotaExceededError:
                raise
            except DartApiError as exc:
                logger.warning("동명 그룹 교정 — company.json 실패 %s: %s", corp_code, exc)
                stats["failed"] += 1
            stats["checked"] += 1
        if not companies:
            continue

        with session_factory() as db:
            rows = {
                row.corp_code: row
                for row in db.execute(
                    select(DartCorpIndex).where(DartCorpIndex.corp_code.in_(codes))
                ).scalars()
            }
            # 교정 전 스냅샷 — 그룹이 보유한 속성 집합을 jurir_no로 색인한다.
            snapshots = {
                row.jurir_no: {field: getattr(row, field) for field in _PERMUTED_FIELDS}
                for row in rows.values()
                if row.jurir_no
            }
            for corp_code, company in companies.items():
                row = rows.get(corp_code)
                if row is None:
                    continue
                want = _normalize_jurir_no(company.get("jurir_no"))
                source = snapshots.get(want) if want else None
                if source is not None:
                    if source["jurir_no"] != row.jurir_no:
                        stats["repaired"] += 1
                    for field, value in source.items():
                        setattr(row, field, value)
                else:
                    # 그룹 안에 이 법인등록번호를 가진 행이 없다 — 엑셀이 애초에
                    # 다른 회사를 담고 있었던 경우다. 정본으로 덮어쓴다.
                    _apply_company_json(row, company)
                    stats["fallback"] += 1
            db.commit()
        stats["groups"] += 1

    if max_groups is None:
        # 전체를 돌았을 때만 "교정 완료"로 기록한다 — 파일럿(`max_groups`)은
        # 일부만 손대므로 완료로 표시하면 남은 위험 그룹이 조용히 묻힌다.
        with session_factory() as db:
            _set_meta(db, _META_KEY_RECONCILED_AT, datetime.now().isoformat(timespec="seconds"))
            db.commit()
    return stats


def _set_checkpoint(factory: sessionmaker[Session], code: str) -> None:
    with factory() as db:
        _set_meta(db, _META_KEY_LAST_INDUSTRY, code)
        db.commit()


def _get_meta(db: Session, key: str) -> str | None:
    row = db.execute(select(CacheMeta).where(CacheMeta.key == key)).scalar_one_or_none()
    return row.value if row else None


def _set_meta(db: Session, key: str, value: str) -> None:
    row = db.execute(select(CacheMeta).where(CacheMeta.key == key)).scalar_one_or_none()
    if row is None:
        db.add(CacheMeta(key=key, value=value))
    else:
        row.value = value


# ---------------------------------------------------------------------------
# 크롤 오케스트레이션
# ---------------------------------------------------------------------------


async def crawl_dart_corp_index(
    *,
    session_factory: sessionmaker[Session] | None = None,
    max_industries: int | None = None,
    force: bool = False,
    client: httpx.AsyncClient | None = None,
) -> dict[str, Any]:
    """전수 크롤(실측 약 23분). 중분류 단위로 체크포인트를 남겨 재개할 수 있다.

    `force=False`면 `cache_meta`의 마지막 완료 업종 이후부터 이어서 진행한다
    (A1의 체크포인트 재개 설계와 동일). `max_industries`는 파일럿/테스트용이다.
    """
    factory = session_factory or get_session_factory()
    owns_client = client is None
    http = client or httpx.AsyncClient(timeout=httpx.Timeout(120.0), follow_redirects=True)

    processed = 0
    total_inserted = total_updated = 0
    failed_industries: list[dict[str, str]] = []
    try:
        with factory() as db:
            if force:
                _set_meta(db, _META_KEY_LAST_INDUSTRY, "")
            # 크롤이 도는 동안에는 "완료 시각"을 비워 진행 중임을 드러낸다
            # (A1에서 두 번째 이후 크롤이 항상 완료로 보고되던 버그와 같은 처리).
            _set_meta(db, _META_KEY_UPDATED_AT, "")
            db.commit()
            last_done = _get_meta(db, _META_KEY_LAST_INDUSTRY) or ""

        tree = await fetch_industry_tree(http)
        induty_lookup = build_induty_code_lookup(tree)
        codes = mid_level_codes(tree)
        pending = [c for c in codes if c > last_done] if last_done else codes
        if max_industries is not None:
            pending = pending[:max_industries]

        logger.info(
            "DART 기업개황 크롤 시작: 전체 중분류 %d개, 이번 대상 %d개(체크포인트 %r)",
            len(codes),
            len(pending),
            last_done or None,
        )

        for code in pending:
            excel_rows = await fetch_industry_excel(http, code)
            await asyncio.sleep(_REQUEST_INTERVAL_SEC)
            code_pairs = await fetch_industry_corp_codes(http, code)
            try:
                merged = merge_by_position(
                    excel_rows, code_pairs, business_code=code, induty_lookup=induty_lookup
                )
            except DartIndexCrawlError as exc:
                # 정합성 검증 실패는 **결정적**이다 — 같은 입력으로 다시 시도해도
                # 반드시 같은 결과다. 예외를 위로 던지면 바깥 감독 루프가 동일한
                # 실패를 재시도로 소진한 뒤 크롤 전체를 종료시킨다(2026-07-20 실측:
                # 업종 65와 84에서 각각 20회 재시도 후 사망, 그 뒤 업종은 아예
                # 처리되지 못했다). 해당 업종만 건너뛰고 계속 진행하되,
                # **결과에 실패 목록을 담아 조용한 누락이 되지 않게** 한다.
                logger.error("업종 %s 적재 건너뜀 — %s", code, exc)
                failed_industries.append({"industry": code, "reason": str(exc)})
                processed += 1
                _set_checkpoint(factory, code)
                await asyncio.sleep(_REQUEST_INTERVAL_SEC)
                continue
            with factory() as db:
                inserted, updated = upsert_dart_corp_index(db, merged)
                _set_meta(db, _META_KEY_LAST_INDUSTRY, code)
                db.commit()
            total_inserted += inserted
            total_updated += updated
            processed += 1
            logger.info(
                "업종 %s 완료: %d행 (신규 %d / 갱신 %d)", code, len(merged), inserted, updated
            )
            await asyncio.sleep(_REQUEST_INTERVAL_SEC)

        completed = max_industries is None and processed == len(pending)
        with factory() as db:
            if completed:
                _set_meta(db, _META_KEY_UPDATED_AT, datetime.now().isoformat(timespec="seconds"))
                _set_meta(db, _META_KEY_LAST_INDUSTRY, "")
            row_count = db.execute(select(func.count(DartCorpIndex.corp_code))).scalar_one()
            db.commit()

        if failed_industries:
            logger.warning(
                "검증 실패로 미적재된 업종 %d개: %s",
                len(failed_industries),
                [f["industry"] for f in failed_industries],
            )
        return {
            "industries_processed": processed,
            "inserted": total_inserted,
            "updated": total_updated,
            "row_count": row_count,
            "completed": completed,
            "failed_industries": failed_industries,
        }
    finally:
        if owns_client:
            await http.aclose()


# ---------------------------------------------------------------------------
# A2 — 지역/업종 로컬 필터 (M8 3단계)
# ---------------------------------------------------------------------------


def _expand_industry_prefixes(cond_industry: list[str] | None) -> list[str]:
    """조건의 업종 코드를 `induty_code` prefix 목록으로 정규화한다.

    `GET /api/meta/industries`는 대분류를 **알파벳**(A~U)으로 주는데
    `dart_corp_index.induty_code`는 KSIC 숫자 코드(2~5자리)라, 대분류를 고른
    조건을 그대로 prefix 매칭하면 **조용히 0건**이 된다 — §4-10-C에서 폐기한
    "조용한 누락"과 같은 종류의 실패다. 대분류 코드는 그 대분류에 속한 중분류
    2자리 코드 전체로 펼쳐서 넘긴다.
    """
    prefixes: list[str] = []
    for raw in cond_industry or []:
        code = (raw or "").strip()
        if not code:
            continue
        if code.isdigit():
            prefixes.append(code)
            continue
        children = next(
            (entry["children"] for entry in INDUSTRIES if entry["code"] == code), []
        )
        prefixes.extend(child["code"] for child in children)
    return list(dict.fromkeys(prefixes))


def filter_local_candidates(
    db: Session,
    *,
    cond_region: dict[str, Any] | None,
    cond_industry: list[str] | None,
) -> list[DartCorpIndex]:
    """A2: `dart_corp_index`에서 지역/업종/비상장을 DB 쿼리만으로 확정한다(외부 호출 0).

    `fsc_corp_index` 기반 구현과 결정적으로 다른 점:

    - **업종이 정밀하다.** FSC는 업종을 코드 없는 자유 텍스트(`sic_name`)로만
      줘서 KSIC 라벨과 느슨한 문자열 매칭을 할 수밖에 없었고, 그 결과 어떤
      중분류를 골라도 사실상 대분류 전체가 통과했다(2026-07-18 회귀).
      여기서는 DART가 부여한 `induty_code`에 대해 기존
      `filters.industry_matches()`(prefix 매칭)를 그대로 쓴다 —
      `_industry_labels_for_codes()`/`_sic_name_matches()`가 필요 없다.
    - **상장사를 여기서 뺀다.** `corp_cache`를 뒤져 상장 여부를 알아내던
      A4의 `_build_listed_corp_codes()` 없이, 인덱스의 `stock_code`만 보면 된다.
      상장사는 감사보고서를 별도 공시하지 않아 Phase 2에서 전부 FAILED가 된다.

    회사별 `induty_code` 부여 깊이가 2~5자리로 갈리므로(실측 2자리 5.18% /
    3자리 20.35% / 4자리 15.72% / 5자리 58.75%) 화면에서는 소분류(3자리)까지만
    고르게 한다 — 더 깊이 고르게 하면 얕게 분류된 회사가 조용히 누락된다.
    """
    # 비상장 판정은 **법인구분(corp_cls)** 으로 한다 — `stock_code` 유무로 하면 안 된다.
    # 실측(2026-07-20): `기타법인`인데 `stock_code`가 남아 있는 회사가 1,219개다
    # ((주)프리젠, 영풍산업 등 **상장폐지** 기업). 이들은 지금은 비상장 외감법인이라
    # 감사보고서를 별도 공시하는 정당한 타깃인데, stock_code로 거르면 통째로 누락된다.
    stmt = select(DartCorpIndex).where(
        or_(DartCorpIndex.corp_cls.is_(None), DartCorpIndex.corp_cls.notin_(LISTED_CORP_CLS))
    )

    cond_region = cond_region or {}
    cond_sidos = cond_sido_list(cond_region)
    if cond_sidos:
        stmt = stmt.where(DartCorpIndex.sido.in_(cond_sidos))

    prefixes = _expand_industry_prefixes(cond_industry)
    if prefixes:
        # SQL 단계에서 먼저 좁힌다(인덱스 전체를 메모리로 올리지 않기 위해).
        stmt = stmt.where(
            or_(*[DartCorpIndex.induty_code.like(f"{prefix}%") for prefix in prefixes])
        )

    rows = db.execute(stmt).scalars().all()
    return [
        row
        for row in rows
        if region_matches(row.sido, row.sigungu, cond_region)
        and industry_matches(row.induty_code, prefixes or None)
    ]


def is_dart_index_stale(
    session_factory: sessionmaker[Session] | None = None,
    ttl_days: int | None = None,
    settings: Settings | None = None,
) -> bool:
    """`dart_corp_index`가 비어있거나 TTL(기본 180일)이 지났으면 True.

    `fsc_index.is_fsc_index_stale()`/`corp_cache.is_cache_stale()`와 동일한 패턴.
    """
    settings = settings or get_settings()
    ttl_days = ttl_days if ttl_days is not None else settings.dart_index_ttl_days
    factory = session_factory or get_session_factory()

    with factory() as db:
        updated_at_raw = _get_meta(db, _META_KEY_UPDATED_AT)
        has_rows = db.execute(select(DartCorpIndex.corp_code).limit(1)).first() is not None

    if not has_rows or not updated_at_raw:
        return True
    try:
        updated_at = datetime.fromisoformat(updated_at_raw)
    except ValueError:
        return True
    return datetime.now() - updated_at > timedelta(days=ttl_days)


def get_dart_index_status(
    session_factory: sessionmaker[Session] | None = None,
) -> dict[str, Any]:
    """행 수 / 마지막 완료 시각 / 진행 중 여부 / 동명 그룹 교정이 밀렸는지."""
    factory = session_factory or get_session_factory()
    with factory() as db:
        row_count = db.execute(select(func.count(DartCorpIndex.corp_code))).scalar_one()
        updated_at = _get_meta(db, _META_KEY_UPDATED_AT) or None
        checkpoint = _get_meta(db, _META_KEY_LAST_INDUSTRY) or None
        reconciled_at = _get_meta(db, _META_KEY_RECONCILED_AT) or None
    return {
        "row_count": row_count,
        "last_completed_at": updated_at,
        "crawl_in_progress": updated_at is None and checkpoint is not None,
        "checkpoint_industry": checkpoint,
        "last_reconciled_at": reconciled_at,
        # 크롤은 끝났는데 그 뒤로 교정을 한 적이 없으면 동명 회사의 주소/업종이
        # 교차된 채 남아 있을 수 있다(M8 6단계 실측: 위험군 불일치 42.5%).
        # 교정은 크롤 완료 후 자동으로 이어지지만, 쿼터 소진 등으로 중단되면
        # 이 플래그가 남아 화면에서 재실행을 유도한다.
        "reconcile_pending": bool(row_count)
        and (reconciled_at is None or (updated_at is not None and reconciled_at < updated_at)),
    }
